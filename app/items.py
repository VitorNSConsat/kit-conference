import io
from database import db, now_brt


# ── Tipos de item ──────────────────────────────────────────────────────────────

def listar_tipos(apenas_ativos: bool = False) -> list:
    with db() as conn:
        if apenas_ativos:
            rows = conn.execute(
                "SELECT * FROM item_tipo WHERE ativo = 1 ORDER BY nome"
            ).fetchall()
        else:
            rows = conn.execute(
                "SELECT * FROM item_tipo ORDER BY nome"
            ).fetchall()
    return [dict(r) for r in rows]


def listar_tipos_para_kit(template_id: int) -> list:
    """Retorna os tipos presentes no template disponíveis para classificação manual
    de um código desconhecido — excluindo tipos com código fixo (têm fluxo próprio)
    e mostrando apenas tipos marcados como "Item de Patrimônio" (controle_externo=1).
    Só tipos individualmente rastreados por patrimônio devem ser opções nessa tela;
    tipos não marcados ficam ocultos dela."""
    with db() as conn:
        rows = conn.execute(
            "SELECT it.id, it.nome FROM item_tipo it "
            "JOIN kit_template_items ki ON ki.item_tipo_id = it.id "
            "WHERE ki.kit_template_id = ? AND it.ativo = 1 "
            "AND (it.codigo_fixo IS NULL OR it.codigo_fixo = '') "
            "AND COALESCE(it.controle_externo, 0) = 1 "
            "ORDER BY it.nome",
            (template_id,)
        ).fetchall()
    return [dict(r) for r in rows]


def buscar_tipo_por_codigo_fixo(codigo: str) -> dict | None:
    with db() as conn:
        row = conn.execute(
            "SELECT id, nome, reutilizavel FROM item_tipo WHERE codigo_fixo = ? AND ativo = 1",
            (codigo,)
        ).fetchone()
    return dict(row) if row else None


def definir_codigo_fixo(tipo_id: int, codigo: str | None):
    valor = codigo.strip() if codigo and codigo.strip() else None
    with db() as conn:
        conn.execute(
            "UPDATE item_tipo SET codigo_fixo = ? WHERE id = ?", (valor, tipo_id)
        )


def criar_tipo(nome: str, unidade: str = "un") -> int:
    unidade = unidade if unidade in ("un", "m") else "un"
    with db() as conn:
        cur = conn.execute(
            "INSERT INTO item_tipo (nome, unidade, criado_em) VALUES (?, ?, ?)",
            (nome.strip(), unidade, now_brt())
        )
        return cur.lastrowid


def alternar_reutilizavel_tipo(tipo_id: int):
    with db() as conn:
        conn.execute(
            "UPDATE item_tipo SET reutilizavel = 1 - COALESCE(reutilizavel, 0) WHERE id = ?",
            (tipo_id,)
        )


def alternar_controle_externo(tipo_id: int):
    with db() as conn:
        conn.execute(
            "UPDATE item_tipo SET controle_externo = 1 - COALESCE(controle_externo, 0) WHERE id = ?",
            (tipo_id,)
        )


def alternar_requer_serial(tipo_id: int):
    with db() as conn:
        conn.execute(
            "UPDATE item_tipo SET requer_serial = 1 - COALESCE(requer_serial, 0) WHERE id = ?",
            (tipo_id,)
        )


def alternar_unidade_tipo(tipo_id: int):
    with db() as conn:
        conn.execute(
            "UPDATE item_tipo SET unidade = CASE WHEN unidade = 'm' THEN 'un' ELSE 'm' END WHERE id = ?",
            (tipo_id,)
        )


def buscar_dependencias_tipo(tipo_id: int) -> dict:
    with db() as conn:
        tipo = conn.execute("SELECT nome FROM item_tipo WHERE id = ?", (tipo_id,)).fetchone()
        patrimonios = conn.execute(
            "SELECT COUNT(*) AS n FROM item_master WHERE item_tipo_id = ?", (tipo_id,)
        ).fetchone()["n"]
        templates_rows = conn.execute(
            "SELECT DISTINCT kt.nome FROM kit_template_items ki "
            "JOIN kit_template kt ON kt.id = ki.kit_template_id "
            "WHERE ki.item_tipo_id = ?", (tipo_id,)
        ).fetchall()
        estoque_n = conn.execute(
            "SELECT COUNT(*) AS n FROM estoque WHERE item_tipo_id = ?", (tipo_id,)
        ).fetchone()["n"]
    return {
        "tipo_id": tipo_id,
        "tipo_nome": tipo["nome"] if tipo else "?",
        "patrimonios": patrimonios,
        "templates": [r["nome"] for r in templates_rows],
        "estoque": estoque_n,
    }


def deletar_tipo_cascade(tipo_id: int):
    with db() as conn:
        conn.execute("DELETE FROM scan_session_items WHERE item_tipo_id = ?", (tipo_id,))
        conn.execute(
            "DELETE FROM estoque_movimentos WHERE estoque_id IN "
            "(SELECT id FROM estoque WHERE item_tipo_id = ?)", (tipo_id,)
        )
        conn.execute("DELETE FROM item_master WHERE item_tipo_id = ?", (tipo_id,))
        conn.execute("DELETE FROM kit_template_items WHERE item_tipo_id = ?", (tipo_id,))
        conn.execute("DELETE FROM estoque WHERE item_tipo_id = ?", (tipo_id,))
        conn.execute("DELETE FROM item_tipo WHERE id = ?", (tipo_id,))


def renomear_tipo(tipo_id: int, novo_nome: str):
    with db() as conn:
        conn.execute("UPDATE item_tipo SET nome = ? WHERE id = ?", (novo_nome.strip(), tipo_id))


def deletar_tipo(tipo_id: int):
    with db() as conn:
        conn.execute("DELETE FROM item_tipo WHERE id = ?", (tipo_id,))


def toggle_tipo(tipo_id: int):
    with db() as conn:
        conn.execute(
            "UPDATE item_tipo SET ativo = 1 - ativo WHERE id = ?", (tipo_id,)
        )


def importar_tipos_xlsx(conteudo: bytes) -> dict:
    """Lê um arquivo .xlsx e importa a primeira coluna (a partir da linha 2) como tipos de item.
    Retorna {'criados': N, 'ignorados': M} onde ignorados = duplicatas já existentes."""
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(conteudo), read_only=True, data_only=True)
    ws = wb.active
    criados = 0
    ignorados = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        valor = row[0] if row else None
        if not valor:
            continue
        nome = str(valor).strip()
        if not nome:
            continue
        try:
            criar_tipo(nome)
            criados += 1
        except Exception:
            ignorados += 1
    wb.close()
    return {"criados": criados, "ignorados": ignorados}


# ── Patrimônios (item_master) ──────────────────────────────────────────────────

# Por que um patrimônio aparece sem veículo. Não é erro por si só — a
# maioria dos casos é situação normal (item novo, kit ainda em montagem).
SITUACOES = {
    "ok": "",
    "nunca_bipado": "Cadastrado mas nunca bipado em nenhum kit",
    "em_bipagem": "Está numa bipagem em andamento — ganha veículo ao finalizar",
    "kit_sem_veiculo": "Bipado num kit que foi finalizado sem veículo definido",
    "kit_removido": "O kit onde foi bipado não existe mais (excluído)",
}


def listar_itens(veiculo_id: int | None = None, situacao: str = "") -> list:
    """Patrimônios cadastrados, com o veículo, serial e operador do kit mais
    recente em que cada um foi bipado (o kit 'ativo' mais novo, então pra
    item reutilizável reflete a atribuição atual, não o histórico inteiro).

    Cada item traz também `situacao`, que explica por que está sem veículo
    quando for o caso — a lista sozinha não distinguia "nunca foi bipado"
    de "o kit foi finalizado sem veículo", que pedem ações bem diferentes.

    veiculo_id filtra os itens atribuídos a esse veículo; situacao filtra
    por um dos códigos de SITUACOES."""
    with db() as conn:
        rows = conn.execute("""
            WITH ult_kit AS (
                -- O JOIN com item_master não é decorativo: a maior parte
                -- das linhas de bipagem é de conjunto (COMP:) e estoque
                -- (ESTOQUE:), que nunca casam com um patrimônio. Filtrar
                -- antes de ordenar corta o volume que a janela precisa
                -- percorrer, e o resultado é idêntico.
                SELECT si.codigo_barra, si.serial_number,
                       kr.kit_id, kr.veiculo_id, kr.veiculo, kr.garagem,
                       kr.finalizado_em, kr.operador_id,
                       ROW_NUMBER() OVER (
                           PARTITION BY si.codigo_barra
                           ORDER BY kr.finalizado_em DESC
                       ) AS rn
                FROM scan_session_items si
                JOIN item_master im ON im.codigo_barra = si.codigo_barra
                JOIN kit_record kr ON kr.sessao_id = si.sessao_id
                -- Linha movida/retirada nao diz mais onde o item esta: ela
                -- conta onde ele ESTEVE.
                WHERE kr.status = 'ativo'
                  AND (si.status IS NULL OR si.status NOT IN ('movido', 'retirado'))
            )
            SELECT i.*, t.nome AS descricao, u.nome AS criado_por_nome,
                   k.veiculo_id AS veiculo_id_atual,
                   COALESCE(v.numero, k.veiculo) AS veiculo_atual,
                   k.kit_id  AS kit_id_atual,
                   k.garagem AS garagem_atual,
                   k.serial_number AS serial_atual,
                   k.finalizado_em AS bipado_em_atual,
                   op.nome AS operador_atual,
                   (SELECT COUNT(*) FROM scan_session_items s2
                     WHERE s2.codigo_barra = i.codigo_barra) AS total_bipagens,
                   (SELECT ss2.status FROM scan_session_items s2
                     JOIN scan_session ss2 ON ss2.id = s2.sessao_id
                     WHERE s2.codigo_barra = i.codigo_barra
                     ORDER BY s2.id DESC LIMIT 1) AS status_ultima_sessao
            FROM item_master i
            JOIN item_tipo t ON t.id = i.item_tipo_id
            LEFT JOIN users u ON u.id = i.criado_por
            LEFT JOIN ult_kit k ON k.codigo_barra = i.codigo_barra AND k.rn = 1
            LEFT JOIN veiculos v ON v.id = k.veiculo_id
            LEFT JOIN users op ON op.id = k.operador_id
            ORDER BY t.nome, i.codigo_barra
        """).fetchall()

    itens = []
    for r in rows:
        d = dict(r)
        if d["veiculo_atual"]:
            d["situacao"] = "ok"
        elif not d["total_bipagens"]:
            d["situacao"] = "nunca_bipado"
        elif d["status_ultima_sessao"] == "em_andamento":
            d["situacao"] = "em_bipagem"
        elif d["kit_id_atual"]:
            d["situacao"] = "kit_sem_veiculo"
        else:
            d["situacao"] = "kit_removido"
        d["situacao_texto"] = SITUACOES[d["situacao"]]
        itens.append(d)

    if veiculo_id:
        itens = [i for i in itens if i["veiculo_id_atual"] == veiculo_id]
    if situacao in SITUACOES:
        itens = [i for i in itens if i["situacao"] == situacao]
    return itens


def historico_patrimonio(codigo_barra: str) -> list[dict]:
    """Toda vez que este código foi bipado: quando, em que sessão/kit, por
    qual operador e pra qual veículo. Responde 'onde foi bipado?' sem
    depender de o kit ainda ter veículo."""
    with db() as conn:
        rows = conn.execute("""
            SELECT si.id AS si_id, si.bipado_em, si.serial_number, si.observacao,
                   si.sessao_id, ss.status AS sessao_status,
                   kt.nome AS kit_nome, kt.cliente,
                   kr.kit_id, kr.status AS kit_status,
                   COALESCE(v.numero, kr.veiculo, ss.veiculo, '') AS veiculo,
                   COALESCE(kr.garagem, ss.garagem, '') AS garagem,
                   COALESCE(opi.nome, ops.nome) AS operador_nome
            FROM scan_session_items si
            JOIN scan_session ss ON ss.id = si.sessao_id
            JOIN kit_template kt ON kt.id = ss.kit_template_id
            LEFT JOIN kit_record kr ON kr.sessao_id = ss.id
            LEFT JOIN veiculos v ON v.id = kr.veiculo_id
            LEFT JOIN users opi ON opi.id = si.operador_id
            LEFT JOIN users ops ON ops.id = ss.operador_id
            WHERE si.codigo_barra = ?
            ORDER BY si.bipado_em DESC, si.id DESC
        """, (codigo_barra,)).fetchall()
    return [dict(r) for r in rows]


def bipados_na_mesma_sessao(sessao_id: int, codigo_barra: str) -> list[dict]:
    """Os itens que formaram O MESMO KIT deste patrimônio na última sessão.

    O corte é a SESSÃO de bipagem (scan_session_items.sessao_id), não a
    proximidade de horário no leitor: uma sessão é exatamente um kit sendo
    montado, então dois kits montados em paralelo — ou um item bipado logo
    depois, já em outra sessão — nunca se misturam aqui. É por isso que
    esta seção responde "o que está na mesma caixa", enquanto
    historico_patrimonio() responde "por onde este item já passou"."""
    with db() as conn:
        rows = conn.execute("""
            SELECT si.codigo_barra, si.bipado_em, si.serial_number, si.quantidade,
                   it.nome AS descricao, u.nome AS operador_nome
            FROM scan_session_items si
            JOIN item_tipo it ON it.id = si.item_tipo_id
            LEFT JOIN users u ON u.id = si.operador_id
            WHERE si.sessao_id = ?
              AND (si.status IS NULL OR si.status NOT IN ('movido', 'retirado'))
            ORDER BY si.bipado_em, si.id
        """, (sessao_id,)).fetchall()
    itens = []
    for r in rows:
        d = dict(r)
        d["e_o_item"] = d["codigo_barra"] == codigo_barra
        d["situacao"] = "Item selecionado" if d["e_o_item"] else "Mesmo kit"
        itens.append(d)
    return itens


def kit_da_sessao(sessao_id: int) -> dict | None:
    """Qual kit foi formado nesta sessão — nome, veículo, garagem e estado.
    Usado pra dizer na tela de qual kit são os itens de "Bipado junto"."""
    with db() as conn:
        row = conn.execute("""
            SELECT ss.id AS sessao_id, ss.status AS sessao_status,
                   kt.nome AS kit_nome, kt.cliente,
                   kr.kit_id, kr.status_producao,
                   COALESCE(v.numero, kr.veiculo, ss.veiculo, '') AS veiculo,
                   COALESCE(kr.garagem, ss.garagem, '') AS garagem
            FROM scan_session ss
            JOIN kit_template kt ON kt.id = ss.kit_template_id
            LEFT JOIN kit_record kr ON kr.sessao_id = ss.id
            LEFT JOIN veiculos v ON v.id = kr.veiculo_id
            WHERE ss.id = ?
        """, (sessao_id,)).fetchone()
    return dict(row) if row else None


# Linhas de bipagem que NÃO contam mais como conteúdo do kit: o item saiu
# dali (foi pra outro veículo ou foi retirado). A linha continua no banco de
# propósito — é o que prova onde o item esteve e por quê.
STATUS_FORA_DO_KIT = ("movido", "retirado")


def onde_esta(codigo_barra: str) -> dict | None:
    """Onde este patrimônio está AGORA: o kit/veículo da bipagem mais recente
    que ainda vale (linha não movida nem retirada).

    É a resposta que faltava antes de deixar alguém mexer no patrimônio: sem
    ela, o operador só descobria que o código estava em uso quando levava um
    erro seco — e não dizia onde."""
    codigo_barra = (codigo_barra or "").strip()
    if not codigo_barra:
        return None
    with db() as conn:
        row = conn.execute("""
            SELECT si.id AS si_id, si.sessao_id, si.bipado_em, si.serial_number,
                   si.item_tipo_id, it.nome AS tipo_nome,
                   ss.status AS sessao_status,
                   kt.nome AS kit_nome, kt.cliente,
                   kr.kit_id, kr.status_producao,
                   COALESCE(v.numero, kr.veiculo, ss.veiculo, '') AS veiculo,
                   COALESCE(kr.garagem, ss.garagem, '') AS garagem,
                   COALESCE(opi.nome, ops.nome) AS operador_nome
            FROM scan_session_items si
            JOIN item_tipo it ON it.id = si.item_tipo_id
            JOIN scan_session ss ON ss.id = si.sessao_id
            JOIN kit_template kt ON kt.id = ss.kit_template_id
            LEFT JOIN kit_record kr ON kr.sessao_id = ss.id
            LEFT JOIN veiculos v ON v.id = kr.veiculo_id
            LEFT JOIN users opi ON opi.id = si.operador_id
            LEFT JOIN users ops ON ops.id = ss.operador_id
            WHERE si.codigo_barra = ?
              AND (si.status IS NULL OR si.status NOT IN ('movido', 'retirado'))
            ORDER BY si.bipado_em DESC, si.id DESC LIMIT 1
        """, (codigo_barra,)).fetchone()
    return dict(row) if row else None


def kit_do_veiculo(numero: str) -> dict | None:
    """O kit mais recente de um veículo, procurando pelo NÚMERO — que é como
    o operador pensa ("o 1219"), não pelo kit_id."""
    numero = (numero or "").strip()
    if not numero:
        return None
    with db() as conn:
        row = conn.execute("""
            SELECT kr.kit_id, kr.sessao_id, kr.kit_template_id, kr.status_producao,
                   kr.finalizado_em, kt.nome AS kit_nome, kt.cliente,
                   COALESCE(v.numero, kr.veiculo) AS veiculo,
                   COALESCE(kr.garagem, '') AS garagem
            FROM kit_record kr
            JOIN kit_template kt ON kt.id = kr.kit_template_id
            LEFT JOIN veiculos v ON v.id = kr.veiculo_id
            WHERE kr.status = 'ativo'
              AND (UPPER(TRIM(COALESCE(v.numero, ''))) = UPPER(?)
                   OR UPPER(TRIM(COALESCE(kr.veiculo, ''))) = UPPER(?))
            ORDER BY kr.finalizado_em DESC LIMIT 1
        """, (numero, numero)).fetchone()
    return dict(row) if row else None


def _validar_motivo(motivo: str) -> str:
    """Motivo é obrigatório em toda mexida manual de patrimônio: seis meses
    depois, "por que este item mudou de veículo?" só tem resposta se alguém
    escreveu na hora."""
    motivo = (motivo or "").strip()
    if len(motivo) < 5:
        raise ValueError("Informe o motivo da alteração (pelo menos 5 letras) — "
                         "ele fica gravado no histórico do item.")
    return motivo


def previa_mover(codigo_barra: str, numero_destino: str) -> dict:
    """O que vai acontecer se este patrimônio for pro veículo informado.

    Não grava nada: existe pra a pessoa confirmar vendo de onde sai, pra onde
    vai e o que cada lado perde ou ganha."""
    codigo_barra = (codigo_barra or "").strip()
    origem = onde_esta(codigo_barra)
    destino = kit_do_veiculo(numero_destino)
    avisos = []
    bloqueio = None
    if not destino:
        bloqueio = (f"Nenhum kit encontrado para o veículo '{numero_destino}'. "
                    "Confira o número — o veículo precisa ter um kit montado.")
    elif origem and origem.get("sessao_id") == destino["sessao_id"]:
        bloqueio = "Este patrimônio já está neste kit."
    elif not origem:
        bloqueio = ("Este patrimônio não está em nenhum kit. Use "
                    "\"Atribuir a um kit\" no kit de destino.")
    if destino and not bloqueio and origem:
        with db() as conn:
            pertence = conn.execute(
                "SELECT 1 FROM kit_template_items WHERE kit_template_id = ? "
                "AND item_tipo_id = ?",
                (destino["kit_template_id"], origem["item_tipo_id"])
            ).fetchone()
        if not pertence:
            avisos.append(
                f"O tipo '{origem['tipo_nome']}' não faz parte do modelo "
                f"'{destino['kit_nome']}'. O item vai entrar assim mesmo (foi o que "
                "aconteceu de verdade), mas vai aparecer como sobrando na verificação.")
        if origem.get("status_producao"):
            avisos.append(
                f"O kit de origem ({origem['veiculo'] or '—'}) fica FALTANDO este item "
                "e passa a aparecer na lista de pendências até receber outro.")
    return {"origem": origem, "destino": destino, "avisos": avisos, "bloqueio": bloqueio}


def mover_patrimonio(codigo_barra: str, numero_destino: str, motivo: str,
                     user_id: int | None = None) -> dict:
    """Passa um patrimônio do kit onde ele está para o kit de outro veículo.

    É o caso real que o sistema não cobria: o item saiu do kit do veículo A e
    foi instalado no veículo B na hora da instalação. Antes só dava pra
    resolver apagando o item do A, o que sumia com o rastro.

    A linha antiga NÃO é apagada — vira 'movido', com o motivo gravado, e
    continua no histórico do item provando onde ele esteve. Uma linha nova
    entra no kit de destino com a data de agora."""
    motivo = _validar_motivo(motivo)
    previa = previa_mover(codigo_barra, numero_destino)
    if previa["bloqueio"]:
        raise ValueError(previa["bloqueio"])
    origem, destino = previa["origem"], previa["destino"]
    carimbo = now_brt()
    nota = (f"Movido para o veículo {destino['veiculo']} em {carimbo} — {motivo}")
    with db() as conn:
        conn.execute(
            "UPDATE scan_session_items SET status = 'movido', observacao = ? WHERE id = ?",
            (nota, origem["si_id"]))
        conn.execute(
            "INSERT INTO scan_session_items "
            "(sessao_id, codigo_barra, item_tipo_id, status, bipado_em, operador_id, "
            " serial_number, observacao, quantidade, estoque_debitado) "
            "VALUES (?, ?, ?, 'completo', ?, ?, ?, ?, 1, 0)",
            (destino["sessao_id"], codigo_barra, origem["item_tipo_id"], carimbo,
             user_id, origem.get("serial_number"),
             f"Veio do veículo {origem['veiculo'] or '—'} em {carimbo} — {motivo}"))
    return {"origem": origem, "destino": destino, "motivo": motivo}


def atribuir_patrimonio(codigo_barra: str, kit_id: str, item_tipo_id: int,
                        motivo: str, serial: str = "",
                        user_id: int | None = None) -> dict:
    """Cadastra (se preciso) um patrimônio e coloca num kit já finalizado.

    É a outra metade do caso: o veículo que ficou sem o item precisa receber
    outro, e isso acontece fora da bipagem — o kit já está fechado, às vezes
    já no cliente. Não mexe em estágio de produção nem em estoque; só diz
    que este item passou a fazer parte deste kit."""
    motivo = _validar_motivo(motivo)
    codigo_barra = (codigo_barra or "").strip()
    if not codigo_barra:
        raise ValueError("Informe o código do patrimônio.")
    ocupado = onde_esta(codigo_barra)
    if ocupado:
        raise ValueError(
            f"O patrimônio {codigo_barra} já está no kit do veículo "
            f"{ocupado['veiculo'] or '—'} ({ocupado['kit_nome']}). Use "
            "\"Mover para outro veículo\" na página desse patrimônio.")
    with db() as conn:
        kit = conn.execute(
            "SELECT kr.kit_id, kr.sessao_id, kr.kit_template_id, "
            "COALESCE(v.numero, kr.veiculo, '') AS veiculo "
            "FROM kit_record kr LEFT JOIN veiculos v ON v.id = kr.veiculo_id "
            "WHERE kr.kit_id = ?", (kit_id,)).fetchone()
        if not kit:
            raise ValueError("Kit não encontrado.")
        tipo = conn.execute("SELECT nome FROM item_tipo WHERE id = ?",
                            (item_tipo_id,)).fetchone()
        if not tipo:
            raise ValueError("Tipo de item não encontrado.")
        carimbo = now_brt()
        # O cadastro do patrimônio nasce junto quando o código é novo — é o
        # mesmo que a bipagem faz ao encontrar um código desconhecido.
        ja_cadastrado = conn.execute(
            "SELECT id FROM item_master WHERE codigo_barra = ?", (codigo_barra,)).fetchone()
        if not ja_cadastrado:
            conn.execute(
                "INSERT INTO item_master (codigo_barra, item_tipo_id, criado_por, criado_em) "
                "VALUES (?, ?, ?, ?)", (codigo_barra, item_tipo_id, user_id, carimbo))
        conn.execute(
            "INSERT INTO scan_session_items "
            "(sessao_id, codigo_barra, item_tipo_id, status, bipado_em, operador_id, "
            " serial_number, observacao, quantidade, estoque_debitado) "
            "VALUES (?, ?, ?, 'completo', ?, ?, ?, ?, 1, 0)",
            (kit["sessao_id"], codigo_barra, item_tipo_id, carimbo, user_id,
             (serial or "").strip() or None,
             f"Atribuído manualmente em {carimbo} — {motivo}"))
    return {"kit_id": kit["kit_id"], "veiculo": kit["veiculo"],
            "tipo_nome": tipo["nome"], "novo_cadastro": not ja_cadastrado}


def retirar_do_kit(codigo_barra: str, motivo: str, user_id: int | None = None) -> dict:
    """Tira o patrimônio do kit sem colocar em outro — peça que voltou pro
    estoque, quebrou ou sumiu. A linha vira 'retirado' com o motivo; o kit
    passa a acusar o item faltando, que é a verdade."""
    motivo = _validar_motivo(motivo)
    origem = onde_esta(codigo_barra)
    if not origem:
        raise ValueError("Este patrimônio não está em nenhum kit.")
    with db() as conn:
        conn.execute(
            "UPDATE scan_session_items SET status = 'retirado', observacao = ? WHERE id = ?",
            (f"Retirado do kit em {now_brt()} — {motivo}", origem["si_id"]))
    return {"origem": origem, "motivo": motivo}


def corrigir_patrimonio(codigo_atual: str, novo_codigo: str = "",
                        novo_serial: str | None = None) -> dict:
    """Correção CADASTRAL de um patrimônio, válida em qualquer estágio —
    inclusive com o veículo já no cliente e finalizado.

    Só mexe em identificação (código e número de série). Não toca em
    status_producao, kit_record, scan_session nem estoque: corrigir um
    número digitado errado não é motivo pra reabrir produção, e o veículo
    continua exatamente no estágio em que estava.

    O código novo é propagado pras bipagens que usam o antigo, na MESMA
    transação — scan_session_items referencia o patrimônio por texto, então
    renomear só o cadastro deixaria o histórico órfão. O histórico não é
    apagado nem recriado: as mesmas linhas passam a apontar pro código
    novo, preservando data, operador, sessão e kit de cada passagem."""
    codigo_atual = (codigo_atual or "").strip()
    novo_codigo = (novo_codigo or "").strip()
    if not codigo_atual:
        raise ValueError("Patrimônio não informado.")

    with db() as conn:
        item = conn.execute(
            "SELECT * FROM item_master WHERE codigo_barra = ?", (codigo_atual,)
        ).fetchone()
        if not item:
            raise ValueError(f"Patrimônio {codigo_atual} não encontrado.")

        renomeou = False
        if novo_codigo and novo_codigo != codigo_atual:
            existe = conn.execute(
                "SELECT 1 FROM item_master WHERE codigo_barra = ?", (novo_codigo,)
            ).fetchone()
            if existe:
                # Diz ONDE está o dono do código: só "já pertence a outro"
                # deixava o operador sem saber o que fazer — foi o que
                # obrigou, no campo, a apagar o item pra conseguir seguir.
                dono = onde_esta(novo_codigo)
                if dono:
                    raise ValueError(
                        f"O código {novo_codigo} já é de outro patrimônio, que está "
                        f"no kit do veículo {dono['veiculo'] or '—'} "
                        f"({dono['kit_nome']}, {dono['cliente']}). Se o item trocou de "
                        "veículo, use \"Mover para outro veículo\" em vez de renomear.")
                raise ValueError(
                    f"O código {novo_codigo} já pertence a outro patrimônio "
                    "(sem kit no momento).")
            conn.execute("UPDATE item_master SET codigo_barra = ? WHERE id = ?",
                         (novo_codigo, item["id"]))
            conn.execute(
                "UPDATE scan_session_items SET codigo_barra = ? WHERE codigo_barra = ?",
                (novo_codigo, codigo_atual))
            renomeou = True

        codigo_final = novo_codigo if renomeou else codigo_atual
        seriais = 0
        if novo_serial is not None:
            # Corrige o serial da bipagem MAIS RECENTE deste patrimônio: é a
            # que descreve onde ele está agora. As anteriores continuam com
            # o que foi registrado na época — histórico não se reescreve.
            ultima = conn.execute(
                "SELECT id FROM scan_session_items WHERE codigo_barra = ? "
                "ORDER BY bipado_em DESC, id DESC LIMIT 1", (codigo_final,)
            ).fetchone()
            if ultima:
                conn.execute(
                    "UPDATE scan_session_items SET serial_number = ? WHERE id = ?",
                    ((novo_serial or "").strip() or None, ultima["id"]))
                seriais = 1

    return {"codigo": codigo_final, "renomeou": renomeou, "seriais_atualizados": seriais}


def buscar_item(codigo_barra: str) -> dict | None:
    with db() as conn:
        row = conn.execute(
            "SELECT i.*, t.nome AS descricao, "
            "COALESCE(t.unidade, 'un') AS unidade, "
            "COALESCE(t.reutilizavel, 0) AS reutilizavel "
            "FROM item_master i "
            "JOIN item_tipo t ON t.id = i.item_tipo_id "
            "WHERE i.codigo_barra = ? AND i.ativo = 1",
            (codigo_barra,)
        ).fetchone()
    return dict(row) if row else None


def criar_item(codigo_barra: str, item_tipo_id: int, criado_por: int) -> int:
    with db() as conn:
        cur = conn.execute(
            "INSERT INTO item_master (codigo_barra, item_tipo_id, criado_por, criado_em) "
            "VALUES (?, ?, ?, ?)",
            (codigo_barra, item_tipo_id, criado_por, now_brt())
        )
        return cur.lastrowid


def deletar_item(item_id: int):
    with db() as conn:
        conn.execute("DELETE FROM item_master WHERE id = ?", (item_id,))


def apagar_todos_itens():
    with db() as conn:
        conn.execute("DELETE FROM item_master")


def toggle_item(item_id: int):
    with db() as conn:
        conn.execute(
            "UPDATE item_master SET ativo = 1 - ativo WHERE id = ?", (item_id,)
        )


def importar_bom_xlsx(conteudo: bytes, criado_por: int) -> dict:
    """Importa tipos e patrimônios a partir de um BOM Excel.

    Detecta automaticamente a linha de cabeçalho procurando por 'Description'.
    Colunas usadas: Code → item_master.codigo_barra, Description → item_tipo.nome.
    Rows without a description are skipped; rows without a code create only the tipo.
    """
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(conteudo), data_only=True)
    ws = wb.active

    # Detecta header row e índices de colunas
    header_row = None
    col_desc = col_code = None
    for row in ws.iter_rows(values_only=True):
        cells = [str(c).strip().lower() if c else "" for c in row]
        if "description" in cells:
            header_row = True
            col_desc = next(i for i, c in enumerate(cells) if c == "description")
            # Code pode se chamar 'code', 'part number', 'código', etc.
            for label in ("code", "part number", "código", "codigo", "part no"):
                if label in cells:
                    col_code = next(i for i, c in enumerate(cells) if c == label)
                    break
            break

    if header_row is None:
        wb.close()
        return {"tipos_criados": 0, "itens_criados": 0, "ignorados": 0,
                "erro": "Cabeçalho 'Description' não encontrado na planilha."}

    tipos_criados = itens_criados = ignorados = 0

    with db() as conn:
        past_header = False
        for row in ws.iter_rows(values_only=True):
            cells = [str(c).strip().lower() if c else "" for c in row]
            # Pula até depois do header
            if not past_header:
                if "description" in cells:
                    past_header = True
                continue

            desc = str(row[col_desc]).strip() if col_desc is not None and row[col_desc] else ""
            code = (str(row[col_code]).strip() if col_code is not None and row[col_code] else "")
            # Limpa values como "None" ou "no part number"
            if desc.lower() in ("none", "") or not desc:
                continue
            if code.lower() in ("none", "no part number", "n/a", ""):
                code = ""

            # Cria ou recupera o tipo
            existing_tipo = conn.execute(
                "SELECT id FROM item_tipo WHERE nome = ?", (desc,)
            ).fetchone()
            if existing_tipo:
                tipo_id = existing_tipo["id"]
                ignorados += 1
            else:
                cur = conn.execute(
                    "INSERT INTO item_tipo (nome, criado_em) VALUES (?, ?)",
                    (desc, now_brt())
                )
                tipo_id = cur.lastrowid
                tipos_criados += 1

            # Cria patrimônio se houver código
            if code:
                exists = conn.execute(
                    "SELECT 1 FROM item_master WHERE codigo_barra = ?", (code,)
                ).fetchone()
                if not exists:
                    conn.execute(
                        "INSERT INTO item_master (codigo_barra, item_tipo_id, criado_por, criado_em) "
                        "VALUES (?, ?, ?, ?)",
                        (code, tipo_id, criado_por, now_brt())
                    )
                    itens_criados += 1

    wb.close()
    return {"tipos_criados": tipos_criados, "itens_criados": itens_criados, "ignorados": ignorados}
