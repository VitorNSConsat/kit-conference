from database import db, now_brt


def listar(cliente: str | None = None, ativo: bool = True,
           modelo: str | None = None) -> list[dict]:
    """`modelo` filtra os veículos daquele modelo — que é o NOME do kit.
    Comparação sem caixa/espaço extra, pra "mercedes euro 5 diesel" e
    "Mercedes Euro 5 Diesel" serem o mesmo modelo."""
    sql = """
        SELECT v.*,
               COUNT(kr.kit_id) AS total_kits,
               MAX(kr.finalizado_em) AS ultimo_kit_em
        FROM veiculos v
        LEFT JOIN kit_record kr ON kr.veiculo_id = v.id
        WHERE v.ativo = ?
    """
    params: list = [1 if ativo else 0]
    if cliente:
        sql += " AND v.cliente = ?"
        params.append(cliente)
    if modelo is not None:
        sql += " AND LOWER(TRIM(COALESCE(v.modelo, ''))) = ?"
        params.append(modelo.strip().lower())
    sql += " GROUP BY v.id ORDER BY v.cliente, v.numero"
    with db() as conn:
        rows = conn.execute(sql, params).fetchall()
    return [dict(r) for r in rows]


def buscar(veiculo_id: int) -> dict | None:
    with db() as conn:
        row = conn.execute(
            "SELECT * FROM veiculos WHERE id = ?", (veiculo_id,)
        ).fetchone()
    return dict(row) if row else None


def numero_em_uso(numero: str, excluir_id: int | None = None) -> dict | None:
    """O número do veículo é único no sistema INTEIRO — não só dentro do
    cliente. O mesmo 31001-50071 em dois clientes é o mesmo ônibus físico
    cadastrado duas vezes (o número vem da frota, não do cliente), e a
    duplicata deixava kits e histórico espalhados entre dois cadastros.
    Compara sem caixa/espaço e SEM filtrar por ativo: um desativado ainda
    ocupa o número. Devolve o veículo que já usa o número, ou None."""
    sql = "SELECT * FROM veiculos WHERE LOWER(TRIM(numero)) = ?"
    params: list = [(numero or "").strip().lower()]
    if excluir_id is not None:
        sql += " AND id != ?"
        params.append(excluir_id)
    with db() as conn:
        row = conn.execute(sql, params).fetchone()
    return dict(row) if row else None


def numeros_duplicados() -> list[dict]:
    """Números que já estão cadastrados mais de uma vez (dados de antes da
    regra de número único). A tela lista pra o admin limpar — o sistema não
    apaga sozinho porque cada cadastro pode ter kits no histórico."""
    with db() as conn:
        rows = conn.execute("""
            SELECT LOWER(TRIM(numero)) AS chave, MIN(numero) AS numero,
                   COUNT(*) AS repeticoes,
                   GROUP_CONCAT(cliente, ' / ') AS clientes
            FROM veiculos
            GROUP BY chave HAVING COUNT(*) > 1
            ORDER BY numero
        """).fetchall()
    return [dict(r) for r in rows]


def criar(numero: str, cliente: str, garagem: str, modelo: str = "") -> int:
    ocupado = numero_em_uso(numero)
    if ocupado:
        raise ValueError(
            f"O número {numero.strip()} já está cadastrado no cliente "
            f"{ocupado['cliente']}. O número do veículo é único no sistema — "
            "edite o cadastro existente em vez de criar outro.")
    with db() as conn:
        cur = conn.execute(
            "INSERT INTO veiculos (numero, cliente, garagem, modelo, criado_em) "
            "VALUES (?, ?, ?, ?, ?)",
            (numero.strip(), cliente.strip(), garagem.strip().upper(),
             modelo.strip(), now_brt())
        )
        return cur.lastrowid


def atualizar(veiculo_id: int, numero: str, cliente: str, garagem: str,
              modelo: str | None = None):
    """modelo=None mantém o que já está gravado — deixa outras telas
    atualizarem número/cliente/garagem sem apagar o modelo sem querer."""
    ocupado = numero_em_uso(numero, excluir_id=veiculo_id)
    if ocupado:
        raise ValueError(
            f"O número {numero.strip()} já está cadastrado no cliente "
            f"{ocupado['cliente']}. O número do veículo é único no sistema.")
    with db() as conn:
        if modelo is None:
            conn.execute(
                "UPDATE veiculos SET numero=?, cliente=?, garagem=? WHERE id=?",
                (numero.strip(), cliente.strip(), garagem.strip().upper(), veiculo_id)
            )
        else:
            conn.execute(
                "UPDATE veiculos SET numero=?, cliente=?, garagem=?, modelo=? WHERE id=?",
                (numero.strip(), cliente.strip(), garagem.strip().upper(),
                 modelo.strip(), veiculo_id)
            )


def modelos_disponiveis() -> list[str]:
    """Nomes de kit ativos — são eles que valem como modelo de veículo."""
    with db() as conn:
        rows = conn.execute(
            "SELECT DISTINCT nome FROM kit_template WHERE ativo = 1 ORDER BY nome"
        ).fetchall()
    return [r["nome"] for r in rows]


def casar_modelo(texto: str, modelos: list[str] | None = None) -> str | None:
    """Casa o que foi digitado na planilha com um kit EXISTENTE — a planilha
    só trabalha em cima dos kits que já existem, nunca inventa um.

    Primeiro tenta bater exato (ignorando caixa e espaços duplicados):
    "mercedes Euro 6 diesel" é "Mercedes Euro 6 Diesel". Se não bater,
    procura o MAIS PRÓXIMO por similaridade — cobre abreviação e erro de
    digitação ("Mercedes Euro 6" → "Mercedes Euro 6 Diesel"). O corte de
    0.6 evita casar coisa que não tem nada a ver: aí devolve None e a
    importação reporta o erro em vez de chutar."""
    import difflib
    if modelos is None:
        modelos = modelos_disponiveis()
    chave = " ".join((texto or "").split()).lower()
    if not chave:
        return None
    por_chave = {" ".join(m.split()).lower(): m for m in modelos}
    if chave in por_chave:
        return por_chave[chave]
    proximos = difflib.get_close_matches(chave, list(por_chave), n=1, cutoff=0.6)
    return por_chave[proximos[0]] if proximos else None


def contar_sem_modelo(cliente: str | None = None) -> int:
    """Quantos veículos ativos ainda estão sem modelo — eles não aparecem
    em nenhum kit, então a tela precisa avisar."""
    sql = ("SELECT COUNT(*) FROM veiculos WHERE ativo = 1 "
           "AND TRIM(COALESCE(modelo, '')) = ''")
    params: list = []
    if cliente:
        sql += " AND cliente = ?"
        params.append(cliente)
    with db() as conn:
        return conn.execute(sql, params).fetchone()[0]


def definir_modelo(veiculo_id: int, modelo: str) -> None:
    """Grava só o modelo. Existe separado de atualizar() porque na tela do
    veículo o modelo vive junto do vínculo de kit (é o que decide em qual
    bipagem o veículo aparece), longe do formulário de número/cliente."""
    with db() as conn:
        conn.execute("UPDATE veiculos SET modelo = ? WHERE id = ?",
                     ((modelo or "").strip(), veiculo_id))


def atualizar_garagem(veiculo_id: int, garagem: str):
    """Atualiza só a garagem — usado ao finalizar um kit vinculado a um
    veículo cadastrado, propagando o que foi digitado na bipagem (que na
    prática também guarda o modelo do veículo) para o cadastro dele."""
    garagem = (garagem or "").strip().upper()
    if not garagem:
        return
    with db() as conn:
        conn.execute("UPDATE veiculos SET garagem = ? WHERE id = ?", (garagem, veiculo_id))


def esta_ocupado(veiculo_id: int) -> bool:
    """Um veículo fica bloqueado pra nova bipagem assim que tem qualquer
    kit associado — em andamento ou já finalizado, não importa em que
    estágio da esteira está, mesmo já entregue ao cliente. Só libera com
    liberar() (manual); a liberação vale só até a próxima sessão nova ser
    criada pra ele (consumir_liberacao)."""
    with db() as conn:
        v = conn.execute(
            "SELECT liberado_em FROM veiculos WHERE id = ?", (veiculo_id,)
        ).fetchone()
        if not v:
            return False
        if v["liberado_em"]:
            return False
        tem_sessao_ativa = conn.execute(
            "SELECT 1 FROM scan_session WHERE veiculo_id = ? AND status = 'em_andamento' LIMIT 1",
            (veiculo_id,)
        ).fetchone()
        if tem_sessao_ativa:
            return True
        tem_kit = conn.execute(
            "SELECT 1 FROM kit_record WHERE veiculo_id = ? LIMIT 1", (veiculo_id,)
        ).fetchone()
        return tem_kit is not None


def liberar(veiculo_id: int) -> None:
    with db() as conn:
        conn.execute(
            "UPDATE veiculos SET liberado_em = ? WHERE id = ?", (now_brt(), veiculo_id)
        )


def consumir_liberacao(veiculo_id: int) -> None:
    """Chamado ao criar uma sessão/kit novo pra esse veículo — a liberação
    manual vale só pra essa próxima atribuição, não pra sempre."""
    with db() as conn:
        conn.execute(
            "UPDATE veiculos SET liberado_em = NULL WHERE id = ?", (veiculo_id,)
        )


def desativar(veiculo_id: int):
    with db() as conn:
        conn.execute("UPDATE veiculos SET ativo=0 WHERE id=?", (veiculo_id,))


def reativar(veiculo_id: int):
    with db() as conn:
        conn.execute("UPDATE veiculos SET ativo=1 WHERE id=?", (veiculo_id,))


def deletar(veiculo_id: int):
    """Exclui o veículo preservando o histórico, que é o que a tela promete
    ao pedir a confirmação.

    kit_record e scan_session referenciam veiculos(id), então o DELETE
    direto batia na FOREIGN KEY e derrubava a requisição com erro 500. Em
    vez de bloquear a exclusão (como fazemos com usuário), aqui dá pra
    preservar tudo: o número do veículo já vive também num campo de texto
    nas duas tabelas, então basta garantir que esse texto esteja preenchido
    antes de soltar o vínculo. Os kits continuam existindo e continuam
    mostrando pra qual veículo foram."""
    with db() as conn:
        # Garante o número em texto antes de perder a referência — se o
        # texto já estiver preenchido, COALESCE/NULLIF mantém o que havia.
        conn.execute(
            "UPDATE kit_record SET veiculo = COALESCE(NULLIF(veiculo, ''), "
            "  (SELECT numero FROM veiculos WHERE id = ?)) "
            "WHERE veiculo_id = ?",
            (veiculo_id, veiculo_id)
        )
        conn.execute(
            "UPDATE scan_session SET veiculo = COALESCE(NULLIF(veiculo, ''), "
            "  (SELECT numero FROM veiculos WHERE id = ?)) "
            "WHERE veiculo_id = ?",
            (veiculo_id, veiculo_id)
        )
        conn.execute("UPDATE kit_record SET veiculo_id = NULL WHERE veiculo_id = ?", (veiculo_id,))
        conn.execute("UPDATE scan_session SET veiculo_id = NULL WHERE veiculo_id = ?", (veiculo_id,))
        conn.execute("DELETE FROM veiculos WHERE id=?", (veiculo_id,))


def importar_excel(file_bytes: bytes) -> dict:
    """Importa veículos da planilha. O veículo é identificado SÓ pelo
    número — ele é único no sistema inteiro, então reimportar nunca duplica,
    nem quando a planilha traz o mesmo número em outro cliente.

    A planilha MANDA: célula preenchida sobrescreve o que está no cadastro
    (cliente, garagem e modelo), célula vazia preserva. É o que deixa
    corrigir um modelo errado subindo a planilha de novo — antes ela só
    completava o que estava em branco e a correção nunca chegava ao cadastro.

    O modelo só trabalha em cima dos kits que JÁ EXISTEM — a planilha não
    cria template. O nome é casado com o kit mais próximo (caixa, espaço,
    abreviação e erro de digitação são tolerados); o que não casar com
    nenhum vira erro reportado na tela, e a linha importa sem mexer no
    modelo. Os casamentos não literais são listados no resultado
    (modelos_ajustados) pra o operador conferir se o "mais próximo" era o
    certo mesmo.

    A coluna Garagem é opcional — planilhas antigas, sem ela, continuam
    funcionando."""
    import openpyxl, io, sqlite3
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
    ws = wb.active
    headers = [str(c.value or "").strip().lower() for c in next(ws.iter_rows(min_row=1, max_row=1))]
    try:
        col_num = next(i for i, h in enumerate(headers) if "número" in h or "numero" in h or "veículo" in h or "veiculo" in h)
        col_cli = next(i for i, h in enumerate(headers) if "cliente" in h)
    except StopIteration:
        return {"inseridos": 0, "atualizados": 0, "ignorados": 0, "inativos": 0,
                "modelos_ajustados": [],
                "erros": ["Cabeçalhos não encontrados. Use 'Número do Veículo' e 'Cliente'."]}
    col_gar = next((i for i, h in enumerate(headers) if "garagem" in h), None)
    col_mod = next((i for i, h in enumerate(headers) if "modelo" in h or "kit" in h), None)

    modelos = modelos_disponiveis()
    inseridos = atualizados = ignorados = inativos = 0
    modelos_ajustados: list[dict] = []
    ajustes_vistos: set[str] = set()
    erros: list[str] = []
    with db() as conn:
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            if len(row) <= max(col_num, col_cli):
                ignorados += 1
                continue
            numero = str(row[col_num] or "").strip()
            cliente = str(row[col_cli] or "").strip()
            garagem = ""
            if col_gar is not None and len(row) > col_gar:
                garagem = str(row[col_gar] or "").strip()
            modelo_digitado = ""
            if col_mod is not None and len(row) > col_mod:
                modelo_digitado = str(row[col_mod] or "").strip()
            if not numero or not cliente:
                ignorados += 1
                continue

            # Casa com o kit existente mais próximo. Não casou = erro na
            # tela e a linha segue SEM tocar no modelo — melhor nenhum
            # modelo (a tela avisa) do que um modelo que não abre kit.
            modelo = ""
            if modelo_digitado:
                modelo = casar_modelo(modelo_digitado, modelos) or ""
                if not modelo:
                    erros.append(
                        f"Linha {row_idx}: modelo \"{modelo_digitado}\" não é "
                        f"parecido com nenhum kit existente — o veículo {numero} "
                        "foi importado sem alterar o modelo.")
                elif (" ".join(modelo.split()).lower()
                        != " ".join(modelo_digitado.split()).lower()
                        and modelo_digitado not in ajustes_vistos):
                    ajustes_vistos.add(modelo_digitado)
                    modelos_ajustados.append({"de": modelo_digitado, "para": modelo})

            # Cliente e garagem são criados na MESMA conexão/transação —
            # chamar clientes.criar()/garagens.criar() aqui abriria uma
            # segunda conexão SQLite enquanto esta ainda está com uma
            # transação aberta, travando o banco ("database is locked").
            try:
                conn.execute(
                    "INSERT INTO clientes (nome, criado_em) VALUES (?, ?)",
                    (cliente, now_brt())
                )
            except sqlite3.IntegrityError:
                pass  # cliente já existe
            if garagem:
                try:
                    conn.execute(
                        "INSERT INTO garagens (nome, criado_em) VALUES (?, ?)",
                        (garagem, now_brt())
                    )
                except sqlite3.IntegrityError:
                    pass  # garagem já existe

            # SÓ pelo número (sem cliente, sem filtro de ativo): o número é
            # único no sistema, então o mesmo número em outro cliente é o
            # mesmo veículo — a planilha atualiza o cadastro em vez de criar
            # a duplicata.
            existe = conn.execute(
                "SELECT id, cliente, garagem, modelo, ativo FROM veiculos "
                "WHERE LOWER(TRIM(numero)) = ?",
                (numero.lower(),)
            ).fetchone()

            if existe:
                # Célula preenchida sobrescreve; vazia preserva o cadastro.
                mudou = False
                if cliente and cliente != (existe["cliente"] or ""):
                    conn.execute("UPDATE veiculos SET cliente=? WHERE id=?",
                                 (cliente, existe["id"]))
                    mudou = True
                if garagem and garagem.upper() != (existe["garagem"] or "").strip().upper():
                    conn.execute("UPDATE veiculos SET garagem=? WHERE id=?",
                                 (garagem, existe["id"]))
                    mudou = True
                if modelo and modelo != (existe["modelo"] or "").strip():
                    conn.execute("UPDATE veiculos SET modelo=? WHERE id=?",
                                 (modelo, existe["id"]))
                    mudou = True
                if mudou:
                    atualizados += 1
                elif not existe["ativo"]:
                    inativos += 1
                else:
                    ignorados += 1
            else:
                conn.execute(
                    "INSERT INTO veiculos (numero, cliente, garagem, modelo, criado_em) "
                    "VALUES (?, ?, ?, ?, ?)",
                    (numero, cliente, garagem, modelo, now_brt())
                )
                inseridos += 1
    return {"inseridos": inseridos, "atualizados": atualizados,
            "ignorados": ignorados, "inativos": inativos,
            "modelos_ajustados": modelos_ajustados, "erros": erros}


def historico_kits(veiculo_id: int) -> list[dict]:
    with db() as conn:
        rows = conn.execute("""
            SELECT kr.kit_id, kt.nome AS kit_nome, kt.cliente,
                   kr.veiculo AS veiculo_texto, kr.garagem,
                   kr.finalizado_em, u.nome AS operador_nome,
                   CASE WHEN COUNT(kv.id) > 0 THEN 1 ELSE 0 END AS tem_verificacao
            FROM kit_record kr
            JOIN kit_template kt ON kt.id = kr.kit_template_id
            JOIN users u ON u.id = kr.operador_id
            LEFT JOIN kit_validacoes kv ON kv.kit_id = kr.kit_id
            WHERE kr.veiculo_id = ?
            GROUP BY kr.kit_id
            ORDER BY kr.finalizado_em DESC
        """, (veiculo_id,)).fetchall()
    return [dict(r) for r in rows]


def kits_para_vincular(veiculo_id: int, limite: int = 100) -> list[dict]:
    """Kits finalizados que podem ser vinculados a este veículo: os que não
    têm veículo nenhum e os que estão em OUTRO veículo (troca). Cada linha
    traz `veiculo_atual` preenchido quando já há vínculo — é o que faz a
    tela exigir motivo antes de trocar. Limitado aos mais recentes:
    atribuição manual é correção pontual, não varredura de histórico."""
    with db() as conn:
        rows = conn.execute("""
            SELECT kr.kit_id, kt.nome AS kit_nome, kt.cliente,
                   kr.finalizado_em,
                   kr.veiculo_id AS veiculo_id_atual,
                   COALESCE(v.numero, kr.veiculo, '') AS veiculo_atual
            FROM kit_record kr
            JOIN kit_template kt ON kt.id = kr.kit_template_id
            LEFT JOIN veiculos v ON v.id = kr.veiculo_id
            WHERE kr.veiculo_id IS NULL OR kr.veiculo_id != ?
            ORDER BY kr.finalizado_em DESC
            LIMIT ?
        """, (veiculo_id, limite)).fetchall()
    return [dict(r) for r in rows]


def clientes_disponiveis() -> list[str]:
    with db() as conn:
        rows = conn.execute(
            "SELECT DISTINCT cliente FROM kit_template WHERE ativo=1 ORDER BY cliente"
        ).fetchall()
    return [r[0] for r in rows]
