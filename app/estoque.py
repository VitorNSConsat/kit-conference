import re
from database import db, now_brt
import app.datas as datas_mod

# Status de compra — independente do status de quantidade (abaixo/proximo/ok).
# Vazio ('') = sem pendencia, nao aparece nenhum aviso.
STATUS_COMPRA = {
    "pedido": "Pedido ao fornecedor",
    "andamento": "Em andamento",
    "recebido": "Recebido",
}


def listar_estoque() -> list:
    with db() as conn:
        rows = conn.execute(
            "SELECT e.*, it.nome AS tipo_nome "
            "FROM estoque e "
            "JOIN item_tipo it ON it.id = e.item_tipo_id "
            "ORDER BY it.nome"
        ).fetchall()
    return [dict(r) for r in rows]


def buscar_por_codigo(codigo_barra: str) -> dict | None:
    with db() as conn:
        row = conn.execute(
            "SELECT e.*, it.nome AS tipo_nome "
            "FROM estoque e "
            "JOIN item_tipo it ON it.id = e.item_tipo_id "
            "WHERE e.codigo_barra = ?",
            (codigo_barra,)
        ).fetchone()
    return dict(row) if row else None


def buscar_por_tipo(item_tipo_id: int) -> dict | None:
    with db() as conn:
        row = conn.execute(
            "SELECT e.*, it.nome AS tipo_nome "
            "FROM estoque e "
            "JOIN item_tipo it ON it.id = e.item_tipo_id "
            "WHERE e.item_tipo_id = ?",
            (item_tipo_id,)
        ).fetchone()
    return dict(row) if row else None


def buscar_por_referencia(texto: str) -> dict | None:
    """Busca item de estoque pelo código de barras direto, ou pela URL do QR
    da etiqueta (formato .../estoque/<id>) — permite que o mesmo QR seja lido
    tanto durante a bipagem (desconta como um código normal) quanto fora dela
    (mostra a quantidade atual)."""
    texto = (texto or "").strip()
    if not texto:
        return None
    direto = buscar_por_codigo(texto)
    if direto:
        return direto
    m = re.search(r'/estoque/(\d+)/?$', texto)
    if m:
        return buscar_por_id(int(m.group(1)))
    return None


def buscar_por_id(estoque_id: int) -> dict | None:
    with db() as conn:
        row = conn.execute(
            "SELECT e.*, it.nome AS tipo_nome "
            "FROM estoque e "
            "JOIN item_tipo it ON it.id = e.item_tipo_id "
            "WHERE e.id = ?",
            (estoque_id,)
        ).fetchone()
    return dict(row) if row else None


def criar_estoque(item_tipo_id: int, codigo_barra: str,
                  quantidade_inicial: int, quantidade_minima: int,
                  criado_por: int) -> int:
    with db() as conn:
        cur = conn.execute(
            "INSERT INTO estoque (item_tipo_id, codigo_barra, quantidade_atual, quantidade_minima, criado_em) "
            "VALUES (?, ?, ?, ?, ?)",
            (item_tipo_id, codigo_barra, quantidade_inicial, quantidade_minima, now_brt())
        )
        estoque_id = cur.lastrowid
        if quantidade_inicial > 0:
            conn.execute(
                "INSERT INTO estoque_movimentos "
                "(estoque_id, tipo, quantidade, criado_por, observacao, criado_em) "
                "VALUES (?, 'entrada', ?, ?, 'Estoque inicial', ?)",
                (estoque_id, quantidade_inicial, criado_por, now_brt())
            )
    return estoque_id


def repor_estoque(estoque_id: int, quantidade: int,
                  criado_por: int, observacao: str = "") -> None:
    with db() as conn:
        conn.execute(
            "UPDATE estoque SET quantidade_atual = quantidade_atual + ? WHERE id = ?",
            (quantidade, estoque_id)
        )
        conn.execute(
            "INSERT INTO estoque_movimentos "
            "(estoque_id, tipo, quantidade, criado_por, observacao, criado_em) "
            "VALUES (?, 'entrada', ?, ?, ?, ?)",
            (estoque_id, quantidade, criado_por, observacao or "Reposição", now_brt())
        )


def registrar_saida(estoque_id: int, quantidade: int,
                    sessao_id: int, criado_por: int) -> None:
    with db() as conn:
        conn.execute(
            "UPDATE estoque SET quantidade_atual = quantidade_atual - ? WHERE id = ?",
            (quantidade, estoque_id)
        )
        conn.execute(
            "INSERT INTO estoque_movimentos "
            "(estoque_id, tipo, quantidade, sessao_id, criado_por, observacao, criado_em) "
            "VALUES (?, 'saida', ?, ?, ?, 'Kit', ?)",
            (estoque_id, quantidade, sessao_id, criado_por, now_brt())
        )


def registrar_sobressalente(estoque_id: int, quantidade: int, cliente: str,
                            criado_por: int, observacao: str = "") -> None:
    """Baixa manual de peça sobressalente enviada numa instalação — fora do
    que o kit bipado já contabiliza. Desconta do estoque e grava o cliente,
    pra dar pra consultar depois em Relatórios quanto já foi enviado de
    sobressalente pra cada um. Bloqueia se não tiver saldo suficiente (ação
    deliberada de escritório, não bipagem em campo — vale travar e pedir
    pra corrigir o estoque antes, ao contrário do desconto automático do
    kit que prefere deixar negativo a travar o operador)."""
    cliente = cliente.strip()
    if not cliente:
        raise ValueError("Informe o cliente.")
    if quantidade <= 0:
        raise ValueError("Quantidade deve ser maior que zero.")
    with db() as conn:
        atual = conn.execute(
            "SELECT quantidade_atual FROM estoque WHERE id = ?", (estoque_id,)
        ).fetchone()
        if atual is None:
            raise ValueError("Item de estoque não encontrado.")
        if atual["quantidade_atual"] < quantidade:
            raise ValueError(
                f"Estoque insuficiente ({atual['quantidade_atual']} disponíveis, "
                f"{quantidade} necessários)."
            )
        conn.execute(
            "UPDATE estoque SET quantidade_atual = quantidade_atual - ? WHERE id = ?",
            (quantidade, estoque_id)
        )
        conn.execute(
            "INSERT INTO estoque_movimentos "
            "(estoque_id, tipo, quantidade, cliente, criado_por, observacao, criado_em) "
            "VALUES (?, 'sobressalente', ?, ?, ?, ?, ?)",
            (estoque_id, quantidade, cliente, criado_por, observacao.strip() or None, now_brt())
        )


def registrar_sobressalentes_em_lote(linhas: list[dict], cliente: str,
                                     criado_por: int) -> dict:
    """Vários sobressalentes pro mesmo cliente num envio só.

    TUDO OU NADA: primeiro confere o saldo de todas as linhas, e só então
    aplica. registrar_sobressalente() já bloqueia por estoque insuficiente —
    aplicar linha a linha faria as primeiras saírem e as últimas falharem,
    deixando um envio pela metade que o operador teria que reconstruir de
    cabeça. Validando antes, ele corrige o estoque e reenvia a lista inteira
    sem risco de mandar nada duas vezes.

    Linhas repetidas do mesmo item são SOMADAS antes da conferência: pedir
    3 e depois 4 do mesmo item é um pedido de 7, e checar 3 e 4 em separado
    passaria mesmo com só 5 em estoque.

    O registro em si continua saindo por registrar_sobressalente(), item a
    item — o movimento, o cliente e a auditoria ficam idênticos ao envio
    individual."""
    cliente = (cliente or "").strip()
    if not cliente:
        raise ValueError("Informe o cliente.")

    # Agrupa por item, guardando as observações de cada pedido.
    pedidos: dict[int, dict] = {}
    for l in linhas:
        try:
            eid = int(l.get("estoque_id") or 0)
            qtd = int(l.get("quantidade") or 0)
        except (TypeError, ValueError):
            continue
        if eid <= 0 or qtd <= 0:
            continue
        p = pedidos.setdefault(eid, {"quantidade": 0, "observacoes": []})
        p["quantidade"] += qtd
        obs = (l.get("observacao") or "").strip()
        if obs and obs not in p["observacoes"]:
            p["observacoes"].append(obs)
    if not pedidos:
        raise ValueError("Nenhum item informado.")

    # 1) Confere TODOS os saldos antes de mexer em qualquer um.
    faltas = []
    with db() as conn:
        for eid, p in pedidos.items():
            row = conn.execute(
                "SELECT e.quantidade_atual, it.nome AS tipo_nome FROM estoque e "
                "JOIN item_tipo it ON it.id = e.item_tipo_id WHERE e.id = ?",
                (eid,)).fetchone()
            if row is None:
                faltas.append(f"item de estoque {eid} não encontrado")
            elif row["quantidade_atual"] < p["quantidade"]:
                faltas.append(f"{row['tipo_nome']}: pedido {p['quantidade']}, "
                              f"disponível {row['quantidade_atual']}")
    if faltas:
        raise ValueError("Estoque insuficiente — " + "; ".join(faltas)
                         + ". Nada foi enviado.")

    # 2) Só agora registra, pela mesma função do envio individual.
    enviados = []
    for eid, p in pedidos.items():
        registrar_sobressalente(eid, p["quantidade"], cliente, criado_por,
                                " | ".join(p["observacoes"]))
        enviados.append({"estoque_id": eid, "quantidade": p["quantidade"]})
    return {"enviados": enviados, "itens": len(enviados),
            "unidades": sum(e["quantidade"] for e in enviados)}


def listar_sobressalentes(data_ini: str = "", data_fim: str = "", cliente: str = "") -> list[dict]:
    query = (
        "SELECT em.*, e.codigo_barra, it.nome AS tipo_nome, u.nome AS operador_nome "
        "FROM estoque_movimentos em "
        "JOIN estoque e ON e.id = em.estoque_id "
        "JOIN item_tipo it ON it.id = e.item_tipo_id "
        "LEFT JOIN users u ON u.id = em.criado_por "
        "WHERE em.tipo = 'sobressalente'"
    )
    params: list = []
    sql_data, p_data = datas_mod.clausula("em.criado_em", data_ini, data_fim)
    query += sql_data
    params += p_data
    if cliente:
        query += " AND em.cliente = ?"
        params.append(cliente)
    query += " ORDER BY em.criado_em DESC"
    with db() as conn:
        rows = conn.execute(query, params).fetchall()
    return [dict(r) for r in rows]


def importar_ajustes_xlsx(conteudo: bytes, criado_por: int) -> dict:
    """Atualização em massa do estoque pela planilha baixada da própria tela.

    O item é identificado pelo ID (coluna oculta da planilha modelo) e, na
    falta dele, pelo código de barras — nunca pelo nome do tipo, que pode
    repetir e não é chave de nada.

    Cada linha alterada passa pelas MESMAS funções da tela
    (corrigir_quantidade / atualizar_minimo / atualizar_status_compra), então
    o movimento de correção e o log saem idênticos ao ajuste feito à mão: uma
    planilha não é um caminho paralelo pra mexer no estoque sem deixar
    rastro. Linha sem mudança não gera movimento nenhum — reimportar a mesma
    planilha duas vezes não polui o histórico.

    Devolve o que mudou, item a item, pra tela poder mostrar a lista em vez
    de só um número."""
    import openpyxl, io
    wb = openpyxl.load_workbook(io.BytesIO(conteudo))
    ws = wb.active
    cabec = [str(c.value or "").strip().lower()
             for c in next(ws.iter_rows(min_row=1, max_row=1))]

    def coluna(*pistas):
        for i, h in enumerate(cabec):
            if any(p in h for p in pistas):
                return i
        return None

    col_id = coluna("id")
    col_cod = coluna("código de barras", "codigo de barras", "código", "codigo")
    col_qtd = coluna("quantidade atual", "qtd atual", "quantidade")
    col_min = coluna("mínima", "minima")
    col_sts = coluna("status de compra", "status")
    if col_id is None and col_cod is None:
        return {"atualizados": [], "ignorados": 0,
                "erros": ["A planilha precisa da coluna 'ID' ou 'Código de Barras'. "
                          "Baixe o modelo pela própria tela de estoque."]}

    with db() as conn:
        por_id = {r["id"]: dict(r) for r in conn.execute(
            "SELECT e.*, it.nome AS tipo_nome FROM estoque e "
            "JOIN item_tipo it ON it.id = e.item_tipo_id").fetchall()}
    por_codigo = {(v["codigo_barra"] or "").strip().lower(): v for v in por_id.values()}

    def celula(linha, idx):
        if idx is None or idx >= len(linha):
            return None
        return linha[idx]

    def inteiro(valor):
        """Célula vazia = 'não mexer nisso', não zero. Zerar estoque é uma
        decisão, e apagar a célula por engano não pode virar zero."""
        if valor is None or str(valor).strip() == "":
            return None
        try:
            return int(float(str(valor).strip().replace(",", ".")))
        except ValueError:
            raise ValueError(f"'{valor}' não é um número")

    atualizados, erros = [], []
    ignorados = 0
    for n, linha in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        if not any(c is not None and str(c).strip() for c in linha):
            continue
        item = None
        bruto_id = celula(linha, col_id)
        if bruto_id is not None and str(bruto_id).strip().isdigit():
            item = por_id.get(int(str(bruto_id).strip()))
        if item is None:
            codigo = str(celula(linha, col_cod) or "").strip().lower()
            item = por_codigo.get(codigo) if codigo else None
        if item is None:
            ignorados += 1
            erros.append(f"Linha {n}: item de estoque não encontrado "
                         "(ID e código de barras não batem com nada cadastrado).")
            continue

        try:
            nova_qtd = inteiro(celula(linha, col_qtd))
            novo_min = inteiro(celula(linha, col_min))
        except ValueError as e:
            ignorados += 1
            erros.append(f"Linha {n} ({item['tipo_nome']}): {e}.")
            continue
        novo_sts = celula(linha, col_sts)
        novo_sts = str(novo_sts).strip() if novo_sts is not None else None

        mudancas = []
        if nova_qtd is not None and nova_qtd != item["quantidade_atual"]:
            if nova_qtd < 0:
                ignorados += 1
                erros.append(f"Linha {n} ({item['tipo_nome']}): quantidade negativa.")
                continue
            corrigir_quantidade(item["id"], nova_qtd, criado_por)
            mudancas.append(f"quantidade {item['quantidade_atual']} → {nova_qtd}")
        if novo_min is not None and novo_min != item["quantidade_minima"]:
            atualizar_minimo(item["id"], novo_min, criado_por)
            mudancas.append(f"mínima {item['quantidade_minima']} → {novo_min}")
        if novo_sts is not None and novo_sts != (item["status_compra"] or ""):
            if novo_sts and novo_sts not in STATUS_COMPRA:
                ignorados += 1
                erros.append(f"Linha {n} ({item['tipo_nome']}): status de compra "
                             f"'{novo_sts}' inválido. Use um de: "
                             + ", ".join(STATUS_COMPRA) + " (ou deixe vazio).")
                continue
            atualizar_status_compra(item["id"], novo_sts, criado_por)
            mudancas.append(f"status de compra → {novo_sts or '(nenhum)'}")

        if mudancas:
            atualizados.append({"tipo_nome": item["tipo_nome"],
                                "codigo_barra": item["codigo_barra"],
                                "mudancas": mudancas})
        else:
            ignorados += 1

    return {"atualizados": atualizados, "ignorados": ignorados, "erros": erros}


def reverter_saidas_sessao(sessao_id: int) -> None:
    """Restaura estoque das saídas de uma sessão cancelada.

    Devolve o SALDO LÍQUIDO, não a soma das saídas: a troca de kit de um kit
    pronto já devolve ao estoque o que sai do kit, gravando uma 'entrada'
    com o sessao_id. Somando só as saídas, essas unidades voltariam duas
    vezes e o estoque ganharia peça que não existe. As entradas de estorno
    da bipagem (remover_item) não têm sessao_id, então não entram nesta
    conta e continuam valendo como sempre."""
    with db() as conn:
        saldos = conn.execute(
            "SELECT estoque_id, "
            "  SUM(CASE WHEN tipo = 'saida'   THEN quantidade ELSE 0 END) "
            "- SUM(CASE WHEN tipo = 'entrada' THEN quantidade ELSE 0 END) AS total "
            "FROM estoque_movimentos "
            "WHERE sessao_id = ? AND tipo IN ('saida', 'entrada') "
            "GROUP BY estoque_id",
            (sessao_id,)
        ).fetchall()
        for s in saldos:
            if s["total"] > 0:
                conn.execute(
                    "UPDATE estoque SET quantidade_atual = quantidade_atual + ? WHERE id = ?",
                    (s["total"], s["estoque_id"])
                )
        # Marca saídas E entradas: sem marcar as entradas, uma segunda
        # chamada recalcularia o líquido contra saídas já canceladas e
        # devolveria a diferença de novo.
        conn.execute(
            "UPDATE estoque_movimentos SET tipo = 'saida_cancelada' "
            "WHERE sessao_id = ? AND tipo = 'saida'",
            (sessao_id,)
        )
        conn.execute(
            "UPDATE estoque_movimentos SET tipo = 'entrada_cancelada' "
            "WHERE sessao_id = ? AND tipo = 'entrada'",
            (sessao_id,)
        )


def listar_historico(estoque_id: int, limit: int = 100) -> list:
    with db() as conn:
        rows = conn.execute(
            "SELECT em.*, u.nome AS operador_nome "
            "FROM estoque_movimentos em "
            "LEFT JOIN users u ON u.id = em.criado_por "
            "WHERE em.estoque_id = ? "
            "ORDER BY em.criado_em DESC LIMIT ?",
            (estoque_id, limit)
        ).fetchall()
    return [dict(r) for r in rows]


def listar_historico_completo() -> list:
    """Todas as movimentações de todos os itens de estoque — usado na
    exportação em Excel (aba Histórico)."""
    with db() as conn:
        rows = conn.execute(
            "SELECT em.*, e.codigo_barra, it.nome AS tipo_nome, u.nome AS operador_nome "
            "FROM estoque_movimentos em "
            "JOIN estoque e ON e.id = em.estoque_id "
            "JOIN item_tipo it ON it.id = e.item_tipo_id "
            "LEFT JOIN users u ON u.id = em.criado_por "
            "ORDER BY it.nome, em.criado_em DESC"
        ).fetchall()
    return [dict(r) for r in rows]


def alertas_abaixo_minimo() -> list:
    """Retorna itens de estoque com quantidade abaixo ou igual ao mínimo."""
    with db() as conn:
        rows = conn.execute(
            "SELECT e.*, it.nome AS tipo_nome "
            "FROM estoque e "
            "JOIN item_tipo it ON it.id = e.item_tipo_id "
            "WHERE e.quantidade_atual <= e.quantidade_minima "
            "ORDER BY (e.quantidade_atual - e.quantidade_minima), it.nome"
        ).fetchall()
    return [dict(r) for r in rows]


# ── Aviso de estoque baixo — o que aparece, onde e por quanto tempo ──────────
# O aviso é a faixa amarela do topo. Antes era fixo: todos os itens no mínimo,
# em todas as telas, pra sempre. Numa lista grande virava paisagem — e o que
# vira paisagem ninguém lê.
ALERTA_PADRAO = {
    "alerta_ativo": "1",
    # % acima do mínimo que já conta como "atenção". 0 = só no mínimo ou abaixo.
    "alerta_margem": "0",
    # Quantos itens cabem na faixa (0 = todos). O resto vira "e mais N".
    "alerta_limite": "8",
    # Some sozinho depois de N segundos (0 = fica até fechar).
    "alerta_segundos": "0",
    # Em quais telas: todas | estoque (Itens & Estoque) | inicio (só a raiz).
    "alerta_telas": "todas",
    "alerta_cor_critico": "#c0392b",
    "alerta_cor_atencao": "#f0ad4e",
}
ALERTA_TELAS = {
    "todas": "Em todas as telas",
    "estoque": "Só em Itens & Estoque",
    "inicio": "Só na tela inicial",
}
_COR_HEX = re.compile(r"^#[0-9a-fA-F]{6}$")


def get_alerta_config() -> dict:
    cfg = dict(ALERTA_PADRAO)
    with db() as conn:
        for r in conn.execute("SELECT chave, valor FROM estoque_config").fetchall():
            if r["chave"] in cfg:
                cfg[r["chave"]] = r["valor"]
    # Números saem prontos pra usar; o resto continua texto.
    for chave in ("alerta_ativo", "alerta_margem", "alerta_limite", "alerta_segundos"):
        try:
            cfg[chave] = max(0, int(cfg[chave]))
        except (TypeError, ValueError):
            cfg[chave] = int(ALERTA_PADRAO[chave])
    if cfg["alerta_telas"] not in ALERTA_TELAS:
        cfg["alerta_telas"] = ALERTA_PADRAO["alerta_telas"]
    for chave in ("alerta_cor_critico", "alerta_cor_atencao"):
        if not _COR_HEX.match(str(cfg[chave] or "")):
            cfg[chave] = ALERTA_PADRAO[chave]
    return cfg


def salvar_alerta_config(valores: dict) -> None:
    """Grava só as chaves conhecidas. Valor inválido cai no padrão em vez de
    quebrar a faixa de aviso em todas as telas do sistema."""
    limpos = {}
    for chave, padrao in ALERTA_PADRAO.items():
        if chave not in valores:
            continue
        v = valores[chave]
        if chave in ("alerta_ativo", "alerta_margem", "alerta_limite", "alerta_segundos"):
            try:
                v = str(max(0, int(v)))
            except (TypeError, ValueError):
                v = padrao
        elif chave == "alerta_telas":
            v = v if v in ALERTA_TELAS else padrao
        else:
            v = v if _COR_HEX.match(str(v or "")) else padrao
        limpos[chave] = v
    with db() as conn:
        for chave, valor in limpos.items():
            conn.execute(
                "INSERT INTO estoque_config (chave, valor) VALUES (?, ?) "
                "ON CONFLICT(chave) DO UPDATE SET valor = excluded.valor",
                (chave, valor)
            )


def nivel_do_item(item: dict, margem: int = 0) -> str:
    """zerado > critico (no mínimo ou abaixo) > atencao (dentro da margem
    acima do mínimo) > ok. Uma função só pra a faixa de aviso, o filtro da
    lista e a cor da linha não discordarem entre si."""
    atual = item.get("quantidade_atual") or 0
    minimo = item.get("quantidade_minima") or 0
    if atual <= 0:
        return "zerado"
    if atual <= minimo:
        return "critico"
    if margem and atual <= minimo * (1 + margem / 100):
        return "atencao"
    return "ok"


def alertas_para_banner(caminho: str = "") -> dict:
    """O que a faixa de aviso deve mostrar nesta tela.

    Devolve {itens, total, cfg}: `itens` já cortado pelo limite e `total` com
    quantos existem de verdade, pra a faixa poder dizer "e mais N" em vez de
    esconder o resto sem avisar."""
    cfg = get_alerta_config()
    vazio = {"itens": [], "total": 0, "cfg": cfg}
    if not cfg["alerta_ativo"]:
        return vazio
    escopo = cfg["alerta_telas"]
    if escopo == "inicio" and caminho != "/":
        return vazio
    if escopo == "estoque" and not (caminho.startswith("/admin/items")
                                    or caminho.startswith("/admin/estoque")):
        return vazio
    margem = cfg["alerta_margem"]
    with db() as conn:
        rows = conn.execute(
            "SELECT e.*, it.nome AS tipo_nome "
            "FROM estoque e JOIN item_tipo it ON it.id = e.item_tipo_id "
            "WHERE e.quantidade_atual <= e.quantidade_minima * (1 + ? / 100.0) "
            "ORDER BY (e.quantidade_atual - e.quantidade_minima), it.nome",
            (margem,)
        ).fetchall()
    itens = [dict(r) for r in rows]
    for i in itens:
        i["nivel"] = nivel_do_item(i, margem)
    total = len(itens)
    limite = cfg["alerta_limite"]
    return {"itens": itens[:limite] if limite else itens, "total": total, "cfg": cfg}


def atualizar_codigo_barra(estoque_id: int, novo_codigo: str) -> None:
    novo_codigo = (novo_codigo or "").strip()
    if not novo_codigo:
        raise ValueError("Código de barras não pode ser vazio.")
    with db() as conn:
        conn.execute(
            "UPDATE estoque SET codigo_barra = ? WHERE id = ?",
            (novo_codigo, estoque_id)
        )


def corrigir_quantidade(estoque_id: int, nova_quantidade: int, criado_por: int) -> None:
    """Corrige a quantidade atual diretamente para o valor informado (ajuste
    absoluto) — para acertar contagens divergentes sem precisar calcular a
    diferença e usar repor/ajustar."""
    if nova_quantidade < 0:
        raise ValueError("Quantidade não pode ser negativa.")
    with db() as conn:
        atual = conn.execute(
            "SELECT quantidade_atual FROM estoque WHERE id = ?", (estoque_id,)
        ).fetchone()
        if atual is None:
            raise ValueError("Item de estoque não encontrado.")
        antigo = atual["quantidade_atual"]
        conn.execute(
            "UPDATE estoque SET quantidade_atual = ? WHERE id = ?",
            (nova_quantidade, estoque_id)
        )
        conn.execute(
            "INSERT INTO estoque_movimentos "
            "(estoque_id, tipo, quantidade, criado_por, observacao, criado_em) "
            "VALUES (?, 'correcao', ?, ?, ?, ?)",
            (estoque_id, nova_quantidade, criado_por,
             f"Quantidade corrigida: {antigo} → {nova_quantidade}", now_brt())
        )


def atualizar_minimo(estoque_id: int, novo_minimo: int, criado_por: int) -> None:
    with db() as conn:
        atual = conn.execute(
            "SELECT quantidade_minima FROM estoque WHERE id = ?", (estoque_id,)
        ).fetchone()
        if atual is None:
            return
        antigo = atual["quantidade_minima"]
        conn.execute(
            "UPDATE estoque SET quantidade_minima = ? WHERE id = ?",
            (novo_minimo, estoque_id)
        )
        conn.execute(
            "INSERT INTO estoque_movimentos "
            "(estoque_id, tipo, quantidade, criado_por, observacao, criado_em) "
            "VALUES (?, 'ajuste_minimo', ?, ?, ?, ?)",
            (estoque_id, novo_minimo, criado_por,
             f"Mínimo alterado: {antigo} → {novo_minimo}", now_brt())
        )


def atualizar_status_compra(estoque_id: int, novo_status: str, criado_por: int) -> None:
    """Marca se o item já foi pedido ao fornecedor / está a caminho / chegou.
    Independente do status de quantidade (abaixo/proximo/ok) — um item pode
    estar OK em quantidade e ainda ter uma reposição futura já encomendada."""
    novo_status = novo_status if novo_status in STATUS_COMPRA else ""
    with db() as conn:
        atual = conn.execute(
            "SELECT status_compra FROM estoque WHERE id = ?", (estoque_id,)
        ).fetchone()
        if atual is None:
            return
        antigo = atual["status_compra"] or ""
        if antigo == novo_status:
            return
        conn.execute(
            "UPDATE estoque SET status_compra = ? WHERE id = ?",
            (novo_status, estoque_id)
        )
        conn.execute(
            "INSERT INTO estoque_movimentos "
            "(estoque_id, tipo, quantidade, criado_por, observacao, criado_em) "
            "VALUES (?, 'status_compra', 0, ?, ?, ?)",
            (estoque_id, criado_por,
             f"Status de compra: {STATUS_COMPRA.get(antigo, 'Sem pendência')} → "
             f"{STATUS_COMPRA.get(novo_status, 'Sem pendência')}", now_brt())
        )


def ajustar_quantidade(estoque_id: int, tipo: str, quantidade: int,
                       motivo: str, criado_por: int) -> int:
    """Ajuste manual (entrada ou saída). Retorna a nova quantidade."""
    if tipo not in ("entrada", "saida"):
        raise ValueError("Tipo inválido.")
    with db() as conn:
        est = conn.execute("SELECT * FROM estoque WHERE id = ?", (estoque_id,)).fetchone()
        if not est:
            raise ValueError("Item não encontrado.")
        nova = est["quantidade_atual"] + quantidade if tipo == "entrada" \
               else est["quantidade_atual"] - quantidade
        if nova < 0:
            raise ValueError(
                f"Estoque insuficiente. Disponível: {est['quantidade_atual']}, "
                f"solicitado: {quantidade}."
            )
        conn.execute("UPDATE estoque SET quantidade_atual = ? WHERE id = ?", (nova, estoque_id))
        conn.execute(
            "INSERT INTO estoque_movimentos "
            "(estoque_id, tipo, quantidade, criado_por, observacao, criado_em) "
            "VALUES (?, ?, ?, ?, ?, ?)",
            (estoque_id, tipo, quantidade, criado_por, motivo, now_brt())
        )
    return nova


def deletar_estoque(estoque_id: int) -> None:
    with db() as conn:
        conn.execute("DELETE FROM estoque_movimentos WHERE estoque_id = ?", (estoque_id,))
        conn.execute("DELETE FROM estoque WHERE id = ?", (estoque_id,))


def reconciliar_saidas_producao(criado_por: int) -> list[dict]:
    """Corrige retroativamente o estoque de duas categorias de bug
    histórico, ambas já corrigidas no código a partir de agora:

    1. Conjunto (codigo_barra 'COMP:...') que nunca descontava estoque.
    2. Item de quantidade em lote (metro ou mais de 1 unidade sob um único
       código, sem conjunto/serial — ex: rolo de cabo) que só descontava 1
       unidade fixa na criação do patrimônio (ou nada, se o patrimônio já
       existia) em vez da quantidade realmente confirmada.

    scan_session_items.estoque_debitado marca quanto já foi debitado por
    linha (0 = nada ainda) — cada linha corrigida é atualizada com
    estoque_debitado = quantidade, então rodar de novo não desconta duas
    vezes. Sem checar saldo antes: mesma filosofia do resto do sistema —
    não trava, só sinaliza que a contagem física precisa ser conferida.

    Ressalva conhecida (categoria 2): quando o patrimônio era NOVO na época
    do bug, 1 unidade cega já tinha sido descontada do total do tipo (não
    fica registrada por linha) — a correção pode ficar até 1 unidade a mais
    por linha afetada nesse caso específico. Ainda assim é muito mais
    preciso que o buraco atual (que pode estar faltando dezenas de
    unidades). Retorna um resumo por tipo de item corrigido."""
    with db() as conn:
        linhas_conjunto = conn.execute(
            "SELECT ssi.id, ssi.item_tipo_id, ssi.quantidade, it.nome AS tipo_nome, e.id AS estoque_id "
            "FROM scan_session_items ssi "
            "JOIN item_tipo it ON it.id = ssi.item_tipo_id "
            "JOIN estoque e ON e.item_tipo_id = ssi.item_tipo_id "
            "WHERE ssi.codigo_barra LIKE 'COMP:%' AND ssi.estoque_debitado = 0"
        ).fetchall()
        linhas_lote = conn.execute(
            "SELECT ssi.id, ssi.item_tipo_id, ssi.quantidade, it.nome AS tipo_nome, e.id AS estoque_id "
            "FROM scan_session_items ssi "
            "JOIN scan_session ss ON ss.id = ssi.sessao_id "
            "JOIN kit_template_items kti ON kti.kit_template_id = ss.kit_template_id "
            "                            AND kti.item_tipo_id = ssi.item_tipo_id "
            "JOIN item_tipo it ON it.id = ssi.item_tipo_id "
            "JOIN estoque e ON e.item_tipo_id = ssi.item_tipo_id "
            "WHERE ssi.estoque_debitado = 0 "
            "  AND ssi.codigo_barra NOT LIKE 'COMP:%' "
            "  AND ssi.codigo_barra NOT LIKE 'ESTOQUE:%' "
            "  AND COALESCE(kti.requer_serial, 0) = 0 "
            "  AND kti.componente_codigo IS NULL "
            "  AND (COALESCE(it.unidade, 'un') = 'm' OR kti.quantidade_exigida > 1)"
        ).fetchall()

    linhas = [dict(r) for r in linhas_conjunto] + [dict(r) for r in linhas_lote]
    if not linhas:
        return []

    # Consolida por tipo — 1 movimento de saída por tipo, não 1 por linha
    por_tipo: dict[int, dict] = {}
    ids_por_tipo: dict[int, list] = {}
    for r in linhas:
        c = por_tipo.setdefault(
            r["item_tipo_id"],
            {"tipo_nome": r["tipo_nome"], "estoque_id": r["estoque_id"], "qtd": 0}
        )
        c["qtd"] += r["quantidade"] or 1
        ids_por_tipo.setdefault(r["item_tipo_id"], []).append(r["id"])

    resumo = []
    for tipo_id, c in por_tipo.items():
        ids = ids_por_tipo[tipo_id]
        placeholders = ",".join("?" * len(ids))
        with db() as conn:
            conn.execute(
                "UPDATE estoque SET quantidade_atual = quantidade_atual - ? WHERE id = ?",
                (c["qtd"], c["estoque_id"])
            )
            conn.execute(
                "INSERT INTO estoque_movimentos "
                "(estoque_id, tipo, quantidade, criado_por, observacao, criado_em) "
                "VALUES (?, 'saida', ?, ?, ?, ?)",
                (c["estoque_id"], c["qtd"], criado_por,
                 "Correção retroativa: produção não debitava estoque corretamente antes da correção do bug",
                 now_brt())
            )
            conn.execute(
                f"UPDATE scan_session_items SET estoque_debitado = quantidade WHERE id IN ({placeholders})",
                ids
            )
        resumo.append({"tipo_nome": c["tipo_nome"], "quantidade": c["qtd"]})
    return resumo
