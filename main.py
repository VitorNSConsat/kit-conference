import json
import os
import re
import uuid
from collections import Counter
from datetime import datetime, timezone, timedelta

# Brasília Time (UTC-3) — garante horário correto independente do fuso do servidor
BRT = timezone(timedelta(hours=-3))
from urllib.parse import quote
from fastapi import (FastAPI, Request, Form, Query, WebSocket, WebSocketDisconnect,
                     UploadFile, File, HTTPException)
from fastapi.responses import HTMLResponse, RedirectResponse, PlainTextResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from starlette.middleware.sessions import SessionMiddleware
from starlette.middleware.base import BaseHTTPMiddleware
from starlette.middleware.gzip import GZipMiddleware
from dotenv import load_dotenv

from database import init_db, db, now_brt
from app.auth import (hash_password, verify_password, get_current_user,
                      require_login, require_admin, require_permission, is_admin)
import app.items as items_mod
import app.kit_templates as templates_mod
import app.sessions as sessions_mod
import app.zpl as zpl_mod
import app.print_queue as pq_mod
import app.estoque as estoque_mod
import app.validacoes as validacoes_mod
import app.veiculos as veiculos_mod
import app.clientes as clientes_mod
import app.garagens as garagens_mod
import app.codigos_gerados as codigos_gerados_mod
import app.prateleira as prateleira_mod
import app.pedidos as pedidos_mod
import app.consumo as consumo_mod
import app.auditoria as auditoria_mod
import app.usuarios as usuarios_mod
import app.producao as producao_mod
import app.permissoes as permissoes_mod
import app.paginacao as paginacao_mod
import app.datas as datas_mod
import app.filtros as filtros_mod
import app.backup as backup_mod
import app.inatividade as inatividade_mod
import app.importacoes as importacoes_mod
import app.remessas as remessas_mod

load_dotenv()

_MOBILE_UA = re.compile(r'(Mobile|Android|iPhone|iPad|iPod)', re.IGNORECASE)

# Rotas GET permitidas em dispositivos móveis (bipagem + estoque)
# /funcionalidades entra aqui porque é só leitura: o portão existe pra manter
# tela de administração fora do celular, não o manual — que é justamente o que
# o operador quer consultar em campo, com o celular na mão.
_MOBILE_OK_EXACT = {'/mobile', '/login', '/logout', '/ping', '/cert', '/estoque',
                    '/funcionalidades'}
_MOBILE_OK_PREFIX = ('/static/', '/session/', '/ws/', '/kit/', '/admin/estoque', '/estoque/', '/prateleira/', '/producao/')


class _MobileGateMiddleware(BaseHTTPMiddleware):
    async def dispatch(self, request: Request, call_next):
        if request.method != 'GET':
            return await call_next(request)
        if not _MOBILE_UA.search(request.headers.get('user-agent', '')):
            return await call_next(request)
        path = request.url.path
        if path in _MOBILE_OK_EXACT or any(path.startswith(p) for p in _MOBILE_OK_PREFIX):
            return await call_next(request)
        return RedirectResponse('/mobile', status_code=302)


class _PaginaPermitidaMiddleware(BaseHTTPMiddleware):
    """Porteiro das telas: esconde uma página inteira de quem não pode vê-la.

    Fica no middleware, e não rota a rota, porque a regra é "esta ÁREA do
    sistema" — cada tela tem várias rotas (lista, detalhe, exportação,
    formulários) e uma delas ficaria de fora na hora de acrescentar rota
    nova. Vale pra qualquer método: esconder a tela e deixar o POST dela
    aberto seria esconder o botão, não a permissão.

    Deslogado não é "sem permissão": aí quem manda é a própria rota, que
    pode ser pública de propósito (o kit lido pelo QR, por exemplo)."""

    async def dispatch(self, request: Request, call_next):
        chave = permissoes_mod.permissao_da_rota(request.url.path)
        if chave:
            user = None
            try:
                user = get_current_user(request)
            except Exception:
                pass
            if user and not permissoes_mod.tem_permissao(user, chave):
                inicio = _primeira_tela(user)
                # Na raiz, mandar 403 seria receber "acesso negado" logo
                # depois de entrar. Cai na primeira tela que a pessoa pode ver.
                if request.url.path == "/" and inicio:
                    return RedirectResponse(inicio, status_code=302)
                atalho = (f"<p style='padding:0 32px'><a href='{inicio}'>Ir para o início</a></p>"
                          if inicio else
                          "<p style='font-family:sans-serif;padding:0 32px'>Nenhuma tela está "
                          "liberada para o seu usuário. <a href='/logout'>Sair</a></p>")
                return HTMLResponse(
                    "<h2 style='font-family:sans-serif;padding:32px'>Tela indisponível</h2>"
                    "<p style='font-family:sans-serif;padding:0 32px'>Seu usuário não tem "
                    "acesso a esta tela. Fale com um administrador.</p>" + atalho,
                    status_code=403,
                )
        return await call_next(request)


def _primeira_tela(user: dict) -> str | None:
    """Primeira tela que este usuário pode abrir, ou None quando não sobrou
    nenhuma.

    Nunca devolve /logout: o redirecionamento da raiz é automático, e
    apontar pra saída significaria deslogar sozinho quem entrou — foi
    exatamente o que aconteceu na primeira versão disto."""
    for chave, _rotulo, destino, _prefixos in permissoes_mod.TELAS:
        if permissoes_mod.tem_permissao(user, chave):
            return destino
    return None


class _InatividadeMiddleware(BaseHTTPMiddleware):
    """Derruba o login parado há tempo demais.

    É o que impede o caso real: o operador sai da estação com o login aberto,
    o próximo senta e bipa no usuário de quem saiu — e o histórico do kit
    passa a responder o nome errado pra sempre.

    O relógio é renovado por QUALQUER uso: navegar aqui e, na bipagem,
    cada código lido (o WebSocket toca o mesmo relógio). Estático fica de
    fora — imagem carregando sozinha não é alguém trabalhando."""

    _IGNORAR = ("/static/", "/login", "/logout", "/ping")

    async def dispatch(self, request: Request, call_next):
        caminho = request.url.path
        if any(caminho.startswith(p) for p in self._IGNORAR):
            return await call_next(request)
        sid = request.session.get("sid", "")
        if sid:
            limite = inatividade_mod.limite_segundos()
            if limite and inatividade_mod.expirou(sid, limite):
                inatividade_mod.esquecer(sid)
                inatividade_mod._limpar_esquecidas(limite)
                request.session.clear()
                # Pedaço de tela pedido por fetch (janela do veículo, prévias)
                # não pode receber a página de login inteira dentro do modal —
                # aí o operador veria um login desenhado dentro da janela.
                if request.headers.get("X-Requested-With") == "fetch":
                    return HTMLResponse(
                        "<div class='alert alert-danger'>Sessão encerrada por "
                        "inatividade. <a href='/login'>Entrar de novo</a></div>",
                        status_code=401)
                return RedirectResponse("/login?expirado=1", status_code=302)
            inatividade_mod.tocar(sid)
        return await call_next(request)


class _CabecalhosSegurancaMiddleware(BaseHTTPMiddleware):
    """Cabeçalhos HTTP de defesa que não têm por que faltar em nenhuma
    resposta, e que não arriscam quebrar nada do sistema atual.

    De propósito NÃO inclui Content-Security-Policy nem X-XSS-Protection:
    o sistema inteiro usa <style> e <script> inline nas telas (é assim que
    o CSS/JS de cada página específica é carregado hoje), e um CSP que
    bloqueasse inline quebraria a aparência e o comportamento de toda tela
    administrativa. Endurecer isso pediria mover todo CSS/JS inline pra
    arquivo — uma reescrita grande demais pra entrar aqui."""

    async def dispatch(self, request: Request, call_next):
        resposta = await call_next(request)
        resposta.headers["X-Content-Type-Options"] = "nosniff"
        # Impede que a tela do sistema seja carregada dentro de um <iframe>
        # de outro site (clickjacking) — nenhuma tela daqui precisa ser
        # embutida em outro lugar.
        resposta.headers["X-Frame-Options"] = "SAMEORIGIN"
        resposta.headers["Referrer-Policy"] = "strict-origin-when-cross-origin"
        return resposta


class _AuditoriaMiddleware(BaseHTTPMiddleware):
    """Grava toda requisição que altera dados.

    Fica no middleware, e não em cada rota, porque cobertura é o requisito:
    rota criada amanhã já nasce auditada. Roda DEPOIS da resposta e nunca
    propaga erro — auditoria com defeito não pode derrubar a operação.
    """

    _IGNORAR = ("/static/", "/ping")

    async def dispatch(self, request: Request, call_next):
        caminho = request.url.path
        if any(caminho.startswith(p) for p in self._IGNORAR):
            return await call_next(request)

        if request.method not in ("POST", "PUT", "PATCH", "DELETE"):
            # GET não altera dados e logar todos inundaria a tabela — mas
            # uma tentativa NEGADA é exatamente o sinal que interessa quando
            # alguém está sondando o que consegue acessar. Logout é a outra
            # exceção: é GET, mas "quem saiu e quando" é justamente o que se
            # procura numa investigação, então precisa ficar registrado.
            # O usuário tem que ser capturado ANTES do call_next() aqui: a
            # própria rota de logout limpa a sessão durante o call_next, e
            # depois disso get_current_user() já não sabe mais quem era.
            eh_logout = caminho == "/logout"
            usuario_antes = None
            if eh_logout:
                try:
                    usuario_antes = get_current_user(request)
                except Exception:
                    pass
            resposta = await call_next(request)
            if resposta.status_code == 403 or eh_logout:
                try:
                    user = usuario_antes if eh_logout else get_current_user(request)
                    auditoria_mod.registrar(
                        user_id=user["id"] if user else None,
                        user_nome=user["nome"] if user else None,
                        acao="LOGOUT" if eh_logout else "ACESSO NEGADO",
                        metodo=request.method, caminho=caminho,
                        detalhe="", ip=_ip_do_cliente(request),
                        status=resposta.status_code,
                    )
                except Exception:
                    pass
            return resposta

        # O corpo precisa ser lido aqui para virar detalhe do log, mas ler
        # consome o stream — então reinjetamos para a rota receber intacto.
        detalhe = ""
        try:
            corpo = await request.body()

            async def _receive():
                return {"type": "http.request", "body": corpo, "more_body": False}

            request._receive = _receive

            tipo = request.headers.get("content-type", "")
            if corpo and ("form-urlencoded" in tipo or "multipart/form-data" in tipo):
                detalhe = auditoria_mod._resumir_form(await request.form())
                request._receive = _receive   # form() reconsome; restaura
        except Exception:
            detalhe = "<corpo nao capturado>"

        resposta = await call_next(request)

        try:
            user = None
            try:
                user = get_current_user(request)
            except Exception:
                pass
            # No POST /login o usuário só existe depois da resposta; o nome
            # digitado já foi para o detalhe, então o log não fica anônimo.
            acao = auditoria_mod.classificar(caminho)
            # A própria rota de login marca o desfecho em request.state antes
            # de responder — sucesso, senha errada, conta desativada ou
            # bloqueio pelo freio de força bruta viram o mesmo status HTTP
            # (200 ou 302) e, sem isso, ficariam indistinguíveis no log.
            desfecho = getattr(request.state, "auditoria_desfecho", None)
            if desfecho:
                acao = f"{acao}: {desfecho}"
            # Rotas importantes escrevem um resumo pronto pra leitura (ex.:
            # "Remessa Consat 07: alvo alterado de 100 para 110"), depois de
            # já saber o resultado — muito mais claro do que os campos crus
            # do formulário. As rotas que ainda não foram customizadas
            # continuam caindo no resumo do formulário, que ao menos cobre
            # tudo sem exigir que cada rota nova pense nisso.
            detalhe_final = getattr(request.state, "auditoria_detalhe", None) or detalhe
            auditoria_mod.registrar(
                user_id=user["id"] if user else None,
                user_nome=user["nome"] if user else None,
                acao=acao,
                metodo=request.method,
                caminho=caminho,
                detalhe=detalhe_final,
                ip=_ip_do_cliente(request),
                status=resposta.status_code,
            )
        except Exception as e:
            print(f"[AUDITORIA] falha ao gravar {request.method} {caminho}: {e}")

        return resposta


app = FastAPI(title="Conferência de Kits")

# COOKIE_SECURE=1 marca o cookie de sessão como "só por HTTPS". Fica
# desligado por padrão porque o acesso pela LAN é HTTP puro (porta 8080) —
# ligado ali, o navegador simplesmente não manda o cookie e ninguém
# consegue logar. Ligue quando o acesso passar a ser só pelo domínio
# HTTPS (ex: atrás do Cloudflare Tunnel).
_COOKIE_SECURE = os.getenv("COOKIE_SECURE", "").strip() in ("1", "true", "True")

# Planilha grande é lida inteira na memória; sem teto, um upload de 1 GB
# derruba o processo. 25 MB cobre com folga qualquer BOM/planilha real.
MAX_UPLOAD_BYTES = int(os.getenv("MAX_UPLOAD_MB", "25")) * 1024 * 1024


# Freio simples contra upload repetido/em massa — não é anti-fraude fino
# como o do login (não bloqueia progressivamente nem por usuário), só evita
# que um script ou um clique múltiplo encha o processo reprocessando
# planilhas sem parar. Guardado em memória pelo mesmo motivo do login.
_UPLOAD_MAX_TENTATIVAS = 10
_UPLOAD_JANELA_SEG = 5 * 60
_upload_tentativas: dict[str, list[float]] = {}


def _upload_liberado(request: Request) -> bool:
    import time
    chave = f"{_ip_do_cliente(request)}|{request.url.path}"
    agora = time.time()
    tentativas = [t for t in _upload_tentativas.get(chave, []) if agora - t < _UPLOAD_JANELA_SEG]
    liberado = len(tentativas) < _UPLOAD_MAX_TENTATIVAS
    tentativas.append(agora)
    _upload_tentativas[chave] = tentativas
    return liberado


async def _ler_upload(request: Request, arquivo) -> bytes:
    """Lê um upload recusando arquivos acima do teto, com extensão errada,
    ou vindos de quem já mandou upload demais nos últimos minutos."""
    if not _upload_liberado(request):
        raise ValueError(
            "Muitos envios de arquivo em pouco tempo. Espere alguns minutos "
            "e tente de novo."
        )
    nome = (getattr(arquivo, "filename", "") or "").lower()
    if nome and not nome.endswith(".xlsx"):
        raise ValueError("O arquivo precisa ser uma planilha .xlsx.")
    conteudo = await arquivo.read()
    if len(conteudo) > MAX_UPLOAD_BYTES:
        raise ValueError(
            f"Arquivo muito grande ({len(conteudo) // (1024*1024)} MB). "
            f"O limite é {MAX_UPLOAD_BYTES // (1024*1024)} MB."
        )
    return conteudo


def _erro_usuario(e: Exception, generico: str = "Não foi possível concluir esta operação.") -> str:
    """Mensagem seguindo pra tela quando uma operação falha.

    ValueError é o tipo que o resto do código já usa de propósito pra
    mensagens pensadas pro usuário (ex.: "Este é o último administrador
    ativo...") — essas passam direto. Qualquer outro tipo de exceção (erro
    de biblioteca, de arquivo, de banco) pode trazer detalhe técnico interno
    que não deveria chegar na tela; essa vai só pro log do servidor, e o
    usuário recebe uma frase genérica."""
    if isinstance(e, ValueError):
        return str(e)
    print(f"[KIT] erro interno ({type(e).__name__}): {e}")
    return generico


def _veiculo_do_kit(kit_id: str) -> str:
    """Número do veículo de um kit, pro log de auditoria falar a língua do
    operador ("veículo 293") em vez do kit_id (um UUID que ninguém decora)."""
    with db() as conn:
        row = conn.execute(
            "SELECT veiculo FROM kit_record WHERE kit_id = ?", (kit_id,)
        ).fetchone()
    return (row["veiculo"] if row and row["veiculo"] else kit_id[:8])


_SECRET_KEY = os.getenv("SECRET_KEY", "").strip()
# O valor de exemplo do .env.example — que fica no repositório. Se o .env de
# verdade ainda tiver ESTE texto, qualquer pessoa que já tenha visto o
# repositório conhece a chave que assina o cookie de sessão e pode forjar um
# login como qualquer usuário, sem senha nenhuma.
_SECRET_KEY_PLACEHOLDER = "troque-por-chave-aleatoria-longa"
if not _SECRET_KEY:
    # Sem SECRET_KEY, qualquer um forja o cookie de sessão e entra como
    # quem quiser. Numa máquina exposta isso é crítico, então o processo
    # se recusa a subir; em uso local o aviso é gritante mas não trava.
    if _COOKIE_SECURE or os.getenv("SERVIDOR_URL", "").startswith("https://"):
        raise RuntimeError(
            "SECRET_KEY nao definido no .env. Como este servidor esta configurado "
            "para acesso externo, subir com a chave padrao permitiria a qualquer "
            "pessoa forjar uma sessao. Defina SECRET_KEY antes de iniciar."
        )
    _SECRET_KEY = "dev-secret"
    print("[KIT] AVISO: SECRET_KEY nao definido — usando chave de desenvolvimento. "
          "NAO exponha este servidor sem definir SECRET_KEY no .env.")
elif _SECRET_KEY == _SECRET_KEY_PLACEHOLDER:
    # De propósito NUNCA travamos por causa disto (ao contrário do caso
    # "vazio" acima): travar aqui poderia derrubar um servidor já no ar sem
    # ninguém saber trocar o .env na hora. O aviso é pra ser impossível de
    # não ver nos logs de inicialização.
    print("=" * 78)
    print("[KIT] ALERTA DE SEGURANCA: SECRET_KEY ainda esta com o valor de exemplo")
    print("      do .env.example, que fica no repositorio. Qualquer pessoa que ja")
    print("      tenha visto esse arquivo pode forjar um cookie de sessao e entrar")
    print("      como qualquer usuario, inclusive admin. Troque AGORA por uma chave")
    print("      aleatoria de verdade, por exemplo:")
    print('      python -c "import secrets; print(secrets.token_hex(32))"')
    print("      Cole o resultado em SECRET_KEY no .env e reinicie o servidor.")
    print("=" * 78)

# Ordem importa: quem é adicionado por último fica por fora. A auditoria
# precisa enxergar a sessão, então entra ANTES do SessionMiddleware para
# ficar por dentro dele. O porteiro das telas entra antes da auditoria pra
# ficar por dentro dela — assim o 403 dele também é registrado.
app.add_middleware(_PaginaPermitidaMiddleware)
app.add_middleware(_AuditoriaMiddleware)
# Depois da auditoria (fica por fora dela) e antes do SessionMiddleware (fica
# por dentro dele, então enxerga a sessão): sessão expirada tem que cair antes
# de a rota rodar, não depois.
app.add_middleware(_InatividadeMiddleware)
app.add_middleware(
    SessionMiddleware,
    secret_key=_SECRET_KEY,
    same_site="lax",
    https_only=_COOKIE_SECURE,
    max_age=12 * 60 * 60,   # 12h — uma jornada; antes eram 14 dias
)
app.add_middleware(_MobileGateMiddleware)
# Por FORA de todo o resto: comprime a resposta já pronta. As telas de lista
# são HTML repetitivo (tabela com milhares de linhas), que encolhe ~10x — a
# Produção sozinha saía com 2,4 MB por carregamento. Numa rede de galpão, com
# vários tablets pedindo a mesma tela, é a diferença entre a página "abrir" e
# "ir chegando". minimum_size evita gastar CPU comprimindo resposta curta.
# compresslevel=6 e nao o 9 que o Starlette usa por padrao: medido aqui,
# o 9 leva o DOBRO do tempo pra deixar a resposta 1,5% menor.
app.add_middleware(GZipMiddleware, minimum_size=1024, compresslevel=6)
# Por FORA de tudo: cabeçalho de segurança tem que voltar em toda resposta,
# inclusive página de erro e redirecionamento — nenhuma lógica de negócio
# depende dele, então não tem por que ficar por dentro de nada.
app.add_middleware(_CabecalhosSegurancaMiddleware)


class _EstaticoComCache(StaticFiles):
    """Diz ao navegador pra guardar CSS/JS/imagem por um ano.

    Só é seguro porque todo estático é pedido com ?v=NN no endereço: mudou o
    arquivo, muda a versão, muda o endereço — o navegador busca de novo
    sozinho. Sem isso, cada troca de tela repetia o download do CSS e do
    zxing.min.js (o leitor por câmera, que é grande).

    É subclasse do StaticFiles, e não um middleware: middleware roda em TODA
    requisição do sistema pra ajudar só as de /static.
    """

    _UM_ANO = 60 * 60 * 24 * 365

    def file_response(self, *args, **kwargs):
        resposta = super().file_response(*args, **kwargs)
        resposta.headers["Cache-Control"] = f"public, max-age={self._UM_ANO}, immutable"
        return resposta


app.mount("/static", _EstaticoComCache(directory="static"), name="static")
jinja = Jinja2Templates(directory="templates")


def _detectar_ip_lan() -> str:
    """Detecta o IP da máquina na LAN local (não localhost)."""
    import socket
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(("8.8.8.8", 80))   # rota padrão — funciona em qualquer rede LAN
        ip = s.getsockname()[0]
        s.close()
        return ip
    except Exception:
        return "127.0.0.1"


@app.on_event("startup")
def startup():
    init_db()
    import app.zpl as _zpl
    _zpl.EMPRESA_NOME = os.getenv("EMPRESA_NOME", "Sua Empresa")

    ip = _detectar_ip_lan()
    _tem_ssl = os.path.exists("certs/cert.pem") and os.path.exists("certs/key.pem")

    app.state.url_http  = f"http://{ip}:8080"
    app.state.url_https = f"https://{ip}:8011" if _tem_ssl else None
    app.state.tem_ssl = _tem_ssl

    if _tem_ssl:
        url_local = f"https://{ip}:8011"
    else:
        url_local = f"http://{ip}:8080"

    # SERVIDOR_URL do .env manda: é o endereço que vai no QR da etiqueta.
    # Sem ele, cai no IP da LAN (funciona só dentro do galpão). Com um
    # domínio público (ex: atrás do Cloudflare Tunnel), a etiqueta impressa
    # abre de qualquer lugar — por isso o valor configurado nunca é
    # sobrescrito pela detecção automática.
    url_publica = (os.getenv("SERVIDOR_URL") or "").strip().rstrip("/")

    _zpl.SERVIDOR_URL = url_publica or url_local
    app.state.servidor_url = _zpl.SERVIDOR_URL

    if url_publica:
        print(f"[KIT] Endereco publico (QR das etiquetas): {url_publica}")
        print(f"[KIT] Acesso local: {url_local}")
    elif _tem_ssl:
        print(f"[KIT] HTTPS (QR + Admin): {url_local}")
        print(f"[KIT] HTTP  (alternativo): {app.state.url_http}")
    else:
        print(f"[KIT] HTTP: {url_local}")


# ── Backup automático ────────────────────────────────────────────────────
# Um laço próprio, dentro do servidor, em vez de tarefa agendada do Windows:
# o sistema já roda como serviço nesta máquina, e uma tarefa a mais seria
# outra coisa pra configurar (e pra alguém desligar sem querer).
_BACKUP_CHECAGEM_SEG = 600       # olha de 10 em 10 minutos
_BACKUP_ESPERA_INICIAL_SEG = 60  # deixa o sistema subir antes da primeira


async def _laco_backup():
    import asyncio
    await asyncio.sleep(_BACKUP_ESPERA_INICIAL_SEG)
    while True:
        try:
            if backup_mod.precisa_agora():
                # to_thread porque a cópia é bloqueante: no laço direto, ela
                # seguraria a bipagem de todo mundo enquanto durasse.
                r = await asyncio.to_thread(backup_mod.criar_backup, "automatico")
                print(f"[KIT] Backup automatico: {r['caminho']}"
                      + (f" — {r['erro']}" if r["erro"] else ""))
        except Exception as e:
            # Backup que falha não pode derrubar o sistema nem parar o laço:
            # a próxima volta tenta de novo, e a tela mostra a data velha.
            print(f"[KIT] Backup automatico falhou: {e}")
        await asyncio.sleep(_BACKUP_CHECAGEM_SEG)


@app.on_event("startup")
async def startup_backup():
    import asyncio
    app.state.tarefa_backup = asyncio.create_task(_laco_backup())


def _parse_itens_form(form) -> list[dict]:
    """Extrai itens do formulário de template sem depender de índices sequenciais."""
    indices = sorted(
        int(m.group(1))
        for k in form.keys()
        for m in [re.match(r'^item_tipo_id_(\d+)$', k)]
        if m
    )
    itens = []
    for i in indices:
        tipo_id = form.get(f"item_tipo_id_{i}", "").strip()
        if not tipo_id:
            continue
        itens.append({
            "item_tipo_id": int(tipo_id),
            "quantidade_exigida": max(1, int(form.get(f"qtd_{i}", 1) or 1)),
            "obrigatorio": bool(form.get(f"obrigatorio_{i}")),
            "componente_codigo": (form.get(f"componente_codigo_{i}", "") or "").strip() or None,
            "requer_serial": bool(form.get(f"requer_serial_{i}")),
        })
    return itens


def _voltar_para(request: Request, padrao: str) -> str:
    """Para onde o botão "← Voltar" de uma tela de detalhe deve apontar.

    Usa o Referer quando ele é uma tela DESTE sistema: assim quem chegou de
    uma busca ou da página 3 volta exatamente pra lá, com filtro e paginação
    preservados, em vez de cair na lista pelada e ter que refazer tudo.
    Fora isso (link colado, aba nova, referer ausente) cai no padrão.

    Só aceita caminho relativo do próprio site — nunca uma URL externa —
    pra o botão não virar um pulo pra fora do sistema."""
    ref = request.headers.get("referer") or ""
    if not ref:
        return padrao
    from urllib.parse import urlparse
    p = urlparse(ref)
    if p.netloc and p.netloc != request.url.netloc:
        return padrao
    destino = p.path + (("?" + p.query) if p.query else "")
    # Voltar pra própria tela seria um botão que não faz nada.
    if not destino.startswith("/") or destino.startswith(request.url.path):
        return padrao
    return destino


def render(request: Request, template: str, ctx: dict = {}):
    user = get_current_user(request)
    # A faixa de aviso obedece à configuração (quais itens, em quais telas,
    # quantos cabem) — por isso quem decide é o estoque, não o template.
    banner = (estoque_mod.alertas_para_banner(request.url.path) if user
              else {"itens": [], "total": 0, "cfg": estoque_mod.ALERTA_PADRAO})
    alertas_estoque = banner["itens"]
    pode = (lambda chave: permissoes_mod.tem_permissao(user, chave)) if user else (lambda chave: False)
    # O menu é montado a partir do mesmo cadastro de telas que o porteiro usa:
    # esconder o link e barrar a rota não podem discordar.
    return jinja.TemplateResponse(template, {"request": request, "user": user,
                                             "alertas_estoque": alertas_estoque,
                                             "alertas_total": banner["total"],
                                             "alerta_cfg": banner["cfg"],
                                             "pode": pode, "telas": permissoes_mod.TELAS,
                                             # Cor por cliente, à mão de qualquer tela: a
                                             # mesma frota tem que ter a mesma cor na lista,
                                             # na janela e em qualquer lugar novo.
                                             "cor_do_cliente": clientes_mod.cor_do_cliente,
                                             **ctx})


# ── Auth ──────────────────────────────────────────────────────────────────────

# Freio de força bruta. Guardado em memória de propósito: o app roda em um
# processo só, e perder a contagem num restart é aceitável — o objetivo é
# tornar inviável varrer senhas, não ser um cofre distribuído.
_LOGIN_MAX_TENTATIVAS = 8
_LOGIN_JANELA_SEG = 15 * 60
# Teto do bloqueio progressivo — nunca passa de 4h, mesmo pra quem insiste
# indefinidamente. Sem teto, um IP com histórico antigo de tentativas
# ficaria bloqueado por dias, o que vira problema pra usuário legítimo
# atrás do mesmo IP (rede corporativa, Cloudflare) tanto quanto pro atacante.
_LOGIN_BLOQUEIO_TETO_SEG = 4 * 60 * 60
_login_tentativas: dict[str, list[float]] = {}
_login_bloqueado_ate: dict[str, float] = {}
_login_nivel_bloqueio: dict[str, int] = {}


def _ip_do_cliente(request: Request) -> str:
    """IP real de quem chamou.

    Atrás de um proxy (Cloudflare Tunnel), request.client.host é o IP do
    proxy — igual para todo mundo — o que faria o freio de login trancar
    todos os usuários de uma vez. Nesse caso o IP verdadeiro vem no
    cabeçalho CF-Connecting-IP.

    Só confiamos no cabeçalho quando TRUST_PROXY_IP=1, porque quem fala
    direto com o app (acesso pela LAN) pode forjar esse cabeçalho e
    escapar do limite trocando o valor a cada tentativa.
    """
    if os.getenv("TRUST_PROXY_IP", "").strip() in ("1", "true", "True"):
        cf = request.headers.get("cf-connecting-ip")
        if cf:
            return cf.strip()
        xff = request.headers.get("x-forwarded-for")
        if xff:
            return xff.split(",")[0].strip()
    return request.client.host if request.client else "?"


def _login_chave(request: Request, username: str) -> str:
    return f"{_ip_do_cliente(request)}|{username.lower()}"


def _login_bloqueado(chave: str) -> int:
    """Segundos restantes de bloqueio, ou 0 se liberado."""
    import time
    agora = time.time()
    ate = _login_bloqueado_ate.get(chave, 0)
    if agora < ate:
        return int(ate - agora)
    return 0


def _login_falhou(chave: str) -> None:
    """Cada ciclo de _LOGIN_MAX_TENTATIVAS erradas dobra o bloqueio do ciclo
    anterior (15min, 30min, 1h, 2h... até o teto de 4h) — quem insiste depois
    de ser liberado espera cada vez mais, em vez de sempre os mesmos 15min."""
    import time
    agora = time.time()
    tentativas = [t for t in _login_tentativas.get(chave, []) if agora - t < _LOGIN_JANELA_SEG]
    tentativas.append(agora)
    if len(tentativas) >= _LOGIN_MAX_TENTATIVAS:
        nivel = _login_nivel_bloqueio.get(chave, 0) + 1
        _login_nivel_bloqueio[chave] = nivel
        duracao = min(_LOGIN_JANELA_SEG * (2 ** (nivel - 1)), _LOGIN_BLOQUEIO_TETO_SEG)
        _login_bloqueado_ate[chave] = agora + duracao
        tentativas = []
    _login_tentativas[chave] = tentativas


def _login_ok(chave: str) -> None:
    _login_tentativas.pop(chave, None)
    _login_bloqueado_ate.pop(chave, None)
    _login_nivel_bloqueio.pop(chave, None)


@app.get("/login", response_class=HTMLResponse)
async def login_page(request: Request):
    if get_current_user(request):
        next_url = request.query_params.get("next", "/")
        return RedirectResponse(next_url if next_url.startswith("/") and not next_url.startswith("//") else "/", status_code=302)
    return render(request, "login.html", {"next": request.query_params.get("next", "")})


@app.post("/login")
async def login_post(request: Request):
    form = await request.form()
    username = str(form.get("username", "")).strip()
    password = str(form.get("password", "")).strip()
    next_url = str(form.get("next", "")).strip()
    if not (next_url.startswith("/") and not next_url.startswith("//")):
        next_url = "/"

    chave = _login_chave(request, username)
    espera = _login_bloqueado(chave)
    if espera:
        minutos = max(1, espera // 60)
        request.state.auditoria_desfecho = "bloqueado"
        return render(request, "login.html", {
            "erro": f"Muitas tentativas. Tente novamente em {minutos} minuto(s).",
            "next": next_url,
        })

    with db() as conn:
        row = conn.execute(
            "SELECT * FROM users WHERE username = ?", (username,)
        ).fetchone()
    if row and not row["ativo"]:
        request.state.auditoria_desfecho = "usuario desativado"
        return render(request, "login.html", {
            "erro": "Este usuário está desativado. Procure um administrador.",
            "next": next_url,
        })

    if row and verify_password(password, row["password_hash"]):
        _login_ok(chave)
        request.state.auditoria_desfecho = "sucesso"
        # Descarta qualquer conteúdo de sessão anterior antes de autenticar,
        # para que um valor plantado na sessão pré-login não sobreviva à
        # troca de identidade (fixação de sessão).
        request.session.clear()
        request.session["user_id"] = row["id"]
        # sid = "esta aba, a partir de agora". É por ele que o relógio de
        # inatividade sabe de quem está contando o tempo.
        import uuid as _uuid
        request.session["sid"] = _uuid.uuid4().hex
        inatividade_mod.tocar(request.session["sid"])
        return RedirectResponse(next_url, status_code=302)

    _login_falhou(chave)
    request.state.auditoria_desfecho = "falha"
    return render(request, "login.html", {"erro": "Usuário ou senha incorretos.", "next": next_url})


@app.get("/logout")
async def logout(request: Request):
    inatividade_mod.esquecer(request.session.get("sid", ""))
    request.session.clear()
    return RedirectResponse("/login", status_code=302)


# ── Usuários (só admin) ───────────────────────────────────────────────────────

@app.get("/admin/usuarios", response_class=HTMLResponse)
@require_admin
async def admin_usuarios(request: Request, pagina: int = 1, busca: str = ""):
    usuarios = usuarios_mod.listar()
    negadas_por_usuario = {u["id"]: permissoes_mod.negadas_do_usuario(u["id"]) for u in usuarios}
    usuarios = paginacao_mod.filtrar(usuarios, busca, ("nome", "username"))
    return render(request, "admin_usuarios.html", {
        "pag_usuarios": paginacao_mod.paginar(usuarios, pagina),
        "busca": busca,
        "permissoes": permissoes_mod.PERMISSOES,
        "grupos": permissoes_mod.GRUPOS,
        "negadas_por_usuario": negadas_por_usuario,
        "inatividade": inatividade_mod.get_config(),
        "inatividade_max": inatividade_mod.MINUTOS_MAX,
    })


@app.get("/admin/padrao-visual", response_class=HTMLResponse)
@require_admin
async def admin_padrao_visual(request: Request):
    return render(request, "admin_padrao_visual.html", {})


@app.post("/admin/usuarios/inatividade")
@require_admin
async def admin_usuarios_inatividade(request: Request):
    """Quanto tempo parado até o login cair. É admin-only de propósito: mexe
    em quem consegue ficar dentro do sistema."""
    form = await request.form()
    minutos = inatividade_mod.salvar_config(form.get("minutos", "0"))
    request.state.auditoria_detalhe = (
        f"Tempo de inatividade definido para {minutos} minuto(s)" if minutos
        else "Expiração por inatividade desligada")
    return RedirectResponse(f"/admin/usuarios?ok=inatividade&m={minutos}", status_code=302)


@app.post("/admin/usuarios")
@require_admin
async def admin_usuarios_criar(request: Request):
    form = await request.form()
    nome = str(form.get("nome", ""))
    username = str(form.get("username", ""))
    eh_admin = bool(form.get("admin"))
    try:
        usuarios_mod.criar(nome=nome, username=username,
                           senha=str(form.get("senha", "")), admin=eh_admin)
    except ValueError as e:
        return RedirectResponse("/admin/usuarios?erro=" + quote(str(e)), status_code=302)
    request.state.auditoria_detalhe = (
        f"Usuário \"{nome}\" ({username}) criado"
        + (" como administrador" if eh_admin else ""))
    return RedirectResponse("/admin/usuarios?ok=criado", status_code=302)


@app.post("/admin/usuarios/{user_id}/nome")
@require_admin
async def admin_usuario_renomear(request: Request, user_id: int):
    """Corrige o nome de exibição. O login fica como foi cadastrado — é ele
    que identifica a pessoa no histórico e é o que ela digita pra entrar."""
    alvo = usuarios_mod.buscar(user_id)
    if not alvo:
        raise HTTPException(status_code=404)
    form = await request.form()
    novo_nome = str(form.get("nome", ""))
    try:
        usuarios_mod.renomear(user_id, novo_nome)
    except ValueError as e:
        return RedirectResponse("/admin/usuarios?erro=" + quote(str(e)), status_code=302)
    request.state.auditoria_detalhe = (
        f"Usuário \"{alvo['nome']}\" ({alvo['username']}) renomeado para \"{novo_nome}\"")
    return RedirectResponse("/admin/usuarios?ok=nome", status_code=302)


@app.post("/admin/usuarios/{user_id}/admin")
@require_admin
async def admin_usuario_toggle_admin(request: Request, user_id: int):
    alvo = usuarios_mod.buscar(user_id)
    if not alvo:
        raise HTTPException(status_code=404)
    novo_admin = not alvo["admin"]
    try:
        usuarios_mod.definir_admin(user_id, novo_admin)
    except ValueError as e:
        return RedirectResponse("/admin/usuarios?erro=" + quote(str(e)), status_code=302)
    request.state.auditoria_detalhe = (
        f"Usuário \"{alvo['nome']}\" ({alvo['username']}) "
        + ("promovido a administrador" if novo_admin else "rebaixado a usuário comum"))
    return RedirectResponse("/admin/usuarios?ok=perfil", status_code=302)


@app.post("/admin/usuarios/{user_id}/ativo")
@require_admin
async def admin_usuario_toggle_ativo(request: Request, user_id: int):
    alvo = usuarios_mod.buscar(user_id)
    if not alvo:
        raise HTTPException(status_code=404)
    novo_ativo = not alvo["ativo"]
    try:
        usuarios_mod.definir_ativo(user_id, novo_ativo)
    except ValueError as e:
        return RedirectResponse("/admin/usuarios?erro=" + quote(str(e)), status_code=302)
    request.state.auditoria_detalhe = (
        f"Usuário \"{alvo['nome']}\" ({alvo['username']}) "
        + ("reativado" if novo_ativo else "desativado"))
    return RedirectResponse("/admin/usuarios?ok=status", status_code=302)


@app.post("/admin/usuarios/{user_id}/permissoes")
@require_admin
async def admin_usuario_permissoes(request: Request, user_id: int):
    alvo = usuarios_mod.buscar(user_id)
    if not alvo:
        raise HTTPException(status_code=404)
    form = await request.form()
    antes = permissoes_mod.negadas_do_usuario(user_id)
    permitidas = {chave for chave in permissoes_mod.PERMISSOES if form.get(chave)}
    permissoes_mod.definir_permissoes(user_id, permitidas)
    depois = permissoes_mod.negadas_do_usuario(user_id)
    # O que a pessoa PERDEU (passou a negado) e o que RECUPEROU (deixou de
    # estar negado) — muito mais direto de revisar do que os ~15 checkboxes
    # crus que o formulário manda.
    passou_a_negar = sorted(permissoes_mod.PERMISSOES.get(c, c) for c in (depois - antes))
    voltou_a_permitir = sorted(permissoes_mod.PERMISSOES.get(c, c) for c in (antes - depois))
    partes = []
    if passou_a_negar:
        partes.append("negou " + ", ".join(passou_a_negar))
    if voltou_a_permitir:
        partes.append("liberou " + ", ".join(voltou_a_permitir))
    resumo = "; ".join(partes) if partes else "sem mudança"
    request.state.auditoria_detalhe = (
        f"Permissões de \"{alvo['nome']}\" ({alvo['username']}): {resumo}")
    return RedirectResponse("/admin/usuarios?ok=permissoes", status_code=302)


@app.post("/admin/usuarios/{user_id}/senha")
@require_admin
async def admin_usuario_senha(request: Request, user_id: int):
    alvo = usuarios_mod.buscar(user_id)
    if not alvo:
        raise HTTPException(status_code=404)
    form = await request.form()
    try:
        usuarios_mod.trocar_senha(user_id, str(form.get("senha", "")))
    except ValueError as e:
        return RedirectResponse("/admin/usuarios?erro=" + quote(str(e)), status_code=302)
    # Nunca a senha em si — só quem teve a senha trocada e por quem (o "por
    # quem" já vem de graça do user_nome da linha da auditoria).
    request.state.auditoria_detalhe = f"Senha de \"{alvo['nome']}\" ({alvo['username']}) alterada"
    return RedirectResponse("/admin/usuarios?ok=senha", status_code=302)


# ── Backup do banco ───────────────────────────────────────────────────────────

@app.get("/admin/backup", response_class=HTMLResponse)
@require_login
async def admin_backup(request: Request):
    cfg = backup_mod.get_config()
    previsto = backup_mod.proximo_previsto()
    return render(request, "admin_backup.html", {
        "cfg": cfg,
        "copias": backup_mod.listar(),
        "ultimo": backup_mod.ultimo(),
        "pasta_padrao": backup_mod.pasta_padrao(),
        "previsto": previsto.strftime("%d/%m/%Y %H:%M") if previsto else "",
        "atrasado": backup_mod.precisa_agora(),
        "intervalo_min": backup_mod.INTERVALO_MIN,
        "intervalo_max": backup_mod.INTERVALO_MAX,
        "manter_max": backup_mod.MANTER_MAX,
    })


@app.post("/admin/backup/config")
@require_permission("backup_configurar")
async def admin_backup_config(request: Request):
    form = await request.form()
    try:
        backup_mod.salvar_config({
            "ativo": "1" if form.get("ativo") else "0",
            "intervalo_horas": form.get("intervalo_horas", "24"),
            "manter": form.get("manter", "30"),
            "pasta_extra": form.get("pasta_extra", ""),
        })
    except ValueError as e:
        return RedirectResponse("/admin/backup?erro=" + quote(str(e)), status_code=302)
    return RedirectResponse("/admin/backup?ok=config", status_code=302)


@app.post("/admin/backup/agora")
@require_permission("backup_configurar")
async def admin_backup_agora(request: Request):
    """Backup na hora — é o que prova que a configuração está de pé (a pasta
    extra grava mesmo?) sem esperar o intervalo virar."""
    import asyncio
    user = get_current_user(request)
    try:
        r = await asyncio.to_thread(backup_mod.criar_backup, "manual",
                                    user["id"] if user else None)
    except Exception as e:
        return RedirectResponse("/admin/backup?erro=" + quote(_erro_usuario(e)), status_code=302)
    if r["erro"]:
        return RedirectResponse("/admin/backup?erro=" + quote(r["erro"]), status_code=302)
    return RedirectResponse("/admin/backup?ok=feito&n=" + quote(r["arquivo"]), status_code=302)


# ── Auditoria (só admin) ──────────────────────────────────────────────────────

@app.get("/admin/auditoria", response_class=HTMLResponse)
@require_admin
async def admin_auditoria(request: Request,
                          data_ini: str = "", data_fim: str = "", pagina: int = 1,
                          busca: str = "",
                          user_id: list[str] = Query(default=[]),
                          acao: list[str] = Query(default=[])):
    # Paginação no SQL, com o total vindo de contar(): a tela pede só a
    # página que vai mostrar. Antes pedia 2000 linhas e paginava esse
    # pedaço — passando disso, os registros mais antigos do período não
    # apareciam em página nenhuma e nada na tela dizia que faltava algo.
    por_pagina = paginacao_mod.POR_PAGINA_PADRAO
    total = auditoria_mod.contar(data_ini, data_fim, user_id, acao, busca=busca)
    total_paginas = max(1, -(-total // por_pagina))
    pagina = max(1, min(pagina, total_paginas))
    offset = (pagina - 1) * por_pagina
    registros = auditoria_mod.listar(data_ini, data_fim, user_id, acao,
                                     limite=por_pagina, offset=offset, busca=busca)
    return render(request, "admin_auditoria.html", {
        "pag_registros": {
            "itens": registros, "pagina": pagina, "total_paginas": total_paginas,
            "total": total, "inicio": offset + 1 if registros else 0,
            "fim": offset + len(registros), "exibindo": len(registros),
            "paginas_visiveis": paginacao_mod.janela_paginas(pagina, total_paginas),
        },
        "busca": busca,
        "usuarios": usuarios_mod.listar(),
        # Pares (valor, rótulo) pro filtro de múltipla escolha — montados aqui
        # porque o Jinja não tem zip.
        "opcoes_usuarios": [(str(u["id"]), u["nome"]) for u in usuarios_mod.listar()],
        "acoes": auditoria_mod.acoes_distintas(),
        "data_ini": data_ini, "data_fim": data_fim,
        "filtro_user_id": user_id, "filtro_acao": acao,
    })


_DIAS_SEMANA = ("segunda", "terça", "quarta", "quinta", "sexta", "sábado", "domingo")


def _dia_da_semana(data_iso: str) -> str:
    """'2026-08-21' → 'sexta'. Coluna própria na exportação pra dar pra
    cruzar volume por dia útil sem fórmula no Excel."""
    try:
        return _DIAS_SEMANA[datetime.strptime(data_iso[:10], "%Y-%m-%d").weekday()]
    except (ValueError, IndexError):
        return ""


@app.get("/admin/auditoria/exportar.xlsx")
@require_admin
async def admin_auditoria_exportar(request: Request,
                                   data_ini: str = "", data_fim: str = "",
                                   busca: str = "",
                                   user_id: list[str] = Query(default=[]),
                                   acao: list[str] = Query(default=[])):
    """Log completo em Excel, com abas de análise já prontas.

    Usa os MESMOS filtros da tela (mesma função _filtros por baixo), então o
    que se exporta é exatamente o que se está vendo — sem teto escondido: o
    período escolhido é que limita o volume.

    Os resumos vêm agregados do banco, não somados em Python sobre as linhas
    exportadas: num período grande isso é a diferença entre gerar o arquivo e
    estourar a memória."""
    from fastapi.responses import Response as _Resp
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from io import BytesIO

    total = auditoria_mod.contar(data_ini, data_fim, user_id, acao, busca=busca)
    registros = auditoria_mod.listar(data_ini, data_fim, user_id, acao,
                                     limite=total or 1, busca=busca)
    resumos = auditoria_mod.resumos_para_analise(data_ini, data_fim, user_id, acao, busca)

    wb = openpyxl.Workbook()
    azul, branco, cinza = "1A3A5C", "FFFFFF", "F4F7FB"

    def montar(ws, colunas, larguras, linhas, zebra=True):
        for col, (h, w) in enumerate(zip(colunas, larguras), 1):
            c = ws.cell(1, col, h)
            c.font = Font(bold=True, color=branco)
            c.fill = PatternFill("solid", fgColor=azul)
            c.alignment = Alignment(horizontal="center", vertical="center")
            ws.column_dimensions[get_column_letter(col)].width = w
        for i, linha in enumerate(linhas):
            r = i + 2
            for col, valor in enumerate(linha, 1):
                ws.cell(r, col, valor)
            if zebra and i % 2 == 0:
                for col in range(1, len(colunas) + 1):
                    ws.cell(r, col).fill = PatternFill("solid", fgColor=cinza)
        ws.freeze_panes = "A2"
        if linhas:
            ws.auto_filter.ref = f"A1:{get_column_letter(len(colunas))}{len(linhas) + 1}"

    # ── Aba 1: o log linha a linha ──────────────────────────────────────────
    ws = wb.active
    ws.title = "Log completo"
    montar(ws,
           ["Data", "Hora", "Dia da semana", "Usuário", "Ação", "Método",
            "Caminho", "Status", "IP", "Detalhe"],
           [12, 10, 15, 22, 24, 9, 46, 9, 16, 80],
           [(
               (r["criado_em"] or "")[:10],
               (r["criado_em"] or "")[11:19],
               _dia_da_semana(r["criado_em"] or ""),
               r["user_nome"] or "(sem usuário)",
               r["acao"] or "",
               r["metodo"] or "",
               r["caminho"] or "",
               r["status"] if r["status"] is not None else "",
               r["ip"] or "",
               (r["detalhe"] or "")[:2000],
           ) for r in registros])

    montar(wb.create_sheet("Por dia"), ["Dia", "Ações"], [16, 12],
           [(x["dia"], x["total"]) for x in resumos["por_dia"]])
    montar(wb.create_sheet("Por usuário"), ["Usuário", "Ações"], [30, 12],
           [(x["usuario"], x["total"]) for x in resumos["por_usuario"]])
    montar(wb.create_sheet("Por ação"), ["Ação", "Ocorrências"], [34, 14],
           [(x["acao"], x["total"]) for x in resumos["por_acao"]])
    montar(wb.create_sheet("Dia x Usuário"), ["Dia", "Usuário", "Ações"], [16, 30, 12],
           [(x["dia"], x["usuario"], x["total"]) for x in resumos["por_dia_usuario"]])
    montar(wb.create_sheet("Por hora"), ["Hora", "Ações"], [12, 12],
           [(x["hora"] + "h", x["total"]) for x in resumos["por_hora"]])
    montar(wb.create_sheet("Por status HTTP"), ["Status", "Ocorrências"], [12, 14],
           [(x["status"], x["total"]) for x in resumos["por_status"]])

    # ── Aba de contexto: o que este arquivo é ───────────────────────────────
    ws_i = wb.create_sheet("Filtros aplicados", 0)
    ws_i.column_dimensions["A"].width = 26
    ws_i.column_dimensions["B"].width = 56
    linhas_info = [
        ("Exportação da auditoria", ""),
        ("Gerado em", now_brt()),
        ("Período", f"{data_ini or 'início'} até {data_fim or 'hoje'}"),
        # Os filtros são listas (múltipla escolha): a capa mostra todos os
        # escolhidos, não só o primeiro.
        ("Usuário filtrado", ", ".join(
            u["nome"] for u in usuarios_mod.listar()
            if str(u["id"]) in {str(x) for x in user_id}) or "todos"),
        ("Ação filtrada", ", ".join(acao) or "todas"),
        ("Busca por texto", busca or "(nenhuma)"),
        ("Registros exportados", total),
        ("", ""),
        ("Observação", "As abas de resumo são calculadas sobre o MESMO filtro "
                       "desta exportação."),
    ]
    for i, (k, v) in enumerate(linhas_info, 1):
        a = ws_i.cell(i, 1, k)
        ws_i.cell(i, 2, v)
        if i == 1:
            a.font = Font(bold=True, size=13, color=branco)
            a.fill = PatternFill("solid", fgColor=azul)
        else:
            a.font = Font(bold=True)

    buf = BytesIO()
    wb.save(buf); buf.seek(0)
    nome = f"auditoria_{(data_ini or 'tudo')}_{(data_fim or now_brt()[:10])}.xlsx"
    return _Resp(content=buf.read(),
                 media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                 headers={"Content-Disposition": f'attachment; filename="{nome}"'})


# ── Rede ──────────────────────────────────────────────────────────────────────

@app.get("/funcionalidades", response_class=HTMLResponse)
@require_login
async def funcionalidades(request: Request):
    """Manual do sistema tela a tela. As permissões e os status de compra vêm
    do próprio código — assim a página não descreve uma lista que já mudou."""
    return render(request, "funcionalidades.html", {
        "permissoes": permissoes_mod.PERMISSOES,
        "grupos": permissoes_mod.GRUPOS,
        "status_compra_opcoes": estoque_mod.STATUS_COMPRA,
    })


@app.get("/rede", response_class=HTMLResponse)
@require_permission("ver_rede")
async def rede(request: Request):
    import app.zpl as _zpl
    url_http  = getattr(app.state, "url_http",  _zpl.SERVIDOR_URL)
    url_https = getattr(app.state, "url_https", None)
    tem_ssl   = getattr(app.state, "tem_ssl",   False)

    def _make_qr_svg(url: str) -> str:
        try:
            import segno, io as _io, re
            qr = segno.make(url, error="q")
            buf = _io.BytesIO()
            qr.save(buf, kind="svg", scale=5, border=2, xmldecl=False, nl=False)
            svg = buf.getvalue().decode("utf-8")
            svg = re.sub(r'\s(width|height)="[^"]*"', '', svg, count=2)
            svg = svg.replace("<svg ", '<svg style="display:block;width:100%;max-width:200px;height:auto;margin:0 auto;" ', 1)
            return svg
        except Exception:
            return ""

    qr_ios     = _make_qr_svg(url_https) if url_https else _make_qr_svg(url_http)
    qr_android = _make_qr_svg(url_http)

    return render(request, "rede.html", {
        "url_http":    url_http,
        "url_https":   url_https,
        "servidor_url": url_https or url_http,
        "qr_ios":      qr_ios,
        "qr_android":  qr_android,
        "tem_ssl":     tem_ssl,
    })


# ── Certificado SSL (para iOS instalar) ──────────────────────────────────────

@app.get("/cert")
async def baixar_cert():
    """Download do certificado SSL para instalar no iOS/Android."""
    from fastapi.responses import Response as _Resp
    cert_path = "certs/cert.pem"
    if not os.path.exists(cert_path):
        return PlainTextResponse("Certificado não encontrado. Execute: python gerar_cert.py", status_code=404)
    with open(cert_path, "rb") as f:
        cert_bytes = f.read()
    return _Resp(
        content=cert_bytes,
        media_type="application/x-x509-ca-cert",
        headers={"Content-Disposition": 'attachment; filename="KitConference.crt"'},
    )


# ── Ping público (sem login) ─────────────────────────────────────────────────

@app.get("/ping")
async def ping():
    import app.zpl as _zpl
    return {"status": "ok", "servidor": _zpl.SERVIDOR_URL}


# ── Home ──────────────────────────────────────────────────────────────────────

@app.get("/", response_class=HTMLResponse)
@require_login
async def home(request: Request):
    templates_ativos = templates_mod.listar_templates_ativos()
    sessoes_em_andamento = sessions_mod.listar_sessoes_em_andamento()
    return render(request, "index.html", {
        "templates_kit": [t for t in templates_ativos if t.get("tipo", "kit") == "kit"],
        "templates_pedido": [t for t in templates_ativos if t.get("tipo") == "pedido"],
        "sessoes_em_andamento": sessoes_em_andamento,
        # A remessa lembrada da última escolha — o operador só troca quando
        # precisa; do contrário segue no mesmo lote kit após kit.
        "remessa_atual": remessas_mod.listar_uma(_remessa_escolhida(request) or 0),
    })


@app.post("/session/start")
@require_login
async def session_start(request: Request, kit_template_id: int = Form(...)):
    user = get_current_user(request)
    sessao_id = sessions_mod.start_session(kit_template_id, user["id"],
                                           _remessa_escolhida(request))
    return RedirectResponse(f"/session/{sessao_id}/destino", status_code=302)


# ── Escolher a remessa no começo da bipagem ───────────────────────────────────

def _remessa_escolhida(request: Request) -> int | None:
    """A remessa que este operador escolheu por último, se ainda estiver
    aberta.

    Fica na sessão de login de propósito: quem monta trinta kits seguidos
    escolhe uma vez e segue. Remessa que fechou (bateu o alvo) sai sozinha da
    escolha, senão o kit seguinte entraria num lote já encerrado."""
    rid = request.session.get("remessa_id")
    if not rid:
        return None
    r = remessas_mod.listar_uma(int(rid))
    if not r or r["status"] != "aberta":
        request.session.pop("remessa_id", None)
        return None
    return r["id"]


@app.get("/session/remessa", response_class=HTMLResponse)
@require_login
async def session_remessa_escolher(request: Request, sessao_id: int = 0):
    """Janela de escolha: as remessas abertas e o formulário de criar uma."""
    return render(request, "_remessa_escolher.html", {
        "abertas": remessas_mod.listar_abertas(),
        "escolhida": _remessa_escolhida(request),
        "clientes_cadastrados": clientes_mod.listar(),
        "sessao_id": sessao_id,
    })


@app.post("/session/remessa")
@require_login
async def session_remessa_definir(request: Request):
    form = await request.form()
    sessao_id = int(form.get("sessao_id") or 0)
    escolha = str(form.get("remessa_id", "")).strip()
    user = get_current_user(request)

    if escolha == "nova":
        if not permissoes_mod.tem_permissao(user, "remessas_gerenciar"):
            return RedirectResponse("/?erro=" + quote(
                "Seu usuário não pode criar remessa. Escolha uma da lista."), status_code=302)
        try:
            r = remessas_mod.abrir(str(form.get("nome", "")), form.get("alvo", ""),
                                   str(form.get("cliente", "")), user["id"] if user else None)
        except ValueError as e:
            destino = f"/session/{sessao_id}" if sessao_id else "/"
            return RedirectResponse(destino + "?erro=" + quote(str(e)), status_code=302)
        rid = r["id"]
    else:
        rid = int(escolha) if escolha.isdigit() else None

    request.session["remessa_id"] = rid
    if sessao_id:
        sessions_mod.definir_remessa(sessao_id, rid)
        return RedirectResponse(f"/session/{sessao_id}?ok=remessa", status_code=302)
    return RedirectResponse("/?ok=remessa", status_code=302)


# ── Admin: Tipos de Item ──────────────────────────────────────────────────────


@app.post("/admin/tipos/importar")
@require_login
async def admin_tipos_importar(request: Request, arquivo: UploadFile = File(...)):
    try:
        conteudo = await _ler_upload(request, arquivo)
        resultado = items_mod.importar_tipos_xlsx(conteudo)
        params = f"importado={resultado['criados']}&ignorado={resultado['ignorados']}"
    except Exception as e:
        params = f"erro_import={quote(_erro_usuario(e))}"
    return RedirectResponse(f"/admin/items?{params}", status_code=302)


@app.post("/admin/tipos/importar-bom")
@require_login
async def admin_tipos_importar_bom(request: Request, arquivo: UploadFile = File(...)):
    user = get_current_user(request)
    try:
        conteudo = await _ler_upload(request, arquivo)
        resultado = items_mod.importar_bom_xlsx(conteudo, user["id"])
        if "erro" in resultado:
            params = f"erro_import={quote(resultado['erro'])}"
        else:
            t, i = resultado["tipos_criados"], resultado["itens_criados"]
            ign = resultado["ignorados"]
            params = f"importado_bom=1&tipos={t}&itens={i}&ignorado={ign}"
    except Exception as e:
        params = f"erro_import={quote(_erro_usuario(e))}"
    return RedirectResponse(f"/admin/items?{params}", status_code=302)


@app.post("/admin/tipos/{tipo_id}/toggle-reutilizavel")
@require_login
async def admin_tipo_toggle_reutilizavel(request: Request, tipo_id: int):
    items_mod.alternar_reutilizavel_tipo(tipo_id)
    return RedirectResponse("/admin/items", status_code=302)


@app.post("/admin/tipos/{tipo_id}/toggle-controle-externo")
@require_login
async def admin_tipo_toggle_controle_externo(request: Request, tipo_id: int):
    items_mod.alternar_controle_externo(tipo_id)
    return RedirectResponse("/admin/items", status_code=302)


@app.post("/admin/tipos/{tipo_id}/toggle-requer-serial")
@require_login
async def admin_tipo_toggle_requer_serial(request: Request, tipo_id: int):
    items_mod.alternar_requer_serial(tipo_id)
    return RedirectResponse("/admin/items", status_code=302)


@app.post("/admin/tipos/{tipo_id}/toggle-unidade")
@require_login
async def admin_tipo_toggle_unidade(request: Request, tipo_id: int):
    items_mod.alternar_unidade_tipo(tipo_id)
    return RedirectResponse("/admin/items", status_code=302)


@app.post("/admin/tipos/{tipo_id}/renomear")
@require_login
async def admin_tipo_renomear(request: Request, tipo_id: int):
    form = await request.form()
    novo_nome = (form.get("nome") or "").strip()
    if novo_nome:
        try:
            items_mod.renomear_tipo(tipo_id, novo_nome)
        except Exception:
            pass
    return RedirectResponse("/admin/items", status_code=302)


@app.post("/admin/tipos/{tipo_id}/delete")
@require_admin
async def admin_tipo_delete(request: Request, tipo_id: int):
    try:
        items_mod.deletar_tipo(tipo_id)
        return RedirectResponse("/admin/items", status_code=302)
    except Exception:
        deps = items_mod.buscar_dependencias_tipo(tipo_id)
        return render(request, "admin_items.html", {
            **_admin_items_context(tab="catalogo"),
            "tipo_com_erro": deps,
        })


@app.post("/admin/tipos/{tipo_id}/delete-force")
@require_admin
async def admin_tipo_delete_force(request: Request, tipo_id: int):
    items_mod.deletar_tipo_cascade(tipo_id)
    return RedirectResponse("/admin/items", status_code=302)


@app.post("/admin/tipos/{tipo_id}/set-codigo-fixo")
@require_login
async def admin_tipo_set_codigo_fixo(request: Request, tipo_id: int):
    form = await request.form()
    codigo = str(form.get("codigo_fixo", "")).strip()
    items_mod.definir_codigo_fixo(tipo_id, codigo or None)
    return RedirectResponse("/admin/items", status_code=302)


# ── Admin: Itens (Patrimônios) ────────────────────────────────────────────────

ABAS_ITENS = ("catalogo", "novo", "patrimonios", "codigos", "sobressalentes")


def _admin_items_context(sobressalente_cliente: str = "",
                         sobressalente_data_ini: str = "",
                         sobressalente_data_fim: str = "",
                         patrimonio_veiculo_id: int | None = None,
                         patrimonio_pagina: int = 1,
                         codigos_pagina: int = 1,
                         tab: str = "",
                         patrimonio_situacao: str = "",
                         busca: str = "",
                         q_situacao: list[str] | None = None,
                         q_min: str = "",
                         q_max: str = "") -> dict:
    aba = tab if tab in ABAS_ITENS else "catalogo"
    busca = (busca or "").strip()
    # Filtros de múltipla escolha chegam como lista.
    q_situacao = filtros_mod.lista(q_situacao)
    patrimonio_situacao = filtros_mod.lista(
        patrimonio_situacao if isinstance(patrimonio_situacao, (list, tuple))
        else [patrimonio_situacao])

    # Só a aba realmente aberta consulta o banco. A lista de patrimônios é
    # de longe a consulta mais cara (cruza toda a tabela de bipagens), e
    # antes ela rodava mesmo quando o usuário estava vendo Estoque ou
    # Sobressalentes. Cada aba paga só o próprio custo.
    vazio = paginacao_mod.paginar([], 1)
    ctx = {
        # Resolvido aqui e não só no JS: se a aba certa só fosse aplicada
        # depois que a página carrega, o navegador pintaria a primeira aba
        # antes de trocar — o "pisca" que aparecia ao recarregar filtrado.
        "tab_ativo": aba,
        "busca": busca,
        "q_situacao": q_situacao,
        "q_min": q_min,
        "q_max": q_max,
        "niveis_estoque": {},
        "total_tipos": 0,
        "alerta_config": estoque_mod.ALERTA_PADRAO,
        "alerta_telas_opcoes": estoque_mod.ALERTA_TELAS,
        "pag_itens": vazio,
        "pag_codigos": vazio,
        "patrimonio_veiculo_id": patrimonio_veiculo_id,
        "patrimonio_situacao": patrimonio_situacao,
        "situacoes_itens": items_mod.SITUACOES,
        "resumo_situacoes": Counter(),
        "veiculos_todos": [],
        "tipos": [],
        "estoque_por_tipo": {},
        "estoque_itens": [],
        "clientes": [],
        "status_compra_opcoes": estoque_mod.STATUS_COMPRA,
        "sobressalente_cliente": sobressalente_cliente,
        "sobressalente_data_ini": sobressalente_data_ini,
        "sobressalente_data_fim": sobressalente_data_fim,
        "sobressalente_itens_enviados": [],
    }

    if aba == "patrimonios":
        # A contagem por situação vem da lista SEM o filtro de situação —
        # senão, depois de filtrar, o select mostraria zero em todas as
        # outras opções e não daria mais pra navegar entre elas.
        itens_do_veiculo = items_mod.listar_itens(veiculo_id=patrimonio_veiculo_id)
        ctx["resumo_situacoes"] = Counter(i["situacao"] for i in itens_do_veiculo)
        alvo_sit = [s for s in patrimonio_situacao if s in items_mod.SITUACOES]
        if alvo_sit:
            itens_do_veiculo = [i for i in itens_do_veiculo
                                if i["situacao"] in set(alvo_sit)]
        # Busca antes de paginar: varre a lista toda, não só a página aberta.
        itens_do_veiculo = paginacao_mod.filtrar(
            itens_do_veiculo, busca,
            ("codigo_barra", "descricao", "veiculo_atual", "serial_atual", "operador_atual"))
        ctx["pag_itens"] = paginacao_mod.paginar(itens_do_veiculo, patrimonio_pagina)
        # Só o que a caixa de seleção usa: listar() contaria os kits de cada
        # um dos 3.000 veículos pra jogar a contagem fora.
        ctx["veiculos_todos"] = veiculos_mod.listar_para_escolha()

    elif aba == "catalogo":
        tipos = items_mod.listar_tipos()
        estoque_por_tipo = {e["item_tipo_id"]: e for e in estoque_mod.listar_estoque()}
        if busca:
            # O código de barras do tipo mora no estoque, não no tipo — junta
            # os dois pra busca achar tanto por nome quanto por código.
            tipos = [t for t in tipos if paginacao_mod.filtrar(
                [{"nome": t["nome"],
                  "codigo": (estoque_por_tipo.get(t["id"]) or {}).get("codigo_barra", "")}],
                busca, ("nome", "codigo"))]

        # Filtro por QUANTIDADE. A situação usa o mesmo nivel_do_item() da
        # faixa de aviso — filtrar por "no vermelho" e o aviso do topo não
        # podem discordar sobre o que é vermelho. A margem de "atenção" vem
        # da mesma configuração.
        cfg_alerta = estoque_mod.get_alerta_config()
        margem = cfg_alerta["alerta_margem"]

        def _nivel(t):
            est = estoque_por_tipo.get(t["id"])
            return estoque_mod.nivel_do_item(est, margem) if est else "sem_estoque"

        if q_situacao:
            # Várias situações somam: "zerado" + "perto do mínimo" traz as duas.
            alvo = set(q_situacao)
            if "alerta" in alvo:      # o mesmo conjunto que a faixa de aviso mostra
                alvo |= {"zerado", "critico", "atencao"}
            tipos = [t for t in tipos
                     if _nivel(t) in alvo
                     or ("sem_estoque" in alvo and t["id"] not in estoque_por_tipo)]

        def _num(texto):
            try:
                return int(str(texto).strip())
            except (TypeError, ValueError):
                return None

        n_min, n_max = _num(q_min), _num(q_max)
        if n_min is not None or n_max is not None:
            # Tipo sem estoque cadastrado não tem quantidade — fica de fora de
            # uma faixa numérica em vez de contar como zero.
            def _na_faixa(t):
                est = estoque_por_tipo.get(t["id"])
                if not est:
                    return False
                q = est["quantidade_atual"]
                return (n_min is None or q >= n_min) and (n_max is None or q <= n_max)
            tipos = [t for t in tipos if _na_faixa(t)]

        ctx["tipos"] = tipos
        ctx["estoque_por_tipo"] = estoque_por_tipo
        ctx["niveis_estoque"] = {t["id"]: _nivel(t) for t in tipos}
        ctx["total_tipos"] = len(items_mod.listar_tipos())
        ctx["alerta_config"] = cfg_alerta
        ctx["alerta_telas_opcoes"] = estoque_mod.ALERTA_TELAS

    elif aba == "novo":
        ctx["tipos"] = items_mod.listar_tipos()

    elif aba == "codigos":
        codigos = paginacao_mod.filtrar(
            codigos_gerados_mod.listar(), busca, ("texto", "criado_por_nome"))
        ctx["pag_codigos"] = paginacao_mod.paginar(codigos, codigos_pagina)

    elif aba == "sobressalentes":
        ctx["clientes"] = clientes_mod.listar()
        # A lista de itens e os envios só aparecem depois de escolher o
        # cliente — antes disso não há o que consultar.
        if sobressalente_cliente:
            ctx["estoque_itens"] = estoque_mod.listar_estoque()
            ctx["sobressalente_itens_enviados"] = estoque_mod.listar_sobressalentes(
                sobressalente_data_ini, sobressalente_data_fim, sobressalente_cliente
            )
    return ctx


@app.get("/admin/items", response_class=HTMLResponse)
@require_login
async def admin_items(request: Request, cliente: str = "", data_ini: str = "", data_fim: str = "",
                      veiculo_id: str = "", pagina: int = 1, codigos_pagina: int = 1,
                      tab: str = "", busca: str = "",
                      q_min: str = "", q_max: str = "",
                      situacao: list[str] = Query(default=[]),
                      q_situacao: list[str] = Query(default=[])):
    return render(request, "admin_items.html", _admin_items_context(
        cliente, data_ini, data_fim,
        patrimonio_veiculo_id=int(veiculo_id) if veiculo_id.isdigit() else None,
        patrimonio_pagina=pagina,
        codigos_pagina=codigos_pagina,
        tab=tab,
        patrimonio_situacao=situacao,
        busca=busca,
        q_situacao=q_situacao, q_min=q_min, q_max=q_max,
    ))


def _patrimonio_context(request: Request, codigo_barra: str) -> dict:
    """Tudo que a tela do patrimônio mostra. Numa função só porque a prévia
    de "mover" reabre a MESMA tela — e ela não pode chegar mais pobre."""
    item = items_mod.buscar_item(codigo_barra)
    historico = items_mod.historico_patrimonio(codigo_barra)
    sessao_recente = historico[0]["sessao_id"] if historico else None
    vizinhos = (items_mod.bipados_na_mesma_sessao(sessao_recente, codigo_barra)
                if sessao_recente else [])
    return {
        # Volta pra tela de onde a pessoa veio (com busca/filtro/página),
        # não pra lista pelada.
        "voltar_para": _voltar_para(request, "/admin/items?tab=patrimonios"),
        "codigo_barra": codigo_barra,
        "item": item,
        "historico": historico,
        "vizinhos": vizinhos,
        # De QUAL kit são os itens de "Bipado junto" — a seção fala do kit
        # formado naquela sessão, não de itens vizinhos no tempo.
        "kit_sessao": items_mod.kit_da_sessao(sessao_recente) if sessao_recente else None,
        "serial_atual": historico[0]["serial_number"] if historico else None,
        # Onde o item está AGORA, antes de qualquer alteração: é a pergunta
        # que o operador precisa ver respondida pra decidir o que fazer.
        "onde_esta": items_mod.onde_esta(codigo_barra),
        "previa_mover": None,
        "mover_destino": "",
        "mover_motivo": "",
        "ok": request.query_params.get("ok", ""),
        "erro": request.query_params.get("erro", ""),
    }


# ATENCAO A ORDEM: esta rota vem ANTES da tela de detalhe porque o
# conversor {codigo:path} e guloso — declarada depois, ela nunca seria
# alcancada: /patrimonio/PAT-A/painel casaria com o detalhe, tratando
# "PAT-A/painel" como se fosse o codigo do patrimonio.
@app.get("/admin/items/patrimonio/{codigo_barra:path}/painel", response_class=HTMLResponse)
@require_login
async def admin_patrimonio_painel(request: Request, codigo_barra: str, voltar: str = ""):
    """Pedaço de tela com tudo que dá pra fazer com um patrimônio — pra abrir
    numa janela sobre a lista, em vez de mandar o operador pra outra página e
    depois pra outra."""
    # O MESMO painel do veículo, com este item em foco: entrar pelo código ou
    # pelo número do veículo tem que dar no mesmo lugar.
    return render(request, "_painel.html",
                  _painel_context(request, None, codigo_barra, voltar, modo="item"))


@app.get("/admin/items/patrimonio/{codigo_barra:path}", response_class=HTMLResponse)
@require_login
async def admin_patrimonio_detalhe(request: Request, codigo_barra: str):
    """Rastreamento de um patrimônio: todo lugar onde ele foi bipado, por
    quem, pra qual veículo — e o que mais foi bipado na mesma sessão."""
    return render(request, "admin_patrimonio.html",
                  _patrimonio_context(request, codigo_barra))


def _destino_apos_acao(request: Request, voltar: str, codigo_barra: str,
                       msg: str, erro: bool = False) -> str:
    """Pra onde voltar depois de mexer no patrimônio.

    Quando a ação saiu da JANELA (aberta em cima de uma lista), a pessoa tem
    que voltar exatamente pra aquela lista, com filtro e página — e não pra
    página do patrimônio, que era um pulo pra fora do que ela estava fazendo.
    `voltar` só é aceito se for caminho do próprio site."""
    chave = "erro" if erro else "ok"
    if voltar.startswith("/") and not voltar.startswith("//"):
        junta = "&" if "?" in voltar else "?"
        return f"{voltar}{junta}{chave}=" + quote(msg)
    return f"/admin/items/patrimonio/{quote(codigo_barra)}?{chave}=" + quote(msg)


@app.post("/admin/items/patrimonio/{codigo_barra:path}/painel", response_class=HTMLResponse)
@require_login
async def admin_patrimonio_painel_previa(request: Request, codigo_barra: str):
    """Prévia dentro da janela, entrando pelo código do patrimônio.

    É a mesma janela do veículo, então trata os dois formulários — quem entra
    por aqui também precisa poder adicionar item. Prévia não grava nada; quem
    grava são as rotas de mover/atribuir, cada uma com sua permissão."""
    form = await request.form()
    ctx = _painel_context(request, None, codigo_barra, str(form.get("voltar", "")),
                          modo="item")
    return render(request, "_painel.html", _preencher_previa(ctx, form, codigo_barra))


@app.post("/admin/items/patrimonio/{codigo_barra:path}/mover")
@require_permission("patrimonio_mover")
async def admin_patrimonio_mover(request: Request, codigo_barra: str):
    """Passa o patrimônio pro kit de outro veículo.

    Dois passos de propósito: o primeiro mostra de onde sai, pra onde vai e o
    que cada lado perde ou ganha; só o segundo grava. Motivo é obrigatório
    nos dois — é ele que responde, meses depois, por que o item mudou de
    veículo."""
    form = await request.form()
    destino = str(form.get("destino", "")).strip()
    motivo = str(form.get("motivo", "")).strip()
    if str(form.get("confirmado", "")) != "1":
        previa = items_mod.previa_mover(codigo_barra, destino)
        return render(request, "admin_patrimonio.html", {
            **_patrimonio_context(request, codigo_barra),
            "previa_mover": previa,
            "mover_destino": destino,
            "mover_motivo": motivo,
        })
    voltar = str(form.get("voltar", ""))
    try:
        r = items_mod.mover_patrimonio(codigo_barra, destino, motivo,
                                       (get_current_user(request) or {}).get("id"))
    except ValueError as e:
        return RedirectResponse(
            _destino_apos_acao(request, voltar, codigo_barra, str(e), erro=True),
            status_code=302)
    msg = (f"Patrimônio {codigo_barra} movido para o veículo {r['destino']['veiculo']}. "
           f"O kit do veículo {r['origem']['veiculo'] or '—'} ficou faltando este item "
           "e aparece na lista de pendências até receber outro.")
    return RedirectResponse(_destino_apos_acao(request, voltar, codigo_barra, msg),
                            status_code=302)


@app.post("/admin/items/patrimonio/{codigo_barra:path}/retirar")
@require_permission("patrimonio_atribuir")
async def admin_patrimonio_retirar(request: Request, codigo_barra: str):
    """Tira o patrimônio do kit sem colocar em outro (voltou pro estoque,
    quebrou, sumiu). O kit passa a acusar o item faltando — que é a verdade."""
    form = await request.form()
    voltar = str(form.get("voltar", ""))
    try:
        r = items_mod.retirar_do_kit(codigo_barra, str(form.get("motivo", "")),
                                     (get_current_user(request) or {}).get("id"))
    except ValueError as e:
        return RedirectResponse(
            _destino_apos_acao(request, voltar, codigo_barra, str(e), erro=True),
            status_code=302)
    msg = (f"Patrimônio {codigo_barra} retirado do kit do veículo "
           f"{r['origem']['veiculo'] or '—'}. Esse kit entrou na lista de pendências "
           "até receber outro item.")
    return RedirectResponse(_destino_apos_acao(request, voltar, codigo_barra, msg),
                            status_code=302)


@app.post("/kit/{kit_id}/trazer-patrimonio")
@require_permission("patrimonio_mover")
async def kit_trazer_patrimonio(request: Request, kit_id: str):
    """Traz pra este kit um patrimônio que está em outro veículo — a mesma
    mudança de "mover", só que escolhida do lado de quem vai receber.

    É como o operador pensa na hora de repor: "esse veículo está sem antena,
    qual antena eu trago?". Antes ele precisava saber o código de cabeça e ir
    até a página daquele patrimônio."""
    form = await request.form()
    try:
        r = items_mod.mover_para_kit(
            str(form.get("codigo_barra", "")), kit_id, str(form.get("motivo", "")),
            (get_current_user(request) or {}).get("id"))
    except ValueError as e:
        return RedirectResponse(
            _destino_apos_acao(request, str(form.get("voltar", "")), "", str(e), erro=True)
            if form.get("voltar") else f"/kit/{quote(kit_id)}?erro=" + quote(str(e)),
            status_code=302)
    msg = (f"{r['origem']['tipo_nome']} {str(form.get('codigo_barra', '')).strip()} "
           f"veio do veículo {r['origem']['veiculo'] or '—'}, que ficou faltando "
           "este item e entrou na lista de pendências.")
    voltar = str(form.get("voltar", ""))
    return RedirectResponse(
        _destino_apos_acao(request, voltar, "", msg) if voltar
        else f"/kit/{quote(kit_id)}?ok=" + quote(msg), status_code=302)


@app.post("/kit/{kit_id}/atribuir-patrimonio")
@require_permission("patrimonio_atribuir")
async def kit_atribuir_patrimonio(request: Request, kit_id: str):
    """Cadastra e coloca um patrimônio num kit JÁ FECHADO — o veículo que
    ficou sem o item recebendo a peça de reposição."""
    form = await request.form()
    try:
        tipo_id = int(str(form.get("item_tipo_id", "0")) or 0)
    except ValueError:
        tipo_id = 0
    try:
        r = items_mod.atribuir_patrimonio(
            str(form.get("codigo_barra", "")), kit_id, tipo_id,
            str(form.get("motivo", "")), str(form.get("serial", "")),
            (get_current_user(request) or {}).get("id"))
    except ValueError as e:
        return RedirectResponse(
            _destino_apos_acao(request, str(form.get("voltar", "")), "", str(e), erro=True)
            if form.get("voltar") else f"/kit/{quote(kit_id)}?erro=" + quote(str(e)),
            status_code=302)
    msg = (f"{r['tipo_nome']} atribuído ao kit do veículo {r['veiculo'] or '—'}"
           + (" (patrimônio cadastrado agora)." if r["novo_cadastro"] else "."))
    voltar = str(form.get("voltar", ""))
    return RedirectResponse(
        _destino_apos_acao(request, voltar, "", msg) if voltar
        else f"/kit/{quote(kit_id)}?ok=" + quote(msg), status_code=302)


@app.post("/admin/items/patrimonio/{codigo_barra:path}/corrigir")
@require_permission("patrimonio_corrigir")
async def admin_patrimonio_corrigir(request: Request, codigo_barra: str):
    """Correção cadastral do patrimônio — vale em qualquer estágio, até com
    o veículo já entregue e finalizado. NÃO mexe em produção: o kit continua
    no estágio em que estava; só a identificação do item muda."""
    form = await request.form()
    novo_codigo = str(form.get("novo_codigo", "")).strip()
    serial_bruto = form.get("novo_serial")
    novo_serial = str(serial_bruto) if serial_bruto is not None else None
    voltar = str(form.get("voltar", ""))
    try:
        # Motivo obrigatório também aqui: renomear patrimônio é mudança de
        # identidade do item, e sem o porquê o histórico não se explica. E
        # agora ele é GRAVADO, não só exigido — vira a entrada de "Alterações
        # depois da montagem" do veículo.
        motivo = items_mod._validar_motivo(str(form.get("motivo", "")))
        user = get_current_user(request)
        r = items_mod.corrigir_patrimonio(codigo_barra, novo_codigo, novo_serial,
                                          motivo, user["id"] if user else None)
    except ValueError as e:
        return RedirectResponse(
            _destino_apos_acao(request, voltar, codigo_barra, str(e), erro=True),
            status_code=302)
    partes = []
    if r["renomeou"]:
        partes.append(f"Código alterado de {codigo_barra} para {r['codigo']} — "
                      "as bipagens antigas acompanharam.")
    if r["seriais_atualizados"]:
        partes.append("Número de série atualizado na bipagem mais recente.")
    msg = " ".join(partes) or "Nada foi alterado."
    return RedirectResponse(_destino_apos_acao(request, voltar, r["codigo"], msg),
                            status_code=302)


@app.get("/admin/gerar-codigo/etiqueta", response_class=HTMLResponse)
@require_login
async def admin_gerar_codigo_etiqueta(request: Request, texto: str = ""):
    """Gera uma etiqueta avulsa (QR + código de barras do mesmo texto livre),
    sem precisar de um item_tipo ou registro de estoque — para conjuntos e
    outros códigos de componente definidos na hora de montar um kit."""
    import app.zpl as _zpl
    texto = texto.strip()
    if not texto:
        raise HTTPException(status_code=400, detail="Texto é obrigatório.")
    user = get_current_user(request)
    codigos_gerados_mod.registrar(texto, user["id"])
    html = _zpl.generate_estoque_html_label(tipo_nome=texto, codigo_barra=texto, url_qr=texto)
    return HTMLResponse(content=html)


@app.post("/admin/codigos-gerados/{codigo_id}/toggle-reciclavel")
@require_login
async def admin_codigo_gerado_toggle_reciclavel(request: Request, codigo_id: int):
    codigos_gerados_mod.toggle_reciclavel(codigo_id)
    return RedirectResponse("/admin/items?tab=codigos", status_code=302)


@app.post("/admin/codigos-gerados/{codigo_id}/renomear")
@require_login
async def admin_codigo_gerado_renomear(request: Request, codigo_id: int):
    form = await request.form()
    try:
        codigos_gerados_mod.renomear(codigo_id, str(form.get("texto", "")))
    except ValueError as e:
        return RedirectResponse(
            "/admin/items?tab=codigos&erro=" + quote(str(e)), status_code=302)
    return RedirectResponse("/admin/items?tab=codigos&ok=codigo_renomeado", status_code=302)


@app.post("/admin/tipos/completo")
@require_login
async def admin_tipos_completo(request: Request):
    """Cria um tipo de item e, opcionalmente, já atribui estoque (quantidade
    + código de barras) num único passo — usado pela aba 'Novo Item'."""
    user = get_current_user(request)
    form = await request.form()
    nome = (form.get("nome") or "").strip()
    unidade = form.get("unidade") or "un"
    reutilizavel = bool(form.get("reutilizavel"))
    codigo_barra = (form.get("codigo_barra") or "").strip()

    if not nome:
        return RedirectResponse(
            "/admin/items?tab=novo&erro=" + quote("Nome do tipo é obrigatório."),
            status_code=302)

    try:
        quantidade = max(0, int(form.get("quantidade") or 0))
        quantidade_minima = max(0, int(form.get("quantidade_minima") or 5))
        tipo_id = items_mod.criar_tipo(nome, unidade)
        if reutilizavel:
            items_mod.alternar_reutilizavel_tipo(tipo_id)
        if codigo_barra:
            estoque_mod.criar_estoque(tipo_id, codigo_barra, quantidade,
                                       quantidade_minima, user["id"])
    except Exception as e:
        return RedirectResponse(
            "/admin/items?tab=novo&erro=" + quote(_erro_usuario(e, "Não foi possível cadastrar o item.")),
            status_code=302)

    return RedirectResponse("/admin/items?ok=item_criado", status_code=302)


@app.post("/admin/estoque/{estoque_id}/codigo")
@require_permission("estoque_editar")
async def admin_estoque_codigo(request: Request, estoque_id: int):
    form = await request.form()
    try:
        estoque_mod.atualizar_codigo_barra(estoque_id, form.get("codigo_barra", ""))
    except Exception as e:
        return RedirectResponse(
            "/admin/items?erro=" + quote(_erro_usuario(e, "Não foi possível atualizar o código.")),
            status_code=302)
    return RedirectResponse("/admin/items?ok=codigo_atualizado", status_code=302)


@app.post("/admin/items")
@require_login
async def admin_items_post(request: Request,
                           codigo_barra: str = Form(...),
                           item_tipo_id: int = Form(...)):
    user = get_current_user(request)
    try:
        codigo_barra = codigo_barra.strip()
        items_mod.criar_item(codigo_barra, item_tipo_id, user["id"])
        codigos_gerados_mod.sincronizar_tipo_se_reciclavel(codigo_barra, item_tipo_id)
        return RedirectResponse("/admin/items?ok=1", status_code=302)
    except Exception as e:
        return render(request, "admin_items.html",
                      {**_admin_items_context(tab="patrimonios"),
                       "erro": _erro_usuario(e, "Não foi possível salvar o item.")})


@app.post("/admin/items/clear")
@require_admin
async def admin_items_clear(request: Request):
    items_mod.apagar_todos_itens()
    return RedirectResponse("/admin/items", status_code=302)


@app.post("/admin/items/{item_id}/delete")
@require_permission("itens_apagar")
async def admin_items_delete(request: Request, item_id: int):
    try:
        items_mod.deletar_item(item_id)
    except ValueError as e:
        # A recusa tem MOTIVO ("está no kit do veículo X") e ele precisa
        # chegar na tela: "não foi possível" mandava o operador tentar de
        # novo sem saber o que fazer diferente.
        return RedirectResponse("/admin/items?tab=patrimonios&erro=" + quote(str(e)),
                                status_code=302)
    except Exception:
        return RedirectResponse(
            "/admin/items?tab=patrimonios&erro=" + quote(
                "Não foi possível excluir o patrimônio."), status_code=302)
    return RedirectResponse("/admin/items?tab=patrimonios&ok=item_excluido",
                            status_code=302)


# ── Admin: Templates ──────────────────────────────────────────────────────────

def _admin_templates_context(pagina_kit: int = 1, pagina_pedido: int = 1,
                             busca: str = "") -> dict:
    todos = templates_mod.listar_todos()
    campos = ("nome", "cliente")
    templates_kit = paginacao_mod.filtrar(
        [t for t in todos if t.get("tipo", "kit") == "kit"], busca, campos)
    templates_pedido = paginacao_mod.filtrar(
        [t for t in todos if t.get("tipo") == "pedido"], busca, campos)
    return {
        "busca": busca,
        "pag_kit": paginacao_mod.paginar(templates_kit, pagina_kit),
        "pag_pedido": paginacao_mod.paginar(templates_pedido, pagina_pedido),
        "tipos_catalogo": items_mod.listar_tipos(apenas_ativos=True),
        "clientes": clientes_mod.listar(),
        "consumo_resumo": consumo_mod.resumo_todos_kits(),
        "conjuntos": templates_mod.listar_todos_conjuntos(),
    }


@app.get("/admin/templates", response_class=HTMLResponse)
@require_login
async def admin_templates(request: Request, pagina_kit: int = 1, pagina_pedido: int = 1,
                          busca: str = ""):
    return render(request, "admin_templates.html", {
        **_admin_templates_context(pagina_kit, pagina_pedido, busca),
        "erro": request.query_params.get("erro"),
    })


@app.post("/admin/templates/{template_id}/conjunto/verificar-em-conjunto")
@require_login
async def admin_conjunto_verifica(request: Request, template_id: int):
    form = await request.form()
    componente_codigo = str(form.get("componente_codigo", "")).strip()
    verifica = form.get("verifica") == "1"
    if componente_codigo:
        templates_mod.definir_verifica_em_conjunto(template_id, componente_codigo, verifica)
    return RedirectResponse("/admin/templates?tab=conjuntos", status_code=302)


@app.post("/admin/templates/import-bom")
@require_login
async def admin_templates_import_bom(request: Request,
                                      nome: str = Form(""),
                                      cliente: str = Form(""),
                                      tipo: str = Form("kit"),
                                      arquivo: UploadFile = File(...)):
    user = get_current_user(request)
    nome, cliente = nome.strip(), cliente.strip()
    tipo = tipo if tipo in ("kit", "pedido") else "kit"
    if not nome or not cliente:
        return render(request, "admin_templates.html", {
            **_admin_templates_context(),
            "erro": "Preencha nome e cliente antes de importar o BOM.",
            "tab_ativo": tipo,
        })
    try:
        conteudo = await _ler_upload(request, arquivo)
        template_id, stats = templates_mod.criar_template_do_bom(
            nome, cliente, user["id"], conteudo, tipo=tipo
        )
        q = f"ok=bom&itens={stats['itens_adicionados']}&tipos={stats['tipos_criados']}"
        return RedirectResponse(f"/admin/templates/{template_id}/edit?{q}", status_code=302)
    except ValueError as e:
        return render(request, "admin_templates.html", {
            **_admin_templates_context(),
            "erro": str(e),
            "tab_ativo": tipo,
        })


@app.post("/admin/templates/{template_id}/unidades")
@require_permission("pedidos_criar_editar")
async def admin_pedido_unidades(request: Request, template_id: int):
    """Salva as unidades do pedido: corrige as existentes e cadastra as
    novas, num envio só. Correção cadastral — não toca em bipagem, kit nem
    estoque; a unidade é a ficha do aparelho que vai naquele pedido."""
    form = await request.form()
    destino = f"/admin/templates/{template_id}/edit"

    # Edição das já cadastradas: listas paralelas indexadas por id.
    ids = form.getlist("unidade_id")
    campos = {c: form.getlist("edit_" + c) for c in
              ("iccid", "telefone", "cdt", "id_hardware")}
    editadas = 0
    for i, uid in enumerate(ids):
        if not str(uid).strip().isdigit():
            continue
        if pedidos_mod.atualizar_unidade(
                int(uid), {c: (campos[c][i] if i < len(campos[c]) else "")
                           for c in campos}):
            editadas += 1

    # Novas linhas, em branco no formulário — as vazias são ignoradas.
    novos = {c: form.getlist("nova_" + c) for c in
             ("iccid", "telefone", "cdt", "id_hardware")}
    total_novas = max((len(v) for v in novos.values()), default=0)
    linhas = [{c: (novos[c][i] if i < len(novos[c]) else "") for c in novos}
              for i in range(total_novas)]
    r = pedidos_mod.adicionar_unidades(template_id, linhas)

    partes = []
    if r["inseridas"]:
        partes.append(f"{r['inseridas']} unidade(s) adicionada(s)")
    if editadas:
        partes.append(f"{editadas} atualizada(s)")
    msg = " · ".join(partes) or "Nada foi alterado."
    return RedirectResponse(destino + "?ok_unidades=" + quote(msg), status_code=302)


@app.post("/admin/templates/{template_id}/unidades/{unidade_id}/remover")
@require_permission("pedidos_criar_editar")
async def admin_pedido_unidade_remover(request: Request, template_id: int, unidade_id: int):
    pedidos_mod.remover_unidade(unidade_id)
    return RedirectResponse(
        f"/admin/templates/{template_id}/edit?ok_unidades="
        + quote("Unidade removida."), status_code=302)


@app.post("/admin/templates/import-pedido")
@require_permission("pedidos_criar_editar")
async def admin_templates_import_pedido(request: Request,
                                         cliente: str = Form(""),
                                         numero_pedido: str = Form(""),
                                         arquivo: UploadFile = File(...)):
    """Cria um ou mais Pedidos a partir da planilha de unidades (ICCID,
    Número de Telefone, CDT, ID Hardware) — diferente do BOM do Kit: não
    cria itens do template, só guarda as unidades para consulta. Uma
    planilha pode conter vários pedidos ao mesmo tempo (agrupados por uma
    coluna 'Pedido'); cada um vira um Pedido separado. Os itens de cada
    pedido são adicionados manualmente depois, na tela de edição."""
    user = get_current_user(request)
    cliente = cliente.strip()
    if not cliente:
        return render(request, "admin_templates.html", {
            **_admin_templates_context(),
            "erro": "Selecione o cliente antes de importar a planilha do pedido.",
            "tab_ativo": "pedido",
        })
    try:
        conteudo = await _ler_upload(request, arquivo)
        template_id, stats = pedidos_mod.importar_planilha(
            cliente, numero_pedido, user["id"], conteudo
        )
        if stats["pedidos"] == 1:
            q = f"ok=pedido&unidades={stats['unidades']}&numero={quote(stats['numeros'][0])}"
            return RedirectResponse(f"/admin/templates/{template_id}/edit?{q}", status_code=302)
        q = (f"ok=pedidos&pedidos={stats['pedidos']}&unidades={stats['unidades']}"
             f"&ignoradas={stats['ignoradas']}&tab=pedido")
        return RedirectResponse(f"/admin/templates?{q}", status_code=302)
    except ValueError as e:
        return render(request, "admin_templates.html", {
            **_admin_templates_context(),
            "erro": str(e),
            "tab_ativo": "pedido",
        })


@app.post("/admin/templates")
@require_login
async def admin_templates_post(request: Request):
    user = get_current_user(request)
    form = await request.form()
    nome = form.get("nome", "").strip()
    cliente = form.get("cliente", "").strip()
    tipo = form.get("tipo", "kit").strip()
    tipo = tipo if tipo in ("kit", "pedido") else "kit"
    if tipo == "pedido" and not permissoes_mod.tem_permissao(user, "pedidos_criar_editar"):
        return render(request, "admin_templates.html", {
            **_admin_templates_context(),
            "erro": "Seu usuário não tem permissão pra criar Pedidos.",
            "tab_ativo": tipo,
        })
    itens = _parse_itens_form(form)
    if not nome or not cliente or not itens:
        return render(request, "admin_templates.html", {
            **_admin_templates_context(),
            "erro": "Preencha nome, cliente e ao menos 1 item.",
            "tab_ativo": tipo,
        })
    templates_mod.criar_template(nome, cliente, user["id"], itens, tipo=tipo)
    return RedirectResponse(f"/admin/templates?ok=1&tab={tipo}", status_code=302)


@app.get("/admin/templates/{template_id}/edit", response_class=HTMLResponse)
@require_login
async def admin_template_edit_page(request: Request, template_id: int):
    template = templates_mod.buscar_template(template_id)
    if not template:
        return RedirectResponse("/admin/templates", status_code=302)
    user = get_current_user(request)
    if template.get("tipo") == "pedido" and not permissoes_mod.tem_permissao(user, "pedidos_criar_editar"):
        return RedirectResponse(
            "/admin/templates?erro=" + quote("Seu usuário não tem permissão pra editar Pedidos.")
            + "&tab=pedido", status_code=302)
    itens = templates_mod.get_itens_template(template_id)
    tipos_ativos = items_mod.listar_tipos(apenas_ativos=True)
    clientes = clientes_mod.listar()
    sessoes_em_andamento = sessions_mod.listar_sessoes_em_andamento(template_id=template_id)
    unidades = pedidos_mod.listar_unidades(template_id) if template.get("tipo") == "pedido" else []
    consumo = consumo_mod.analise_template(template_id)
    return render(request, "admin_template_edit.html", {
        # Voltar tem que cair na aba de onde se veio: editando um Pedido, o
        # botão levava pra lista de Kits — a aba é escolhida pela query, e o
        # padrão agora segue o tipo do que está aberto.
        "voltar_para": _voltar_para(
            request, "/admin/templates?tab=" + ("pedido" if template.get("tipo") == "pedido" else "kit")),
        # Quantos veículos apontam pro NOME atual deste kit. Renomear
        # desliga todos eles da bipagem em silêncio — a tela avisa antes.
        "veiculos_vinculados": veiculos_mod.contar_por_modelo(template["nome"]),
        # Valores repetidos entre as unidades: a tela MARCA, não bloqueia.
        "duplicados": (pedidos_mod.duplicados_do_pedido(template_id)
                       if template.get("tipo") == "pedido" else {"iccid": set(), "id_hardware": set()}),
        "template": template,
        "itens": itens,
        "consumo": consumo,
        "tipos_catalogo": tipos_ativos,
        "clientes": clientes,
        "sessoes_em_andamento": sessoes_em_andamento,
        "unidades": unidades,
    })


@app.post("/admin/templates/{template_id}/edit")
@require_login
async def admin_template_edit_post(request: Request, template_id: int):
    template_atual = templates_mod.buscar_template(template_id)
    if template_atual and template_atual.get("tipo") == "pedido":
        user = get_current_user(request)
        if not permissoes_mod.tem_permissao(user, "pedidos_criar_editar"):
            return RedirectResponse(
                "/admin/templates?erro=" + quote("Seu usuário não tem permissão pra editar Pedidos.")
                + "&tab=pedido", status_code=302)
    form = await request.form()
    nome = form.get("nome", "").strip()
    cliente = form.get("cliente", "").strip()
    itens = _parse_itens_form(form)
    if not nome or not cliente or not itens:
        template = templates_mod.buscar_template(template_id)
        itens_atuais = templates_mod.get_itens_template(template_id)
        tipos_ativos = items_mod.listar_tipos(apenas_ativos=True)
        clientes = clientes_mod.listar()
        return render(request, "admin_template_edit.html", {
            "template": template, "itens": itens_atuais,
            "tipos_catalogo": tipos_ativos,
            "clientes": clientes,
            "erro": "Preencha nome, cliente e ao menos 1 item.",
        })
    templates_mod.atualizar_template(template_id, nome, cliente, itens)
    template = templates_mod.buscar_template(template_id)
    tipo = template.get("tipo", "kit") if template else "kit"
    return RedirectResponse(f"/admin/templates?ok=editado&tab={tipo}", status_code=302)


@app.get("/admin/templates/{template_id}/unidades/exportar.xlsx")
@require_login
async def admin_template_unidades_exportar(request: Request, template_id: int):
    from fastapi.responses import Response as _Resp
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from io import BytesIO

    template = templates_mod.buscar_template(template_id)
    if not template:
        raise HTTPException(status_code=404)
    unidades = pedidos_mod.listar_unidades(template_id)
    itens_template = templates_mod.get_itens_template(template_id)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Unidades"
    azul, branco, cinza = "1A3A5C", "FFFFFF", "F4F7FB"

    for col, h in enumerate(["ICCID", "Número de Telefone", "CDT", "ID Hardware"], 1):
        c = ws.cell(1, col, h)
        c.font = Font(bold=True, color=branco)
        c.fill = PatternFill("solid", fgColor=azul)
        c.alignment = Alignment(horizontal="center", vertical="center")

    for i, u in enumerate(unidades):
        row = i + 2
        ws.cell(row, 1, u.get("iccid") or "")
        ws.cell(row, 2, u.get("telefone") or "")
        ws.cell(row, 3, u.get("cdt") or "")
        ws.cell(row, 4, u.get("id_hardware") or "")
        if i % 2 == 0:
            for col in range(1, 5):
                ws.cell(row, col).fill = PatternFill("solid", fgColor=cinza)
    for col, w in zip("ABCD", (24, 22, 18, 22)):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"

    ws2 = wb.create_sheet("Itens do Pedido")
    for col, h in enumerate(["Item", "Quantidade Exigida", "Obrigatório", "Unidade"], 1):
        c = ws2.cell(1, col, h)
        c.font = Font(bold=True, color=branco)
        c.fill = PatternFill("solid", fgColor=azul)
        c.alignment = Alignment(horizontal="center", vertical="center")
    for i, item in enumerate(itens_template):
        row = i + 2
        ws2.cell(row, 1, item["descricao"])
        ws2.cell(row, 2, item["quantidade_exigida"])
        ws2.cell(row, 3, "Sim" if item["obrigatorio"] else "Não")
        ws2.cell(row, 4, item.get("unidade") or "un")
        if i % 2 == 0:
            for col in range(1, 5):
                ws2.cell(row, col).fill = PatternFill("solid", fgColor=cinza)
    for col, w in zip("ABCD", (32, 20, 14, 12)):
        ws2.column_dimensions[col].width = w
    ws2.freeze_panes = "A2"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    import re as _re
    safe = _re.sub(r'[^\w\-]', '_', template["nome"])
    return _Resp(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{safe}_unidades.xlsx"'},
    )


@app.post("/admin/templates/{template_id}/delete")
@require_admin
async def admin_template_delete(request: Request, template_id: int):
    template = templates_mod.buscar_template(template_id)
    tipo = template.get("tipo", "kit") if template else "kit"
    try:
        templates_mod.deletar_template(template_id)
        return RedirectResponse(f"/admin/templates?ok=excluido&tab={tipo}", status_code=302)
    except ValueError as e:
        return render(request, "admin_templates.html", {
            **_admin_templates_context(),
            "erro": str(e),
            "tab_ativo": tipo,
        })


@app.post("/admin/templates/{template_id}/nova-versao")
@require_login
async def admin_template_nova_versao(request: Request, template_id: int):
    user = get_current_user(request)
    template = templates_mod.buscar_template(template_id)
    tipo = template.get("tipo", "kit") if template else "kit"
    templates_mod.nova_versao(template_id, user["id"])
    return RedirectResponse(f"/admin/templates?ok=versao&tab={tipo}", status_code=302)


@app.post("/admin/templates/{template_id}/toggle")
@require_login
async def admin_template_toggle(request: Request, template_id: int):
    template = templates_mod.buscar_template(template_id)
    tipo = template.get("tipo", "kit") if template else "kit"
    templates_mod.toggle_ativo(template_id)
    return RedirectResponse(f"/admin/templates?tab={tipo}", status_code=302)


@app.post("/admin/templates/{template_id}/toggle-concluido")
@require_login
async def admin_template_toggle_concluido(request: Request, template_id: int):
    template = templates_mod.buscar_template(template_id)
    tipo = template.get("tipo", "kit") if template else "kit"
    templates_mod.toggle_concluido(template_id)
    return RedirectResponse(f"/admin/templates?tab={tipo}", status_code=302)


# ── Sessions ──────────────────────────────────────────────────────────────────

@app.get("/session/{sessao_id}", response_class=HTMLResponse)
@require_login
async def session_page(request: Request, sessao_id: int):
    session = sessions_mod.get_session(sessao_id)
    if not session:
        return RedirectResponse("/", status_code=302)
    if session["status"] != "em_andamento":
        return RedirectResponse("/", status_code=302)
    if not session.get("garagem"):
        # Destino ainda não escolhido — obrigatório antes de bipar (pode
        # acontecer com link direto/voltar do navegador, ou sessão antiga
        # de antes dessa mudança existir).
        return RedirectResponse(f"/session/{sessao_id}/destino", status_code=302)
    itens = templates_mod.get_itens_template(session["kit_template_id"])
    contagem = sessions_mod.get_contagem(sessao_id)
    itens_por_operador = sessions_mod.listar_itens_por_operador(sessao_id)
    return render(request, "session.html", {
        "session": session,
        "itens": itens,
        "contagem": contagem,
        "itens_por_operador": itens_por_operador,
        "remessa_atual": remessas_mod.listar_uma(session.get("remessa_id") or 0),
    })


@app.get("/session/{sessao_id}/remessa", response_class=HTMLResponse)
@require_login
async def session_remessa_pagina(request: Request, sessao_id: int):
    """Trocar a remessa no meio da bipagem. Página curta em vez de janela: a
    tela de bipagem não usa a janela única do sistema."""
    session = sessions_mod.get_session(sessao_id)
    if not session or session["status"] != "em_andamento":
        return RedirectResponse("/", status_code=302)
    return render(request, "remessa_pagina.html", {
        "abertas": remessas_mod.listar_abertas(),
        "escolhida": session.get("remessa_id"),
        "clientes_cadastrados": clientes_mod.listar(),
        "sessao_id": sessao_id,
    })


@app.get("/session/{sessao_id}/destino", response_class=HTMLResponse)
@require_login
async def session_destino_page(request: Request, sessao_id: int, erro: str = ""):
    session = sessions_mod.get_session(sessao_id)
    if not session:
        return RedirectResponse("/", status_code=302)
    if session["status"] != "em_andamento":
        return RedirectResponse("/", status_code=302)
    if session.get("garagem"):
        # Destino já escolhido — não pergunta de novo, vai direto pra bipagem.
        return RedirectResponse(f"/session/{sessao_id}", status_code=302)
    # Só os veículos DESTE modelo — o modelo do veículo é o nome do kit.
    # Antes qualquer veículo do cliente aparecia em qualquer kit, o que
    # deixava atribuir um Euro 5 a um kit de Euro 6.
    veiculos_lista = veiculos_mod.listar(
        cliente=session.get("cliente", ""),
        modelo=session.get("kit_nome", ""))
    for v in veiculos_lista:
        v["ocupado"] = veiculos_mod.esta_ocupado(v["id"])
    # Sem lista de garagens: a garagem não é mais escolhida aqui, vem do
    # cadastro do veículo.
    return render(request, "session_destino.html", {
        "session": session,
        "veiculos_lista": veiculos_lista,
        # Pra explicar a lista vazia: sem veículo com este modelo, é preciso
        # definir o modelo no cadastro — não adianta procurar na tela.
        "sem_modelo": veiculos_mod.contar_sem_modelo(session.get("cliente", "")),
        "erro": erro,
    })


@app.post("/session/{sessao_id}/destino")
@require_login
async def session_destino_post(request: Request, sessao_id: int):
    session = sessions_mod.get_session(sessao_id)
    if not session or session["status"] != "em_andamento":
        return RedirectResponse("/", status_code=302)

    # Só o veículo vem do formulário. Garagem e modelo são DERIVADOS aqui —
    # do cadastro do veículo e do nome do kit — e não aceitos do cliente,
    # pra não existir caminho (nem por engano, nem por requisição forjada)
    # que grave um destino diferente do que está cadastrado.
    form = await request.form()
    veiculo_id_str = str(form.get("veiculo_id", "")).strip()
    veiculo_id = int(veiculo_id_str) if veiculo_id_str.isdigit() else None

    if not veiculo_id:
        return RedirectResponse(
            f"/session/{sessao_id}/destino?erro=" +
            quote("Selecione o veículo antes de continuar."),
            status_code=302)

    v = veiculos_mod.buscar(veiculo_id)
    if not v:
        return RedirectResponse(
            f"/session/{sessao_id}/destino?erro=" +
            quote("Veículo não encontrado."),
            status_code=302)

    veiculo_texto = v["numero"]
    garagem = (v["garagem"] or "").strip()
    modelo = session.get("kit_nome", "") or ""

    if not garagem:
        return RedirectResponse(
            f"/session/{sessao_id}/destino?erro=" +
            quote(f"O veículo {veiculo_texto} não tem garagem no cadastro. "
                  "Defina a garagem dele em Veículos e Clientes e volte aqui."),
            status_code=302)

    if veiculos_mod.esta_ocupado(veiculo_id):
        return RedirectResponse(
            f"/session/{sessao_id}/destino?erro=" +
            quote("Esse veículo já tem kit associado. Libere em Veículos e Clientes antes de atribuir de novo."),
            status_code=302)

    # Não regrava a garagem no veículo: ela agora VEM de lá. Antes o
    # operador podia escolher outra na tela e o cadastro era atualizado —
    # hoje isso só reescreveria o mesmo valor em maiúsculas, mexendo no
    # cadastro sem motivo. Alterar garagem é em Veículos e Clientes.
    sessions_mod.definir_destino(sessao_id, veiculo_id, veiculo_texto, garagem, modelo)
    veiculos_mod.consumir_liberacao(veiculo_id)
    return RedirectResponse(f"/session/{sessao_id}", status_code=302)


@app.get("/session/{sessao_id}/trocar-modelo", response_class=HTMLResponse)
@require_admin
async def session_trocar_modelo_page(request: Request, sessao_id: int,
                                     novo: str = "", erro: str = ""):
    """Tela de correção: troca o kit de uma bipagem já em andamento.

    Restrita a admin de propósito — mexer no kit no meio da produção muda o
    que a etiqueta vai dizer. Ao escolher um kit, a tela mostra a PRÉVIA do
    impacto antes de confirmar: o que já bipado continua valendo, o que
    sobra e o que ainda vai faltar."""
    session = sessions_mod.get_session(sessao_id)
    if not session or session["status"] != "em_andamento":
        return RedirectResponse("/", status_code=302)
    novo_id = int(novo) if novo.isdigit() else None
    previa = (sessions_mod.comparar_troca_template(sessao_id, novo_id)
              if novo_id else None)
    veiculo = veiculos_mod.buscar(session["veiculo_id"]) if session.get("veiculo_id") else None
    return render(request, "session_trocar_modelo.html", {
        "session": session,
        "templates": templates_mod.listar_templates_ativos(),
        "novo_id": novo_id,
        "previa": previa,
        "veiculo": veiculo,
        "erro": erro,
    })


@app.post("/session/{sessao_id}/trocar-modelo")
@require_admin
async def session_trocar_modelo_post(request: Request, sessao_id: int):
    form = await request.form()
    novo_str = str(form.get("kit_template_id", "")).strip()
    if not novo_str.isdigit():
        return RedirectResponse(
            f"/session/{sessao_id}/trocar-modelo?erro=" + quote("Escolha o kit novo."),
            status_code=302)

    resultado = sessions_mod.trocar_template(sessao_id, int(novo_str))
    if resultado["resultado"] != "trocado":
        return RedirectResponse(
            f"/session/{sessao_id}/trocar-modelo?erro=" + quote(resultado["mensagem"]),
            status_code=302)

    # O veículo tinha o modelo antigo — foi por ele que apareceu na lista.
    # Corrigir o cadastro junto é o que impede o mesmo erro de se repetir na
    # próxima bipagem desse veículo; sem isso ele voltaria a aparecer no kit
    # errado. Opcional porque às vezes o errado foi o kit, não o veículo.
    session = sessions_mod.get_session(sessao_id)
    if form.get("atualizar_veiculo") and session and session.get("veiculo_id"):
        veiculos_mod.definir_modelo(session["veiculo_id"], resultado["kit_nome"])

    return RedirectResponse(
        f"/session/{sessao_id}?ok=" +
        quote(f"Kit trocado de \"{resultado['kit_anterior']}\" para "
              f"\"{resultado['kit_nome']}\". As bipagens em comum foram mantidas."),
        status_code=302)


@app.post("/session/{sessao_id}/cancel")
@require_login
async def session_cancel(request: Request, sessao_id: int):
    sessions_mod.cancel_session(sessao_id)
    return RedirectResponse("/", status_code=302)


@app.post("/admin/sessoes/{sessao_id}/cancelar")
@require_login
async def admin_sessao_cancelar(request: Request, sessao_id: int):
    """Admin cancela uma sessão em andamento para liberar template para edição/exclusão."""
    session = sessions_mod.get_session(sessao_id)
    template_id = session["kit_template_id"] if session else None
    sessions_mod.cancel_session(sessao_id)
    if template_id:
        return RedirectResponse(f"/admin/templates/{template_id}/edit?cancelou=1", status_code=302)
    return RedirectResponse("/admin/templates", status_code=302)


@app.websocket("/ws/session/{sessao_id}")
async def ws_session(websocket: WebSocket, sessao_id: int):
    session_data = websocket.scope.get("session", {})
    user_id = session_data.get("user_id")
    sid = session_data.get("sid", "")
    await websocket.accept()
    if not user_id:
        await websocket.close(code=1008)
        return
    try:
        while True:
            data = await websocket.receive_text()
            data = data.strip()
            if not data:
                continue
            # Bipar É usar o sistema: sem isto, quem passasse vinte minutos
            # só lendo códigos seria deslogado por "inatividade" no meio do
            # kit — o WebSocket não passa pelo middleware HTTP.
            inatividade_mod.tocar(sid)
            try:
                msg = json.loads(data)
                if not isinstance(msg, dict):
                    raise ValueError("not a JSON object")
                if msg.get("acao") == "identificar":
                    result = sessions_mod.register_scan(
                        sessao_id, msg["codigo"],
                        item_tipo_id=int(msg["item_tipo_id"]),
                        operador_id=user_id,
                    )
                elif msg.get("acao") == "confirmar_quantidade":
                    result = sessions_mod.confirmar_quantidade(
                        sessao_id, msg["codigo_barra"], float(msg.get("quantidade", 1)),
                        operador_id=user_id,
                    )
                elif msg.get("acao") == "confirmar_substituicao":
                    result = sessions_mod.confirmar_substituicao(
                        sessao_id, msg["codigo_barra"], msg.get("motivo", ""),
                        operador_id=user_id,
                    )
                elif msg.get("acao") == "confirmar_componente":
                    result = sessions_mod.confirmar_componente(
                        sessao_id, msg["codigo_barra"], msg.get("quantidades", {}),
                        operador_id=user_id,
                    )
                elif msg.get("acao") == "cancelar_serial":
                    result = sessions_mod.cancelar_serial(sessao_id)
                elif msg.get("acao") == "cancelar_patrimonio_fixo":
                    result = sessions_mod.cancelar_patrimonio_fixo(sessao_id)
                elif msg.get("acao") == "desfazer_ultimo":
                    result = sessions_mod.desfazer_ultimo_item(sessao_id)
                else:
                    result = {"resultado": "rejeitado", "mensagem": "Mensagem inválida."}
            except (json.JSONDecodeError, KeyError, ValueError):
                # Plain barcode scan — priority: serial > patrimônio fixo > componente > normal
                pendente_serial = sessions_mod.get_pendente_serial(sessao_id)
                if pendente_serial:
                    result = sessions_mod.registrar_serial(sessao_id, data, operador_id=user_id)
                else:
                    pendente_fixo = sessions_mod.get_pendente_patrimonio_fixo(sessao_id)
                    if pendente_fixo:
                        result = sessions_mod.registrar_patrimonio_de_fixo(sessao_id, data, operador_id=user_id)
                    else:
                        # Conjunto registra direto, sem modal de confirmação —
                        # a quantidade vem do template e o operador só ouve
                        # o beep, igual a qualquer outro item.
                        result = sessions_mod.registrar_conjunto(sessao_id, data, operador_id=user_id)
                        if result is None:
                            result = sessions_mod.register_scan(sessao_id, data, operador_id=user_id)
            await websocket.send_json(result)
    except WebSocketDisconnect:
        pass

# ── Finalização ───────────────────────────────────────────────────────────────

@app.post("/session/{sessao_id}/finalize")
@require_login
async def session_finalize(request: Request, sessao_id: int):
    user = get_current_user(request)

    session_check = sessions_mod.get_session(sessao_id)
    if not session_check or session_check["status"] != "em_andamento":
        return RedirectResponse("/", status_code=302)
    if not session_check.get("garagem"):
        # Destino não foi definido (não deveria acontecer — a tela de bipagem
        # só é alcançada depois do /destino — mas não finaliza sem isso).
        return RedirectResponse(f"/session/{sessao_id}/destino", status_code=302)

    validation = sessions_mod.validate_kit_complete(sessao_id)
    if validation["status"] != "completo":
        faltam = "; ".join(
            f"{i['descricao']} (faltam {i['faltam']})"
            for i in validation["itens_faltantes"]
        )
        return RedirectResponse(
            f"/session/{sessao_id}?erro={quote(faltam)}", status_code=302
        )

    session = sessions_mod.get_session(sessao_id)
    contagem = sessions_mod.get_contagem(sessao_id)
    itens_template = templates_mod.get_itens_template(session["kit_template_id"])

    itens_label = []
    for it in itens_template:
        qtd = contagem.get(it["item_tipo_id"], 0)
        if qtd > 0:
            itens_label.append({"descricao": it["descricao"], "quantidade": qtd})

    kit_id = str(uuid.uuid4())
    ts = datetime.now(tz=BRT)

    veiculo = session.get("veiculo") or ""
    garagem = session.get("garagem") or ""
    modelo = session.get("modelo") or ""
    veiculo_id = session.get("veiculo_id")

    zpl = zpl_mod.generate_zpl(
        kit_id=kit_id,
        kit_nome=session["kit_nome"],
        cliente=session["cliente"],
        operador=session["operador_nome"],
        timestamp=ts,
        itens=itens_label,
        veiculo=veiculo,
        garagem=garagem,
        modelo=modelo,
    )

    html_label = zpl_mod.generate_html_label(
        kit_id=kit_id,
        kit_nome=session["kit_nome"],
        cliente=session["cliente"],
        operador=session["operador_nome"],
        timestamp=ts,
        itens=itens_label,
        veiculo=veiculo,
        garagem=garagem,
        modelo=modelo,
    )

    with db() as conn:
        ts_str = ts.strftime("%Y-%m-%d %H:%M:%S")
        # operador_id = quem ABRIU a bipagem (segue responsável pelo kit).
        # finalizado_por = quem clicou em finalizar. Quando são pessoas
        # diferentes, o kit foi feito em dupla e as telas mostram os dois.
        conn.execute(
            "INSERT INTO kit_record (kit_id, sessao_id, kit_template_id, "
            "kit_template_versao, operador_id, finalizado_por, veiculo, garagem, modelo, "
            "finalizado_em, veiculo_id) "
            "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
            (kit_id, sessao_id, session["kit_template_id"],
             session["kit_template_versao"], session["operador_id"], user["id"],
             veiculo, garagem, modelo, ts_str, veiculo_id)
        )
        conn.execute(
            "UPDATE scan_session SET status = 'finalizado', "
            "finalizado_em = ? WHERE id = ?",
            (ts_str, sessao_id)
        )
        conn.execute(
            "INSERT INTO print_queue (kit_id, zpl, html_label, solicitado_por, solicitado_em) "
            "VALUES (?, ?, ?, ?, ?)",
            (kit_id, zpl, html_label, user["id"], now_brt())
        )
        # Congela o que este kit exigia AGORA. Editar o template depois muda
        # o kit dali pra frente, não o que já foi montado — sem isso, tirar
        # um item do template fazia todo kit antigo aparecer com aquele item
        # "sobrando", como se estivesse errado.
        sessions_mod.gravar_itens_exigidos(conn, kit_id, session["kit_template_id"])

    # Fora do `with`: vincular_kit abre a própria conexão, e aninhar duas
    # trava o SQLite. O kit entra na remessa escolhida lá no começo — é o que
    # faz o contador andar enquanto o galpão monta, e não só no despacho.
    if session.get("remessa_id"):
        remessas_mod.vincular_kit(kit_id, session["remessa_id"])

    if session.get("kit_tipo") == "pedido":
        templates_mod.marcar_concluido(session["kit_template_id"])

    return RedirectResponse(
        f"/session/{sessao_id}/complete?kit_id={kit_id}", status_code=302
    )


@app.get("/session/{sessao_id}/complete", response_class=HTMLResponse)
@require_login
async def session_complete(request: Request, sessao_id: int, kit_id: str):
    with db() as conn:
        pq_row = conn.execute(
            "SELECT * FROM print_queue WHERE kit_id = ? ORDER BY id DESC LIMIT 1",
            (kit_id,)
        ).fetchone()
    return render(request, "complete.html", {
        "kit_id": kit_id,
        "pq_id": dict(pq_row)["id"] if pq_row else None,
    })


# ── Fila de Impressão ─────────────────────────────────────────────────────────

@app.get("/print-queue", response_class=HTMLResponse)
@require_login
async def print_queue_page(request: Request):
    fila = pq_mod.listar_aguardando_tudo()
    return render(request, "print_queue.html", {"fila": fila})


@app.get("/print-queue/{pq_id}/zpl")
@require_login
async def print_queue_zpl(request: Request, pq_id: int):
    """Retorna o ZPL como download de arquivo .zpl para envio à Zebra."""
    from fastapi.responses import Response
    item = pq_mod.buscar(pq_id)
    if not item:
        return PlainTextResponse("Não encontrado", status_code=404)
    nome = f"etiqueta_{pq_id}.zpl"
    return Response(
        content=item["zpl"].encode("ascii", "replace"),
        media_type="application/octet-stream",
        headers={"Content-Disposition": f'attachment; filename="{nome}"'},
    )


@app.get("/print-queue/{pq_id}/etiqueta")
@require_login
async def print_queue_html_label(request: Request, pq_id: int):
    item = pq_mod.buscar(pq_id)
    if not item or not item.get("html_label"):
        return PlainTextResponse("Etiqueta HTML não disponível.", status_code=404)
    return HTMLResponse(item["html_label"])


@app.get("/print-queue/{pq_id}/preview")
@require_login
async def print_queue_preview(request: Request, pq_id: int):
    """Renderiza o ZPL como imagem PNG via Labelary (validação sem imprimir)."""
    import urllib.request as _urlreq
    from fastapi.responses import Response as _Resp
    item = pq_mod.buscar(pq_id)
    if not item:
        return PlainTextResponse("Não encontrado", status_code=404)
    zpl_bytes = item["zpl"].encode("ascii", "replace")
    # Labelary: 8 dpmm (203 DPI), 100x150mm = 3.94x5.91"
    url = "http://api.labelary.com/v1/printers/8dpmm/labels/3.94x5.91/0/"
    req = _urlreq.Request(url, data=zpl_bytes, method="POST")
    req.add_header("Content-Type", "application/x-www-form-urlencoded")
    req.add_header("Accept", "image/png")
    try:
        with _urlreq.urlopen(req, timeout=10) as resp:
            png = resp.read()
        return _Resp(content=png, media_type="image/png")
    except Exception as exc:
        return HTMLResponse(
            f'<body style="font-family:sans-serif;padding:20px;">'
            f'<h3>Erro ao renderizar via Labelary</h3><pre>{exc}</pre>'
            f'<p>Verifique se há conexão com a internet.</p></body>',
            status_code=502,
        )


@app.post("/print-queue/{pq_id}/impresso")
@require_login
async def print_queue_impresso(request: Request, pq_id: int):
    pq_mod.marcar_impresso(pq_id)
    return RedirectResponse("/print-queue", status_code=302)


@app.post("/print-queue/{pq_id}/cancelar")
@require_login
async def print_queue_cancelar(request: Request, pq_id: int):
    pq_mod.cancelar(pq_id)
    return RedirectResponse("/print-queue", status_code=302)


# ── Fila de etiquetas "Em Andamento" ────────────────────────────────────────────

@app.post("/session/{sessao_id}/imprimir-pausa")
@require_login
async def session_imprimir_pausa(request: Request, sessao_id: int):
    user = get_current_user(request)
    session = sessions_mod.get_session(sessao_id)
    if not session or session["status"] != "em_andamento":
        return RedirectResponse("/", status_code=302)
    sequencia = producao_mod.atribuir_sequencia(sessao_id)
    html_label = zpl_mod.generate_pausa_html_label(
        veiculo=session.get("veiculo") or "",
        kit_nome=session["kit_nome"],
        cliente=session["cliente"],
        garagem=session.get("garagem") or "",
        operador=session["operador_nome"],
        sequencia=sequencia,
        timestamp=datetime.now(tz=BRT),
    )
    pq_mod.adicionar_pausa(sessao_id, html_label, user["id"])
    return RedirectResponse(f"/session/{sessao_id}?ok=etiqueta_pausa", status_code=302)


@app.post("/session/{sessao_id}/item/{item_id}/remover")
@require_permission("bipagem_excluir_item")
async def session_remover_item(request: Request, sessao_id: int, item_id: int):
    resultado = sessions_mod.remover_item(sessao_id, item_id)
    if resultado["resultado"] != "item_removido":
        return RedirectResponse(
            f"/session/{sessao_id}?erro=" + quote(resultado["mensagem"]), status_code=302
        )
    return RedirectResponse(f"/session/{sessao_id}?ok=item_removido", status_code=302)


@app.get("/print-queue/pausa/{pq_id}/etiqueta")
@require_login
async def print_queue_pausa_html_label(request: Request, pq_id: int):
    item = pq_mod.buscar_pausa(pq_id)
    if not item or not item.get("html_label"):
        return PlainTextResponse("Etiqueta HTML não disponível.", status_code=404)
    return HTMLResponse(item["html_label"])


@app.post("/print-queue/pausa/{pq_id}/impresso")
@require_login
async def print_queue_pausa_impresso(request: Request, pq_id: int):
    pq_mod.marcar_impresso_pausa(pq_id)
    return RedirectResponse("/print-queue", status_code=302)


@app.post("/print-queue/pausa/{pq_id}/cancelar")
@require_login
async def print_queue_pausa_cancelar(request: Request, pq_id: int):
    pq_mod.cancelar_pausa(pq_id)
    return RedirectResponse("/print-queue", status_code=302)


# ── Mobile Hub (público) ──────────────────────────────────────────────────────

@app.get("/mobile", response_class=HTMLResponse)
async def mobile_hub(request: Request):
    user = get_current_user(request)
    sessoes_ativas = []
    sessoes_outros = []
    templates_list = []
    if user:
        with db() as conn:
            sessoes_ativas = conn.execute(
                "SELECT ss.id, kt.nome AS kit_nome, kt.cliente, ss.iniciado_em, "
                "ss.veiculo, ss.garagem "
                "FROM scan_session ss "
                "JOIN kit_template kt ON kt.id = ss.kit_template_id "
                "WHERE ss.operador_id = ? AND ss.status = 'em_andamento' "
                "ORDER BY ss.iniciado_em DESC",
                (user["id"],)
            ).fetchall()
            sessoes_outros = conn.execute(
                "SELECT ss.id, kt.nome AS kit_nome, kt.cliente, ss.iniciado_em, "
                "ss.veiculo, ss.garagem, u.nome AS operador_nome "
                "FROM scan_session ss "
                "JOIN kit_template kt ON kt.id = ss.kit_template_id "
                "JOIN users u ON u.id = ss.operador_id "
                "WHERE ss.operador_id != ? AND ss.status = 'em_andamento' "
                "ORDER BY ss.iniciado_em DESC",
                (user["id"],)
            ).fetchall()
            templates_list = conn.execute(
                "SELECT id, nome, cliente FROM kit_template WHERE ativo = 1 ORDER BY nome"
            ).fetchall()

    return render(request, "mobile_hub.html", {
        "user": user,
        "sessoes_ativas": [dict(s) for s in sessoes_ativas],
        "sessoes_outros": [dict(s) for s in sessoes_outros],
        "templates_list": [dict(t) for t in templates_list],
    })


# ── Kit Detail (público — escaneado pelo QR code) ─────────────────────────────

_RE_UUID = re.compile(
    r'[0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12}', re.I)


def _resolver_kit_id(texto: str) -> str | None:
    """Resolve um texto lido (URL do QR da etiqueta, kit_id completo, ou o
    ID curto de 8 caracteres do código de barras) para o kit_id completo
    correspondente, ou None se não encontrar.

    Aceita o texto do jeito que vier do leitor. Antes exigia que a URL
    TERMINASSE no id, e comparava caixa alta/baixa de forma inconsistente —
    então uma URL com ?parâmetro, #fragmento, ou lida em MAIÚSCULAS não
    resolvia. Isso importa porque leitores diferentes devolvem o mesmo QR de
    formas diferentes (o modo alfanumérico do QR, por exemplo, guarda só
    maiúsculas), o que fazia a MESMA etiqueta funcionar num aparelho e
    falhar em outro.

    Agora procura o UUID em qualquer posição do texto, sem depender do
    formato da URL em volta."""
    texto = (texto or "").strip()
    if not texto:
        return None

    # UUID em qualquer lugar do texto (URL, com query, com fragmento...)
    m = _RE_UUID.search(texto)
    if m:
        candidato = m.group(0).lower()
        with db() as conn:
            row = conn.execute(
                "SELECT kit_id FROM kit_record WHERE kit_id = ?", (candidato,)
            ).fetchone()
        if row:
            return row["kit_id"]

    # ID curto do código de barras (8 caracteres) — pega o último trecho
    # "limpo" do texto, pra funcionar mesmo se vier com prefixo.
    curto = re.sub(r'[^0-9a-fA-F]', '', texto.split('/')[-1])
    with db() as conn:
        if curto:
            rows = conn.execute(
                "SELECT kit_id FROM kit_record WHERE kit_id LIKE ?",
                (curto.lower() + '%',)
            ).fetchall()
            if len(rows) == 1:
                return rows[0]["kit_id"]
    return None


@app.get("/kit/buscar")
async def kit_buscar(request: Request, codigo: str = ""):
    """Resolve um código de barras ou o texto de um QR de kit para a
    página de verificação correspondente — usado pelo scanner do /mobile."""
    kit_id = _resolver_kit_id(codigo)
    if not kit_id:
        return RedirectResponse("/mobile?erro=kit_nao_encontrado", status_code=302)
    return RedirectResponse(f"/kit/{kit_id}", status_code=302)


@app.get("/kit/{kit_id}", response_class=HTMLResponse)
async def kit_detail(request: Request, kit_id: str):
    with db() as conn:
        kit = conn.execute(
            "SELECT kr.*, kt.nome AS kit_nome, kt.cliente, kt.versao, kt.tipo AS kit_tipo, "
            "u.nome AS operador_nome "
            "FROM kit_record kr "
            "JOIN kit_template kt ON kt.id = kr.kit_template_id "
            "JOIN users u ON u.id = kr.operador_id "
            "WHERE kr.kit_id = ?",
            (kit_id,)
        ).fetchone()
        if not kit:
            return HTMLResponse("<h2>Kit não encontrado.</h2>", status_code=404)
        kit = dict(kit)

        itens = conn.execute(
            "SELECT si.item_tipo_id, it.nome AS tipo_nome, COUNT(*) AS quantidade, "
            "GROUP_CONCAT(si.codigo_barra, ', ') AS barcodes, "
            # O código da caixa (o lote de onde a peça saiu) na conferência:
            # é o dado que faltava pra responder "de que lote é este CVC?"
            # sem abrir o histórico item por item.
            "GROUP_CONCAT(DISTINCT si.codigo_caixa) AS caixas "
            "FROM scan_session_items si "
            "JOIN item_tipo it ON it.id = si.item_tipo_id "
            "WHERE si.sessao_id = ? AND (si.status IS NULL OR si.status NOT IN ('movido', 'retirado')) "
            "GROUP BY si.item_tipo_id ORDER BY it.nome",
            (kit["sessao_id"],)
        ).fetchall()

    validacoes = validacoes_mod.listar_por_kit(kit_id)
    ok = request.query_params.get("ok", "")
    erro = request.query_params.get("erro", "")
    unidades = pedidos_mod.listar_unidades(kit["kit_template_id"]) if kit.get("kit_tipo") == "pedido" else []
    operadores_kit = sessions_mod.operadores_da_sessao(kit["sessao_id"])
    conferidos = validacoes_mod.listar_conferidos(kit_id)
    grupos = validacoes_mod.grupos_conjunto(kit["kit_template_id"], kit["sessao_id"])

    # O que o template passou a pedir (ou deixou de pedir) depois que este
    # kit foi montado. A conferência segue o template ATUAL: item que saiu do
    # modelo continua na caixa e na lista, mas não é mais cobrado no
    # checklist. A comparação é por tipo — quantidade total igual não basta.
    mudancas = sessions_mod.mudancas_do_template(kit_id)
    fora = {s["item_tipo_id"] for s in (mudancas["sairam"] if mudancas else [])}
    itens = [dict(i) for i in itens]
    for it in itens:
        it["fora_do_template"] = it["item_tipo_id"] in fora

    # Verificação feita ANTES da troca de modelo descreve outro conteúdo —
    # não vale como verificação do kit atual. O histórico continua todo
    # visível; o que muda é o selo de "verificado" e o aviso.
    corte = kit.get("verificacao_corte") or 0
    validacoes_vigentes = [v for v in validacoes if v["id"] > corte]
    validacoes_antigas = [v for v in validacoes if v["id"] <= corte]

    # O que este kit está devendo em relação ao modelo — é isso que vira o
    # aviso de pendência aqui e a marca do veículo nas outras telas.
    faltas = (sessions_mod.kits_incompletos().get(kit_id) or {}).get("faltas", [])
    # A lista de candidatos só é consultada pra quem pode movê-los: é uma
    # varredura de bipagens, e não vale pagar por ela pra quem não vai usar.
    _u = get_current_user(request)
    pode_ver_patrimonio = permissoes_mod.tem_permissao(_u, "patrimonio_mover")

    return render(request, "kit_detail.html", {
        # No celular o kit é aberto pelo hub; no computador, pela Produção,
        # pelo patrimônio ou pelo veículo. Mandar todo mundo pro hub deixava
        # quem veio da Produção sem caminho de volta.
        "voltar_para": _voltar_para(request, "/mobile"),
        "faltas": faltas,
        # Tipos que este kit aceita por patrimônio — as opções do formulário
        # de atribuir um item novo ao kit já fechado.
        "tipos_para_atribuir": items_mod.listar_tipos_para_kit(kit["kit_template_id"]),
        # Patrimônios que estão em OUTROS veículos e caberiam aqui: repor
        # escolhendo da lista, sem precisar decorar o código.
        "candidatos_outros": (items_mod.candidatos_de_outros_kits(kit_id)
                              if pode_ver_patrimonio else []),
        "kit": kit,
        "mudancas": mudancas,
        "validacoes_vigentes": validacoes_vigentes,
        "validacoes_antigas": validacoes_antigas,
        "itens": itens,
        "validacoes": validacoes,
        "ok": ok,
        "erro": erro,
        "unidades": unidades,
        "operadores_kit": operadores_kit,
        "conferidos": conferidos,
        "grupos": grupos,
    })


@app.get("/kit/{kit_id}/trocar-modelo", response_class=HTMLResponse)
@require_admin
async def kit_pronto_trocar_modelo_page(request: Request, kit_id: str,
                                        novo: str = "", erro: str = ""):
    """Troca o kit de um kit JÁ PRONTO — mesma lógica e mesma tela de
    prévia da troca em bipagem, mais o efeito no estoque (o que volta, o
    que sai, o que fica pendente). Só admin, igual à troca em bipagem."""
    with db() as conn:
        kr = conn.execute(
            "SELECT kr.*, kt.nome AS kit_nome, kt.cliente, kt.versao "
            "FROM kit_record kr JOIN kit_template kt ON kt.id = kr.kit_template_id "
            "WHERE kr.kit_id = ?", (kit_id,)).fetchone()
    if not kr:
        return HTMLResponse("<h2>Kit não encontrado.</h2>", status_code=404)
    novo_id = int(novo) if novo.isdigit() else None
    previa = (sessions_mod.previa_troca_kit_pronto(kit_id, novo_id)
              if novo_id and novo_id != kr["kit_template_id"] else None)
    return render(request, "kit_trocar_modelo.html", {
        "kit": dict(kr),
        "templates": templates_mod.listar_templates_ativos(),
        "novo_id": novo_id,
        "previa": previa,
        "erro": erro,
    })


@app.post("/kit/{kit_id}/trocar-modelo")
@require_admin
async def kit_pronto_trocar_modelo_post(request: Request, kit_id: str):
    user = get_current_user(request)
    form = await request.form()
    novo_str = str(form.get("kit_template_id", "")).strip()
    if not novo_str.isdigit():
        return RedirectResponse(
            f"/kit/{kit_id}/trocar-modelo?erro=" + quote("Escolha o kit novo."),
            status_code=302)
    resultado = sessions_mod.trocar_template_kit_pronto(
        kit_id, int(novo_str), operador_id=user["id"])
    if resultado["resultado"] != "trocado":
        return RedirectResponse(
            f"/kit/{kit_id}/trocar-modelo?erro=" + quote(resultado["mensagem"]),
            status_code=302)
    partes = [f"Kit trocado de \"{resultado['kit_anterior']}\" para \"{resultado['kit_nome']}\"."]
    if resultado["estornos"]:
        partes.append("Voltou ao estoque: " + ", ".join(resultado["estornos"]) + ".")
    if resultado["saidas"]:
        partes.append("Saiu do estoque: " + ", ".join(resultado["saidas"]) + ".")
    if resultado["pendentes"]:
        partes.append("Pendente (sem estoque vinculado): " + ", ".join(
            f"{p['descricao']} ({p['faltam']})" for p in resultado["pendentes"]) + ".")
    return RedirectResponse(f"/kit/{kit_id}?ok=" + quote(" ".join(partes)),
                            status_code=302)


@app.post("/kit/{kit_id}/conferir-item")
@require_login
async def kit_conferir_item(request: Request, kit_id: str):
    user = get_current_user(request)
    form = await request.form()
    item_tipo_id = int(form.get("item_tipo_id", 0) or 0)
    with db() as conn:
        kit = conn.execute(
            "SELECT kit_template_id, sessao_id FROM kit_record WHERE kit_id = ?", (kit_id,)
        ).fetchone()
    if not kit:
        return JSONResponse({"ok": False}, status_code=404)
    afetados = validacoes_mod.conferir_item(
        kit_id, kit["kit_template_id"], kit["sessao_id"], item_tipo_id, user["id"])
    return JSONResponse({"ok": True, "item_tipo_ids": afetados})


@app.post("/kit/{kit_id}/desfazer-item")
@require_login
async def kit_desfazer_item(request: Request, kit_id: str):
    user = get_current_user(request)
    form = await request.form()
    item_tipo_id = int(form.get("item_tipo_id", 0) or 0)
    with db() as conn:
        kit = conn.execute(
            "SELECT kit_template_id, sessao_id FROM kit_record WHERE kit_id = ?", (kit_id,)
        ).fetchone()
    if not kit:
        return JSONResponse({"ok": False}, status_code=404)
    afetados = validacoes_mod.desfazer_item(kit_id, kit["kit_template_id"], kit["sessao_id"], item_tipo_id)
    return JSONResponse({"ok": True, "item_tipo_ids": afetados})


@app.post("/kit/{kit_id}/validar")
@require_login
async def kit_validar(request: Request, kit_id: str):
    user = get_current_user(request)
    form = await request.form()
    observacao = str(form.get("observacao", "")).strip()
    with db() as conn:
        kit = conn.execute(
            "SELECT sessao_id, kit_template_id FROM kit_record WHERE kit_id = ?", (kit_id,)
        ).fetchone()
    if not kit:
        return HTMLResponse("<h2>Kit não encontrado.</h2>", status_code=404)
    # Cobra só o que o template ainda pede: item removido do modelo depois
    # que o kit foi fechado não pode travar a verificação.
    tipos_kit = validacoes_mod.tipos_do_kit(kit["sessao_id"], kit["kit_template_id"])
    conferidos = validacoes_mod.listar_conferidos(kit_id)
    if not tipos_kit.issubset(conferidos):
        return RedirectResponse(f"/kit/{kit_id}?erro=itens_pendentes", status_code=302)
    validacoes_mod.registrar(kit_id, user["id"], observacao)
    return RedirectResponse(f"/kit/{kit_id}?ok=validado", status_code=302)


@app.get("/reports/operadores", response_class=HTMLResponse)
@require_permission("ver_relatorios")
async def reports_operadores(request: Request,
                             data_ini: str = "", data_fim: str = "",
                             pagina: int = 1, busca: str = "",
                             operador_id: list[str] = Query(default=[])):
    """Kits por operador — inclui os que ainda estão em montagem, que o
    relatório de kits (só finalizados) não mostra."""
    linhas = sessions_mod.listar_por_operador(operador_id, data_ini, data_fim)
    linhas = paginacao_mod.filtrar(
        linhas, busca,
        ("kit_nome", "cliente", "veiculo", "garagem", "operador_nome", "finalizado_por_nome"))
    with db() as conn:
        usuarios = [dict(u) for u in conn.execute(
            "SELECT id, nome FROM users ORDER BY nome").fetchall()]
    return render(request, "reports_operadores.html", {
        "voltar_para": _voltar_para(request, "/reports"),
        "pag": paginacao_mod.paginar(linhas, pagina),
        "resumo": sessions_mod.resumo_por_operador(data_ini, data_fim),
        "usuarios": usuarios,
        "opcoes_usuarios": [(str(u["id"]), u["nome"]) for u in usuarios],
        "operador_id": operador_id,
        "data_ini": data_ini, "data_fim": data_fim,
        "busca": busca,
    })


# ── Relatórios ────────────────────────────────────────────────────────────────

def _where_relatorio_kits(data_ini: str, data_fim: str,
                          operador_id, tipo) -> tuple[str, list]:
    """Filtros do relatório de kits, num lugar só — a TELA e a EXPORTAÇÃO
    chamam esta função, então não há como uma filtrar diferente da outra.
    Antes cada uma montava o próprio WHERE e elas já divergiam no limite.

    Operador e tipo são listas (filtro de múltipla escolha); um valor solto
    também é aceito, pra link antigo continuar funcionando."""
    where = "WHERE 1=1"
    sql_data, params = datas_mod.clausula("kr.finalizado_em", data_ini, data_fim)
    where += sql_data
    ops = [int(o) for o in filtros_mod.lista(
        operador_id if isinstance(operador_id, (list, tuple)) else [operador_id])
        if str(o).isdigit()]
    sql_op, p_op = filtros_mod.em("kr.operador_id", ops)
    where += sql_op
    params += p_op
    tipos = [t for t in filtros_mod.lista(
        tipo if isinstance(tipo, (list, tuple)) else [tipo]) if t in ("kit", "pedido")]
    sql_tp, p_tp = filtros_mod.em("kt.tipo", tipos)
    where += sql_tp
    params += p_tp
    return where, params


@app.get("/reports", response_class=HTMLResponse)
@require_permission("ver_relatorios")
async def reports(request: Request,
                  data_ini: str = "",
                  data_fim: str = "",
                  pagina: int = 1,
                  operador_id: list[str] = Query(default=[]),
                  tipo: list[str] = Query(default=[])):
    where, params = _where_relatorio_kits(data_ini, data_fim, operador_id, tipo)

    por_pagina = paginacao_mod.POR_PAGINA_PADRAO
    with db() as conn:
        # Os MESMOS JOINs da consulta de baixo. Se as duas divergirem, o
        # total diz um número e a listagem entrega outro — e a diferença
        # vira registro que não aparece em página nenhuma.
        total = conn.execute(
            f"SELECT COUNT(*) FROM kit_record kr "
            f"JOIN kit_template kt ON kt.id = kr.kit_template_id "
            f"LEFT JOIN users u ON u.id = kr.operador_id {where}",
            params
        ).fetchone()[0]
        total_paginas = max(1, -(-total // por_pagina))
        pagina = max(1, min(pagina, total_paginas))
        offset = (pagina - 1) * por_pagina

        query = f"""
            SELECT kr.kit_id, kr.finalizado_em, kr.status,
                   kr.veiculo, kr.garagem,
                   kr.veiculo_id,
                   COALESCE(v.numero, kr.veiculo) AS veiculo_exibido,
                   v.id AS v_id,
                   kt.nome AS kit_nome, kt.cliente, kt.versao, kt.tipo AS kit_tipo,
                   kr.observacao,
                   COALESCE(u.nome, '—') AS operador_nome,
                   CASE WHEN kr.finalizado_por IS NOT NULL
                         AND kr.finalizado_por != kr.operador_id
                        THEN uf.nome END AS finalizado_por_nome,
                   -- Subconsulta, e não LEFT JOIN print_queue: um kit tem
                   -- UMA linha na fila por impressão, e reimprimir cria
                   -- outra. Com o JOIN, cada kit reimpresso 3x virava 4
                   -- linhas aqui — mas o COUNT abaixo (que não tem esse
                   -- join) seguia contando 1. A paginação era calculada
                   -- pelo número certo e aplicada sobre o resultado
                   -- inflado, então os kits do fim eram empurrados pra
                   -- depois da última página e ficavam INALCANÇÁVEIS.
                   -- Pega a etiqueta mais recente, que é a que o botão
                   -- "Etiqueta" deve abrir.
                   (SELECT pq.id FROM print_queue pq
                     WHERE pq.kit_id = kr.kit_id
                     ORDER BY pq.id DESC LIMIT 1) AS pq_id,
                   (SELECT COUNT(*) FROM kit_validacoes kv WHERE kv.kit_id = kr.kit_id) AS num_validacoes
            FROM kit_record kr
            JOIN kit_template kt ON kt.id = kr.kit_template_id
            -- LEFT, não INNER: kit cujo operador não exista mais sumiria do
            -- relatório inteiro em vez de aparecer sem o nome.
            LEFT JOIN users u ON u.id = kr.operador_id
            LEFT JOIN users uf ON uf.id = kr.finalizado_por
            LEFT JOIN veiculos v ON v.id = kr.veiculo_id
            {where}
            ORDER BY kr.finalizado_em DESC, kr.kit_id LIMIT ? OFFSET ?
        """
        rows = conn.execute(query, params + [por_pagina, offset]).fetchall()
        usuarios = conn.execute("SELECT id, nome FROM users ORDER BY nome").fetchall()

    veiculos_todos = veiculos_mod.listar()
    return render(request, "reports.html", {
        "kits": [dict(r) for r in rows],
        "usuarios": [dict(u) for u in usuarios],
        "opcoes_usuarios": [(str(u["id"]), u["nome"]) for u in usuarios],
        "data_ini": data_ini,
        "data_fim": data_fim,
        "operador_id": operador_id,
        "tipo": tipo,
        "ok": request.query_params.get("ok", ""),
        "veiculos_todos": veiculos_todos,
        "pagina": pagina,
        "total_paginas": total_paginas,
        "total_kits": total,
        # Mesma forma do paginacao.paginar() — aqui a paginação é feita no
        # SQL (a lista é grande demais pra carregar inteira), mas a tela usa
        # o mesmo macro de contagem das outras.
        "pag": {"pagina": pagina, "total_paginas": total_paginas, "total": total,
                "inicio": offset + 1 if rows else 0, "fim": offset + len(rows),
                "exibindo": len(rows)},
        "paginas_visiveis": paginacao_mod.janela_paginas(pagina, total_paginas),
    })


@app.post("/kit-record/{kit_id}/veiculo")
@require_login
async def kit_record_vincular_veiculo(request: Request, kit_id: str):
    """Atribuição manual de veículo a um kit. Fica só na tela do veículo
    (/admin/veiculos/{id}) — foi tirada do relatório pra não haver dois
    lugares mexendo na mesma coisa. `voltar_para` traz o id do veículo de
    onde a ação partiu, pra devolver o usuário na mesma tela.

    Se o kit JÁ tinha veículo, trocar ou desvincular exige motivo — mesma
    regra da nota fiscal: primeira atribuição é livre, mexer no que já
    estava definido precisa de justificativa. O motivo não vai pra
    kit_record; fica no log de auditoria, que o middleware grava com todos
    os campos do formulário (inclusive `veiculo_anterior`, enviado oculto
    pela tela justamente pra o log guardar a troca de → para)."""
    form = await request.form()
    veiculo_id_str = str(form.get("veiculo_id", "")).strip()
    veiculo_id = int(veiculo_id_str) if veiculo_id_str.isdigit() else None
    voltar_str = str(form.get("voltar_para", "")).strip()
    motivo = str(form.get("motivo", "")).strip()

    with db() as conn:
        atual = conn.execute(
            "SELECT veiculo_id FROM kit_record WHERE kit_id = ?", (kit_id,)
        ).fetchone()
    if not atual:
        return RedirectResponse("/admin/veiculos?erro=kit_nao_encontrado", status_code=302)

    destino = voltar_str if voltar_str.isdigit() else veiculo_id_str
    base = f"/admin/veiculos/{destino}" if destino.isdigit() else "/admin/veiculos"

    tinha_veiculo = atual["veiculo_id"] is not None
    mudando = atual["veiculo_id"] != veiculo_id
    if tinha_veiculo and mudando and not motivo:
        return RedirectResponse(f"{base}?erro=motivo_veiculo", status_code=302)

    veiculo_texto = ""
    garagem_texto = ""
    if veiculo_id:
        v = veiculos_mod.buscar(veiculo_id)
        if v:
            veiculo_texto = v["numero"]
            garagem_texto = v["garagem"]

    # A observação (opcional) fica no próprio kit, e não só na auditoria,
    # porque precisa aparecer no relatório junto do kit corrigido.
    observacao = str(form.get("observacao", "")).strip()
    with db() as conn:
        conn.execute(
            "UPDATE kit_record SET veiculo_id=?, veiculo=?, garagem=? WHERE kit_id=?",
            (veiculo_id, veiculo_texto, garagem_texto, kit_id)
        )
        if observacao:
            conn.execute(
                "UPDATE kit_record SET observacao=? WHERE kit_id=?",
                (observacao, kit_id)
            )
    return RedirectResponse(f"{base}?ok=veiculo_kit", status_code=302)


@app.post("/reports/reprint/{kit_id}")
@require_permission("ver_relatorios")
async def reprint_kit(request: Request, kit_id: str):
    """Recria a entrada na fila de impressão para um kit já finalizado.
    Se a garagem enviada for diferente da gravada, atualiza o kit_record e
    regenera a etiqueta (ZPL + HTML) com o novo valor; caso contrário,
    reimprime a última etiqueta já gerada, sem recalcular nada."""
    user = get_current_user(request)
    form = await request.form()
    nova_garagem = str(form.get("garagem", "")).strip().upper()

    with db() as conn:
        kit_row = conn.execute(
            "SELECT kr.*, kt.nome AS kit_nome, kt.cliente, u.nome AS operador_nome "
            "FROM kit_record kr "
            "JOIN kit_template kt ON kt.id = kr.kit_template_id "
            "JOIN users u ON u.id = kr.operador_id "
            "WHERE kr.kit_id = ?",
            (kit_id,)
        ).fetchone()
    if not kit_row:
        return RedirectResponse("/reports?erro=Kit+nao+encontrado", status_code=302)
    kit = dict(kit_row)

    if nova_garagem != (kit.get("garagem") or ""):
        with db() as conn:
            conn.execute("UPDATE kit_record SET garagem = ? WHERE kit_id = ?", (nova_garagem, kit_id))
            itens_rows = conn.execute(
                "SELECT it.nome AS descricao, COUNT(*) AS quantidade "
                "FROM scan_session_items si "
                "JOIN item_tipo it ON it.id = si.item_tipo_id "
                "WHERE si.sessao_id = ? AND (si.status IS NULL OR si.status NOT IN ('movido', 'retirado')) "
                "GROUP BY si.item_tipo_id ORDER BY it.nome",
                (kit["sessao_id"],)
            ).fetchall()
        itens_label = [dict(r) for r in itens_rows]
        ts = datetime.strptime(kit["finalizado_em"], "%Y-%m-%d %H:%M:%S")

        zpl = zpl_mod.generate_zpl(
            kit_id=kit_id, kit_nome=kit["kit_nome"], cliente=kit["cliente"],
            operador=kit["operador_nome"], timestamp=ts, itens=itens_label,
            veiculo=kit.get("veiculo") or "", garagem=nova_garagem,
        )
        html_label = zpl_mod.generate_html_label(
            kit_id=kit_id, kit_nome=kit["kit_nome"], cliente=kit["cliente"],
            operador=kit["operador_nome"], timestamp=ts, itens=itens_label,
            veiculo=kit.get("veiculo") or "", garagem=nova_garagem,
        )
    else:
        with db() as conn:
            pq_row = conn.execute(
                "SELECT * FROM print_queue WHERE kit_id = ? ORDER BY id DESC LIMIT 1",
                (kit_id,)
            ).fetchone()
        if not pq_row:
            return RedirectResponse("/reports?erro=Etiqueta+nao+encontrada", status_code=302)
        pq = dict(pq_row)
        zpl = pq["zpl"]
        html_label = pq.get("html_label")

    with db() as conn:
        conn.execute(
            "INSERT INTO print_queue (kit_id, zpl, html_label, solicitado_por, solicitado_em) VALUES (?,?,?,?,?)",
            (kit_id, zpl, html_label, user["id"], now_brt())
        )
    return RedirectResponse("/print-queue?ok=reimpresso", status_code=302)


@app.get("/reports/{kit_id}/excel")
@require_permission("ver_relatorios")
async def report_excel(request: Request, kit_id: str):
    from fastapi.responses import Response as _Resp
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from io import BytesIO

    with db() as conn:
        kit = conn.execute(
            "SELECT kr.*, kt.nome AS kit_nome, kt.cliente, kt.versao, kt.tipo AS kit_tipo, "
            "u.nome AS operador_nome "
            "FROM kit_record kr "
            "JOIN kit_template kt ON kt.id = kr.kit_template_id "
            "JOIN users u ON u.id = kr.operador_id "
            "WHERE kr.kit_id = ?",
            (kit_id,)
        ).fetchone()
        if not kit:
            return PlainTextResponse("Kit não encontrado", status_code=404)
        kit = dict(kit)

        resumo = conn.execute(
            "SELECT it.nome AS tipo_nome, COUNT(*) AS quantidade "
            "FROM scan_session_items si "
            "JOIN item_tipo it ON it.id = si.item_tipo_id "
            "WHERE si.sessao_id = ? AND (si.status IS NULL OR si.status NOT IN ('movido', 'retirado')) "
            "GROUP BY si.item_tipo_id ORDER BY it.nome",
            (kit["sessao_id"],)
        ).fetchall()
        resumo = [dict(r) for r in resumo]

        itens = conn.execute(
            "SELECT it.nome AS tipo_nome, si.codigo_barra, si.serial_number, "
            "si.codigo_caixa, si.bipado_em "
            "FROM scan_session_items si "
            "JOIN item_tipo it ON it.id = si.item_tipo_id "
            "WHERE si.sessao_id = ? AND (si.status IS NULL OR si.status NOT IN ('movido', 'retirado')) "
            "ORDER BY it.nome, si.bipado_em",
            (kit["sessao_id"],)
        ).fetchall()
        itens = [dict(i) for i in itens]

    wb = openpyxl.Workbook()
    azul = "1A3A5C"
    branco = "FFFFFF"
    cinza = "F4F7FB"

    def hdr_cell(ws, row, col, value):
        c = ws.cell(row=row, column=col, value=value)
        c.font = Font(bold=True, color=branco)
        c.fill = PatternFill("solid", fgColor=azul)
        c.alignment = Alignment(horizontal="center", vertical="center")
        return c

    def meta_block(ws):
        meta = [
            ("Veículo", kit.get("veiculo") or "—"),
            ("Cliente", kit["cliente"]),
            ("Garagem", kit.get("garagem") or "—"),
            ("Kit", kit["kit_nome"]),
            ("Versão", f"v{kit['versao']}"),
            ("Operador", kit["operador_nome"]),
            ("Finalizado em", kit["finalizado_em"]),
        ]
        for r, (label, value) in enumerate(meta, 1):
            ws.cell(r, 1, label).font = Font(bold=True)
            ws.cell(r, 2, value)
        return len(meta) + 2  # blank row + next data row

    # ── Aba Resumo ──────────────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Resumo"
    next_row = meta_block(ws1)
    for col, h in enumerate(["Tipo de Item", "Quantidade Bipada"], 1):
        hdr_cell(ws1, next_row, col, h)
    for i, r in enumerate(resumo):
        row = next_row + 1 + i
        ws1.cell(row, 1, r["tipo_nome"])
        ws1.cell(row, 2, r["quantidade"])
        if i % 2 == 0:
            for col in (1, 2):
                ws1.cell(row, col).fill = PatternFill("solid", fgColor=cinza)
    ws1.column_dimensions["A"].width = 32
    ws1.column_dimensions["B"].width = 20

    # ── Aba Detalhes ────────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Detalhes")
    next_row = meta_block(ws2)
    # "Caixa (lote)" entra antes de "Origem": quem lê o relatório procura a
    # peça pelo lote quando um lote inteiro sai com defeito.
    for col, h in enumerate(["Tipo de Item", "Código de Barras", "Serial Number",
                             "Caixa (lote)", "Origem", "Bipado em"], 1):
        hdr_cell(ws2, next_row, col, h)
    for i, item in enumerate(itens):
        row = next_row + 1 + i
        codigo = item["codigo_barra"]
        if codigo.startswith("COMP:"):
            parts = codigo.split(":", 3)
            origem = "Conjunto"
            codigo_display = parts[1] if len(parts) >= 2 else codigo
        else:
            origem = "Bipagem direta"
            codigo_display = codigo
        ws2.cell(row, 1, item["tipo_nome"])
        ws2.cell(row, 2, codigo_display)
        ws2.cell(row, 3, item.get("serial_number") or "")
        ws2.cell(row, 4, item.get("codigo_caixa") or "")
        ws2.cell(row, 5, origem)
        ws2.cell(row, 6, item.get("bipado_em", ""))
        if i % 2 == 0:
            for col in (1, 2, 3, 4, 5, 6):
                ws2.cell(row, col).fill = PatternFill("solid", fgColor=cinza)
    ws2.column_dimensions["A"].width = 32
    ws2.column_dimensions["B"].width = 28
    ws2.column_dimensions["C"].width = 24
    ws2.column_dimensions["D"].width = 22
    ws2.column_dimensions["E"].width = 18
    ws2.column_dimensions["F"].width = 22

    # ── Aba Unidades do Pedido (ICCID/Telefone/CDT/ID Hardware) ────────────────
    if kit.get("kit_tipo") == "pedido":
        unidades = pedidos_mod.listar_unidades(kit["kit_template_id"])
        ws3 = wb.create_sheet("Unidades do Pedido")
        for col, h in enumerate(["ICCID", "Número de Telefone", "CDT", "ID Hardware"], 1):
            hdr_cell(ws3, 1, col, h)
        for i, u in enumerate(unidades):
            row = i + 2
            ws3.cell(row, 1, u.get("iccid") or "")
            ws3.cell(row, 2, u.get("telefone") or "")
            ws3.cell(row, 3, u.get("cdt") or "")
            ws3.cell(row, 4, u.get("id_hardware") or "")
            if i % 2 == 0:
                for col in range(1, 5):
                    ws3.cell(row, col).fill = PatternFill("solid", fgColor=cinza)
        for col, w in zip("ABCD", (24, 22, 18, 22)):
            ws3.column_dimensions[col].width = w
        ws3.freeze_panes = "A2"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    import re as _re
    safe = _re.sub(r'[^\w\-]', '_', kit["kit_nome"])
    data = (kit["finalizado_em"] or "")[:10]
    filename = f"kit_{safe}_{data}.xlsx"
    return _Resp(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.get("/reports/exportar-todos.xlsx")
@require_permission("ver_relatorios")
async def reports_exportar_todos(request: Request,
                                  data_ini: str = "",
                                  data_fim: str = "",
                                  operador_id: list[str] = Query(default=[]),
                                  tipo: list[str] = Query(default=[])):
    """Exporta todos os kits/pedidos finalizados que batem com os filtros
    atuais da tela de Relatórios (mesmos filtros — não é o kit único, é
    o lote inteiro), com uma aba de resumo e uma de itens detalhados."""
    from fastapi.responses import Response as _Resp
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from io import BytesIO

    query = """
        SELECT kr.kit_id, kr.kit_template_id, kr.finalizado_em, kr.veiculo, kr.garagem,
               COALESCE(v.numero, kr.veiculo) AS veiculo_exibido,
               kt.nome AS kit_nome, kt.cliente, kt.versao, kt.tipo AS kit_tipo,
               kr.observacao,
               u.nome AS operador_nome,
               (SELECT COUNT(*) FROM kit_validacoes kv WHERE kv.kit_id = kr.kit_id) AS num_validacoes
        FROM kit_record kr
        JOIN kit_template kt ON kt.id = kr.kit_template_id
        LEFT JOIN users u ON u.id = kr.operador_id
        LEFT JOIN veiculos v ON v.id = kr.veiculo_id
    """
    # Mesmo WHERE da tela, pela mesma função — a exportação tem que trazer
    # exatamente o que a tela mostra. O LIMIT 200 que existia aqui cortava
    # em silêncio: 250 kits no filtro viravam 200 na planilha, sem aviso.
    # Quem limita o volume é o período escolhido, não um teto escondido.
    where, params = _where_relatorio_kits(data_ini, data_fim, operador_id, tipo)
    query += " " + where + " ORDER BY kr.finalizado_em DESC, kr.kit_id"

    with db() as conn:
        kits = [dict(r) for r in conn.execute(query, params).fetchall()]
        kit_ids = [k["kit_id"] for k in kits]
        itens_por_kit = {}
        if kit_ids:
            placeholders = ",".join("?" * len(kit_ids))
            rows_itens = conn.execute(
                "SELECT kr.kit_id, it.nome AS tipo_nome, si.codigo_barra, "
                "si.serial_number, si.codigo_caixa, si.bipado_em "
                "FROM scan_session_items si "
                "JOIN item_tipo it ON it.id = si.item_tipo_id "
                "JOIN kit_record kr ON kr.sessao_id = si.sessao_id "
                f"WHERE kr.kit_id IN ({placeholders}) "
                "AND (si.status IS NULL OR si.status NOT IN ('movido', 'retirado')) "
                "ORDER BY kr.kit_id, it.nome, si.bipado_em",
                kit_ids
            ).fetchall()
            for r in rows_itens:
                itens_por_kit.setdefault(r["kit_id"], []).append(dict(r))

    wb = openpyxl.Workbook()
    azul, branco, cinza = "1A3A5C", "FFFFFF", "F4F7FB"

    def hdr_cell(ws, row, col, value):
        c = ws.cell(row=row, column=col, value=value)
        c.font = Font(bold=True, color=branco)
        c.fill = PatternFill("solid", fgColor=azul)
        c.alignment = Alignment(horizontal="center", vertical="center")
        return c

    # ── Aba Resumo ────────────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Resumo"
    for col, h in enumerate(
        ["Tipo", "Veículo", "Cliente", "Garagem", "Kit", "Operador",
         "Finalizado em", "Verificações", "Observação"], 1):
        hdr_cell(ws1, 1, col, h)
    for i, k in enumerate(kits):
        row = i + 2
        ws1.cell(row, 1, "Pedido" if k.get("kit_tipo") == "pedido" else "Kit")
        ws1.cell(row, 2, k.get("veiculo_exibido") or "")
        ws1.cell(row, 3, k["cliente"])
        ws1.cell(row, 4, k.get("garagem") or "")
        ws1.cell(row, 5, f"{k['kit_nome']} v{k['versao']}")
        ws1.cell(row, 6, k["operador_nome"])
        ws1.cell(row, 7, k.get("finalizado_em") or "")
        ws1.cell(row, 8, k.get("num_validacoes") or 0)
        ws1.cell(row, 9, k.get("observacao") or "")
        if i % 2 == 0:
            for col in range(1, 10):
                ws1.cell(row, col).fill = PatternFill("solid", fgColor=cinza)
    for col, w in zip("ABCDEFGHI", (10, 16, 22, 16, 30, 22, 20, 14, 34)):
        ws1.column_dimensions[col].width = w
    ws1.freeze_panes = "A2"

    # ── Aba Detalhes ──────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Detalhes")
    for col, h in enumerate(
        ["Kit", "Veículo", "Tipo de Item", "Código de Barras", "Serial Number",
         "Caixa (lote)", "Origem", "Bipado em"], 1):
        hdr_cell(ws2, 1, col, h)
    row = 2
    for k in kits:
        veiculo_exibido = k.get("veiculo_exibido") or ""
        kit_label = f"{k['kit_nome']} v{k['versao']} ({k['kit_id'][:8].upper()})"
        for item in itens_por_kit.get(k["kit_id"], []):
            codigo = item["codigo_barra"]
            if codigo.startswith("COMP:"):
                parts = codigo.split(":", 3)
                origem = "Conjunto"
                codigo_display = parts[1] if len(parts) >= 2 else codigo
            else:
                origem = "Bipagem direta"
                codigo_display = codigo
            ws2.cell(row, 1, kit_label)
            ws2.cell(row, 2, veiculo_exibido)
            ws2.cell(row, 3, item["tipo_nome"])
            ws2.cell(row, 4, codigo_display)
            ws2.cell(row, 5, item.get("serial_number") or "")
            ws2.cell(row, 6, item.get("codigo_caixa") or "")
            ws2.cell(row, 7, origem)
            ws2.cell(row, 8, item.get("bipado_em") or "")
            if row % 2 == 0:
                for col in range(1, 9):
                    ws2.cell(row, col).fill = PatternFill("solid", fgColor=cinza)
            row += 1
    for col, w in zip("ABCDEFGH", (34, 16, 28, 24, 20, 20, 16, 20)):
        ws2.column_dimensions[col].width = w
    ws2.freeze_panes = "A2"

    # ── Aba Unidades (ICCID/Telefone/CDT/ID Hardware dos Pedidos) ──────────────
    pedidos_no_lote = [k for k in kits if k.get("kit_tipo") == "pedido"]
    if pedidos_no_lote:
        ws3 = wb.create_sheet("Unidades")
        for col, h in enumerate(
            ["Pedido", "Veículo", "ICCID", "Número de Telefone", "CDT", "ID Hardware"], 1):
            hdr_cell(ws3, 1, col, h)
        row = 2
        for k in pedidos_no_lote:
            pedido_label = f"{k['kit_nome']} v{k['versao']} ({k['kit_id'][:8].upper()})"
            veiculo_exibido = k.get("veiculo_exibido") or ""
            for u in pedidos_mod.listar_unidades(k["kit_template_id"]):
                ws3.cell(row, 1, pedido_label)
                ws3.cell(row, 2, veiculo_exibido)
                ws3.cell(row, 3, u.get("iccid") or "")
                ws3.cell(row, 4, u.get("telefone") or "")
                ws3.cell(row, 5, u.get("cdt") or "")
                ws3.cell(row, 6, u.get("id_hardware") or "")
                if row % 2 == 0:
                    for col in range(1, 7):
                        ws3.cell(row, col).fill = PatternFill("solid", fgColor=cinza)
                row += 1
        for col, w in zip("ABCDEF", (34, 16, 24, 22, 18, 22)):
            ws3.column_dimensions[col].width = w
        ws3.freeze_panes = "A2"

    # ── Aba Sobressalentes (peças extras enviadas, mesmo período do filtro) ────
    sobressalentes = estoque_mod.listar_sobressalentes(data_ini, data_fim)
    if sobressalentes:
        ws4 = wb.create_sheet("Sobressalentes")
        for col, h in enumerate(
            ["Data", "Item", "Código", "Quantidade", "Cliente", "Enviado Por", "Observação"], 1):
            hdr_cell(ws4, 1, col, h)
        for i, r in enumerate(sobressalentes, 2):
            ws4.cell(i, 1, (r["criado_em"] or "")[:16])
            ws4.cell(i, 2, r["tipo_nome"])
            ws4.cell(i, 3, r["codigo_barra"])
            ws4.cell(i, 4, r["quantidade"])
            ws4.cell(i, 5, r["cliente"])
            ws4.cell(i, 6, r["operador_nome"] or "")
            ws4.cell(i, 7, r["observacao"] or "")
            if i % 2 == 0:
                for col in range(1, 8):
                    ws4.cell(i, col).fill = PatternFill("solid", fgColor=cinza)
        for col, w in zip("ABCDEFG", (17, 28, 18, 12, 22, 22, 40)):
            ws4.column_dimensions[col].width = w
        ws4.freeze_panes = "A2"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return _Resp(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="relatorio_kits.xlsx"'},
    )


@app.post("/reports/{kit_id}/delete")
@require_admin
async def report_delete(request: Request, kit_id: str):
    sessions_mod.deletar_kit_record(kit_id)
    return RedirectResponse("/reports?ok=excluido", status_code=302)


@app.get("/reports/validacoes", response_class=HTMLResponse)
@require_permission("ver_relatorios")
async def reports_validacoes(request: Request,
                             data_ini: str = "",
                             data_fim: str = "",
                             pagina: int = 1,
                             validador_id: list[str] = Query(default=[])):
    todas = validacoes_mod.listar_relatorio(data_ini, data_fim, validador_id)
    # Pagina em vez de cortar: o LIMIT 500 que existia na consulta escondia
    # o excedente sem dizer nada. Agora tudo que bate com o filtro é
    # alcançável, e a contagem mostra o total de verdade.
    pag = paginacao_mod.paginar(todas, pagina)
    with db() as conn:
        usuarios = conn.execute("SELECT id, nome FROM users ORDER BY nome").fetchall()
    return render(request, "reports_validacoes.html", {
        "voltar_para": _voltar_para(request, "/reports"),
        "rows": pag["itens"],
        "pag": pag,
        "pagina": pag["pagina"],
        "usuarios": [dict(u) for u in usuarios],
        "opcoes_usuarios": [(str(u["id"]), u["nome"]) for u in usuarios],
        "data_ini": data_ini,
        "data_fim": data_fim,
        "validador_id": validador_id,
    })


@app.get("/reports/validacoes/export")
@require_permission("ver_relatorios")
async def reports_validacoes_export(request: Request,
                                    data_ini: str = "",
                                    data_fim: str = "",
                                    validador_id: list[str] = Query(default=[])):
    """Uma linha por kit (não por verificação) — se o mesmo kit foi
    verificado mais de uma vez, cada verificação vira um bloco extra de
    colunas (Verificação 1, Verificação 2...) na mesma linha, em vez de
    duplicar veículo/cliente/itens numa linha nova por verificação."""
    from fastapi.responses import Response as _Resp
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from io import BytesIO

    grupos = validacoes_mod.listar_relatorio_agrupado(data_ini, data_fim, validador_id)
    max_verificacoes = max((len(g["verificacoes"]) for g in grupos), default=0)

    azul = "1A3A5C"
    branco = "FFFFFF"
    cinza = "F4F7FB"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Verificado"

    headers = [
        "Kit ID", "Template", "Cliente", "Veículo", "Garagem",
        "Operador Conferência", "Data Conferência", "Itens"
    ]
    widths = [14, 28, 22, 14, 14, 22, 20, 50]
    for n in range(1, max_verificacoes + 1):
        headers += [f"Verificação {n} - Por", f"Verificação {n} - Data", f"Verificação {n} - Observação"]
        widths += [22, 20, 30]

    for col, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = Font(bold=True, color=branco)
        c.fill = PatternFill("solid", fgColor=azul)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[ws.cell(1, col).column_letter].width = w

    for i, g in enumerate(grupos, 2):
        ws.cell(i, 1, g["kit_id"][:8].upper())
        ws.cell(i, 2, g["kit_nome"])
        ws.cell(i, 3, g["cliente"])
        ws.cell(i, 4, g.get("veiculo") or "")
        ws.cell(i, 5, g.get("garagem") or "")
        ws.cell(i, 6, g["operador_nome"])
        ws.cell(i, 7, g.get("finalizado_em") or "")
        itens_texto = (g.get("itens_resumo") or "").replace(" | ", "\n")
        c_itens = ws.cell(i, 8, itens_texto)
        c_itens.alignment = Alignment(wrap_text=True, vertical="top")
        col = 9
        for v in g["verificacoes"]:
            ws.cell(i, col, v["validado_por_nome"])
            ws.cell(i, col + 1, v["validado_em"])
            ws.cell(i, col + 2, v.get("observacao") or "")
            col += 3
        if i % 2 == 0:
            for c in range(1, len(headers) + 1):
                ws.cell(i, c).fill = PatternFill("solid", fgColor=cinza)

    ws.freeze_panes = "A2"
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return _Resp(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=verificacoes.xlsx"},
    )


@app.get("/reports/sobressalentes", response_class=HTMLResponse)
@require_permission("ver_relatorios")
async def reports_sobressalentes(request: Request,
                                 data_ini: str = "",
                                 data_fim: str = "",
                                 cliente: list[str] = Query(default=[])):
    rows = estoque_mod.listar_sobressalentes(data_ini, data_fim, cliente)
    return render(request, "reports_sobressalentes.html", {
        "voltar_para": _voltar_para(request, "/reports"),
        "rows": rows,
        "clientes": clientes_mod.listar(),
        "data_ini": data_ini,
        "data_fim": data_fim,
        "cliente": cliente,
    })


@app.get("/reports/sobressalentes/export")
@require_permission("ver_relatorios")
async def reports_sobressalentes_export(request: Request,
                                        data_ini: str = "",
                                        data_fim: str = "",
                                        cliente: list[str] = Query(default=[])):
    from fastapi.responses import Response as _Resp
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from io import BytesIO

    rows = estoque_mod.listar_sobressalentes(data_ini, data_fim, cliente)

    azul, branco = "1A3A5C", "FFFFFF"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sobressalentes"

    headers = ["Data", "Item", "Código", "Quantidade", "Cliente", "Enviado Por", "Observação"]
    widths = [17, 28, 18, 12, 22, 22, 40]
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = Font(bold=True, color=branco)
        c.fill = PatternFill("solid", fgColor=azul)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[ws.cell(1, col).column_letter].width = w

    for i, r in enumerate(rows, 2):
        ws.cell(i, 1, (r["criado_em"] or "")[:16])
        ws.cell(i, 2, r["tipo_nome"])
        ws.cell(i, 3, r["codigo_barra"])
        ws.cell(i, 4, r["quantidade"])
        ws.cell(i, 5, r["cliente"])
        ws.cell(i, 6, r["operador_nome"] or "")
        ws.cell(i, 7, r["observacao"] or "")

    ws.freeze_panes = "A2"
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return _Resp(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=sobressalentes.xlsx"},
    )


# ── Prateleira ────────────────────────────────────────────────────────────────

def _prateleira_context() -> dict:
    layout = prateleira_mod.get_layout()
    blocos = prateleira_mod.listar_blocos()
    return {
        "layout": layout,
        "colunas_nomes": prateleira_mod.listar_colunas(),
        "blocos": blocos,
        "celulas_vazias": prateleira_mod.celulas_vazias(blocos, layout),
        "livre": prateleira_mod.listar_livre(),
        "estoque_itens": estoque_mod.listar_estoque(),
        "max_itens_por_slot": prateleira_mod.MAX_ITENS_POR_SLOT,
    }


@app.get("/admin/prateleira", response_class=HTMLResponse)
@require_login
async def admin_prateleira(request: Request):
    return render(request, "admin_prateleira.html", _prateleira_context())


@app.post("/admin/prateleira/layout")
@require_login
async def admin_prateleira_layout(request: Request):
    form = await request.form()
    try:
        linhas = int(form.get("linhas"))
        colunas = int(form.get("colunas"))
        nomes = [str(form.get(f"nome_coluna_{i}", "")).strip() for i in range(1, colunas + 1)]
        prateleira_mod.atualizar_layout(linhas, colunas, nomes)
    except (ValueError, TypeError) as e:
        return RedirectResponse("/admin/prateleira?erro=" + quote(str(e)), status_code=302)
    return RedirectResponse("/admin/prateleira?ok=layout", status_code=302)


@app.post("/admin/prateleira/blocos")
@require_login
async def admin_prateleira_criar_bloco(request: Request):
    form = await request.form()
    try:
        linha_ini = int(form.get("linha_ini"))
        linha_fim = int(form.get("linha_fim"))
        coluna_ini = int(form.get("coluna_ini"))
        coluna_fim = int(form.get("coluna_fim"))
        estoque_id = int(form.get("estoque_id"))
        prateleira_mod.criar_bloco(linha_ini, linha_fim, coluna_ini, coluna_fim, estoque_id)
    except (ValueError, TypeError) as e:
        return RedirectResponse("/admin/prateleira?erro=" + quote(str(e)), status_code=302)
    return RedirectResponse("/admin/prateleira?ok=bloco", status_code=302)


@app.post("/admin/prateleira/blocos/{bloco_id}/remover")
@require_admin
async def admin_prateleira_remover_bloco(request: Request, bloco_id: int):
    prateleira_mod.remover_bloco(bloco_id)
    return RedirectResponse("/admin/prateleira", status_code=302)


@app.post("/admin/prateleira/livre")
@require_login
async def admin_prateleira_adicionar_livre(request: Request):
    form = await request.form()
    try:
        prateleira_mod.adicionar_livre(int(form.get("estoque_id")))
    except (ValueError, TypeError) as e:
        return RedirectResponse("/admin/prateleira?erro=" + quote(str(e)), status_code=302)
    return RedirectResponse("/admin/prateleira?ok=livre", status_code=302)


@app.post("/admin/prateleira/livre/{livre_id}/remover")
@require_admin
async def admin_prateleira_remover_livre(request: Request, livre_id: int):
    prateleira_mod.remover_livre(livre_id)
    return RedirectResponse("/admin/prateleira", status_code=302)


@app.get("/prateleira/tv", response_class=HTMLResponse)
async def prateleira_tv(request: Request, minutos: int = 5):
    minutos = max(1, minutos)
    layout = prateleira_mod.get_layout()
    blocos = prateleira_mod.listar_blocos()
    livre = prateleira_mod.listar_livre()
    return render(request, "prateleira_tv.html", {
        "layout": layout,
        "colunas_nomes": prateleira_mod.listar_colunas(),
        "blocos": blocos,
        "celulas_vazias": prateleira_mod.celulas_vazias(blocos, layout),
        "livre": livre,
        "minutos": minutos,
        "contagem_status": prateleira_mod.contar_status(blocos, livre),
    })


# ── Produção (Consat → Trânsito → Cliente) ────────────────────────────────────

_PRODUCAO_MOBILE_LIMITE = 40


@app.get("/producao/mobile", response_class=HTMLResponse)
@require_login
async def producao_mobile(request: Request):
    """A esteira no celular, só pra olhar.

    A tela de computador tem seleção em massa, nota fiscal e botões de
    estágio — coisa de mesa, que no telefone vira toque errado. Aqui vão os
    números e as listas; quem precisa mover kit continua no computador.

    Cada etapa mostra até 40 linhas e diz quando cortou, em vez de deixar o
    resto sumir sem aviso."""
    def _linhas(itens, veiculo, sub, quando, destino=None):
        cortadas = itens[:_PRODUCAO_MOBILE_LIMITE]
        return [{
            "veiculo": veiculo(i) or "—",
            "sub": sub(i),
            "quando": (quando(i) or "")[:16],
            "destino": destino(i) if destino else None,
        } for i in cortadas]

    def _cli_gar(i):
        partes = [i.get("cliente") or "", i.get("garagem") or ""]
        return " · ".join(p for p in partes if p)

    a_produzir = producao_mod.listar_a_produzir()
    em_producao = producao_mod.listar_em_producao()
    produzido = producao_mod.listar_produzido()
    transito = producao_mod.listar_transito()
    no_cliente = producao_mod.listar_no_cliente()

    etapas = [
        {"titulo": "Kits a produzir", "total": len(a_produzir),
         "vazio": "Nada pronto pra produzir agora.",
         "linhas": _linhas(a_produzir, lambda i: i.get("veiculo"),
                           lambda i: _cli_gar(i) + (" · " + (i.get("modelo") or "") if i.get("modelo") else ""),
                           lambda i: "")},
        {"titulo": "Em produção", "total": len(em_producao),
         "vazio": "Nenhuma bipagem em andamento.",
         # Única ação que faz sentido no celular: continuar a bipagem.
         "linhas": _linhas(em_producao, lambda i: i.get("veiculo"),
                           lambda i: _cli_gar(i) + " · " + (i.get("operador_nome") or ""),
                           lambda i: i.get("iniciado_em"),
                           lambda i: "/session/%s" % i["sessao_id"])},
        {"titulo": "Produzido", "total": len(produzido),
         "vazio": "Nada esperando envio.",
         "linhas": _linhas(produzido, lambda i: i.get("veiculo") or i.get("kit_nome"),
                           _cli_gar, lambda i: i.get("finalizado_em"),
                           lambda i: "/kit/%s" % i["kit_id"])},
        {"titulo": "Em trânsito", "total": len(transito),
         "vazio": "Nada em trânsito.",
         "linhas": _linhas(transito, lambda i: i.get("veiculo") or i.get("kit_nome"),
                           _cli_gar, lambda i: i.get("transito_em"),
                           lambda i: "/kit/%s" % i["kit_id"])},
        {"titulo": "No cliente", "total": len(no_cliente),
         "vazio": "Nenhum kit no cliente.",
         "linhas": _linhas(
             no_cliente,
             lambda i: i.get("veiculo") or i.get("kit_nome"),
             lambda i: _cli_gar(i) + (" · concluído"
                                      if i.get("status_producao") == "cliente_concluido"
                                      else " · instalando"),
             lambda i: i.get("chegou_em"),
             lambda i: "/kit/%s" % i["kit_id"])},
    ]
    resumo = producao_mod.resumo()
    resumo["total_kits"] = (resumo["produzido"] + resumo["transito"] + resumo["cliente"])
    return render(request, "producao_mobile.html", {"resumo": resumo, "etapas": etapas})


@app.get("/producao/tv", response_class=HTMLResponse)
async def producao_tv(request: Request, minutos: int = 5):
    minutos = max(1, minutos)
    return render(request, "producao_tv.html", {
        **producao_mod.dados_tv(),
        "resumo": producao_mod.resumo(),
        "minutos": minutos,
    })


# Coluna "Veículo" cai pro nome do kit quando o veículo está vazio (mesma
# regra que o template já usa pra exibir a linha) — por isso é uma tupla de
# fallback, e não um campo só. Cada prefixo é o mesmo já usado no pag_XX de
# cada etapa, pra clicar em ordenar não desarrumar a paginação dela.
_COL_VEICULO = ("veiculo", "kit_nome")
_ORDENS_PRODUCAO = {
    "ap":  {"veiculo": "veiculo", "garagem": "garagem", "cliente": "cliente", "modelo": "modelo"},
    "pr":  {"veiculo": _COL_VEICULO, "garagem": "garagem", "cliente": "cliente", "quando": "finalizado_em"},
    "tr":  {"veiculo": _COL_VEICULO, "garagem": "garagem", "cliente": "cliente", "quando": "transito_em"},
    "cli": {"veiculo": _COL_VEICULO, "garagem": "garagem", "cliente": "cliente",
            "modelo": "modelo", "quando": "chegou_em"},
}


def _ler_ordem(request: Request, prefixo: str) -> tuple[str, str]:
    """Lê ord_<prefixo>/dir_<prefixo> da URL — coluna clicada e sentido.
    Coluna que não existe no mapa dessa etapa é ignorada (lista fica na
    ordem de sempre), então um link/URL velho nunca quebra a tela."""
    campo = request.query_params.get(f"ord_{prefixo}", "")
    if campo not in _ORDENS_PRODUCAO[prefixo]:
        campo = ""
    direcao = request.query_params.get(f"dir_{prefixo}", "asc")
    if direcao not in ("asc", "desc"):
        direcao = "asc"
    return campo, direcao


@app.get("/admin/producao", response_class=HTMLResponse)
@require_login
async def admin_producao(request: Request):
    # "Kits possíveis" NÃO é conta nova: é a mesma autonomia que a lista de
    # Kits Cadastrados (Criar Kit/Pedido) mostra, vinda de consumo_mod. Uma
    # chamada só devolve o mapa de TODOS os templates — nada de uma consulta
    # por linha da tabela.
    autonomia = consumo_mod.resumo_todos_kits()

    def _com_autonomia(linhas):
        for l in linhas:
            info = autonomia.get(l.get("kit_template_id")) or {}
            l["kits_possiveis"] = info.get("autonomia_kit")
            l["gargalo"] = info.get("gargalo")
        return linhas

    # Cada etapa pagina SOZINHA: abrir a página 3 de "Produzido" não pode
    # mexer em "Em Trânsito". Sem paginação nenhuma, com 2.500 kits a tela
    # saía com 2,4 MB e 1.800 linhas de uma vez — no tablet do galpão isso
    # é meio minuto de espera pra ver os cinco primeiros.
    pag_ap = max(1, int(request.query_params.get("pag_ap", 1) or 1))
    pag_pr = max(1, int(request.query_params.get("pag_pr", 1) or 1))
    pag_tr = max(1, int(request.query_params.get("pag_tr", 1) or 1))

    # Clicar no cabeçalho da coluna só reordena a lista antes de paginar —
    # é apoio de busca visual, não filtro: nada some, só muda a ordem.
    ord_ap, dir_ap = _ler_ordem(request, "ap")
    ord_pr, dir_pr = _ler_ordem(request, "pr")
    ord_tr, dir_tr = _ler_ordem(request, "tr")
    ord_cli, dir_cli = _ler_ordem(request, "cli")

    lista_ap = paginacao_mod.ordenar(
        _com_autonomia(producao_mod.listar_a_produzir()),
        _ORDENS_PRODUCAO["ap"].get(ord_ap), dir_ap)
    lista_pr = paginacao_mod.ordenar(
        producao_mod.listar_produzido(), _ORDENS_PRODUCAO["pr"].get(ord_pr), dir_pr)
    lista_tr = paginacao_mod.ordenar(
        producao_mod.listar_transito(), _ORDENS_PRODUCAO["tr"].get(ord_tr), dir_tr)
    lista_cli = paginacao_mod.ordenar(
        producao_mod.listar_no_cliente(limite=30), _ORDENS_PRODUCAO["cli"].get(ord_cli), dir_cli)

    return render(request, "admin_producao.html", {
        "a_produzir": paginacao_mod.paginar(lista_ap, pag_ap),
        "em_producao": _com_autonomia(producao_mod.listar_em_producao()),
        "produzido": paginacao_mod.paginar(lista_pr, pag_pr),
        "transito": paginacao_mod.paginar(lista_tr, pag_tr),
        # Card único de Cliente: instalando + concluído na mesma lista. As
        # duas funções antigas seguem existindo (o Painel da TV usa cada
        # coluna separada), então nada foi perdido.
        "no_cliente": lista_cli,
        "resumo": producao_mod.resumo(),
        # Kits devendo item (patrimônio movido pra outro veículo ou retirado):
        # a esteira é onde o kit espera, então é aqui que a pendência precisa
        # aparecer pra alguém repor antes de despachar.
        "faltando_item": sessions_mod.kits_incompletos(),
        "tv_config": producao_mod.get_tv_config(),
        # TODAS as remessas abertas: agora pode haver mais de uma (duas
        # frentes de trabalho), e mostrar só a última esconderia a outra.
        "remessas_abertas": remessas_mod.listar_abertas(),
        "clientes_cadastrados": clientes_mod.listar(),
        "ok": request.query_params.get("ok", ""),
        "ord_ap": ord_ap, "dir_ap": dir_ap,
        "ord_pr": ord_pr, "dir_pr": dir_pr,
        "ord_tr": ord_tr, "dir_tr": dir_tr,
        "ord_cli": ord_cli, "dir_cli": dir_cli,
    })


# ── Remessas (o lote de envio: "mandar 90") ───────────────────────────────────
# Ficam sob /admin/producao de propósito: a permissão de tela da Produção já
# vale pra elas, sem precisar de uma chave nova só pra ver a mesma esteira.

@app.get("/admin/producao/remessas", response_class=HTMLResponse)
@require_login
async def admin_remessas(request: Request, busca: str = "", remessa_id: int = 0,
                         etapa: list = Query(default=[]),
                         imp_ini: str = "", imp_fim: str = "",
                         arquivadas: int = 0):
    """`remessa_id` diz qual remessa esta sendo montada a mao — sem ele, com
    varias abertas, nao da pra saber pra qual a lista de candidatos aponta."""
    abertas = remessas_mod.listar_abertas()
    alvo = remessas_mod.listar_uma(remessa_id) if remessa_id else (abertas[0] if abertas else None)
    alvo = alvo if (alvo and alvo["status"] == "aberta") else None
    return render(request, "admin_remessas.html", {
        "remessas": remessas_mod.listar(incluir_arquivadas=bool(arquivadas)),
        "abertas": abertas,
        "alvo": alvo,
        # Candidatos de TODAS as etapas — o lote pode ser montado antes de o
        # galpao comecar a bipar.
        "candidatos": (remessas_mod.candidatos(busca, list(etapa), imp_ini, imp_fim)
                       if abertas else []),
        "kits_do_alvo": remessas_mod.kits_da_remessa(alvo["id"]) if alvo else [],
        "etapas": remessas_mod.ETAPAS,
        "filtro_etapa": list(etapa),
        "busca": busca,
        "imp_ini": imp_ini,
        "imp_fim": imp_fim,
        "arquivadas": bool(arquivadas),
        "clientes_cadastrados": clientes_mod.listar(),
        "voltar_para": _voltar_para(request, "/admin/producao"),
    })


def _volta_remessa(remessa_id: int, extra: str = "") -> str:
    base = f"/admin/producao/remessas?remessa_id={remessa_id}"
    return base + ("&" + extra if extra else "")


@app.post("/admin/producao/remessas/{remessa_id}/editar")
@require_permission("remessas_gerenciar")
async def admin_remessa_editar(request: Request, remessa_id: int):
    """Corrige nome, quantidade e cliente — em QUALQUER status. Nada aqui
    toca em producao, bipagem, patrimonio ou estoque."""
    form = await request.form()
    try:
        remessas_mod.editar(remessa_id, str(form.get("nome", "")),
                            form.get("alvo", ""), form.get("cliente"))
    except ValueError as e:
        return RedirectResponse(_volta_remessa(remessa_id, "erro=" + quote(str(e))),
                                status_code=302)
    return RedirectResponse(_volta_remessa(remessa_id, "ok=editada"), status_code=302)


@app.post("/admin/producao/remessas/{remessa_id}/reabrir")
@require_permission("remessas_gerenciar")
async def admin_remessa_reabrir(request: Request, remessa_id: int):
    try:
        r = remessas_mod.reabrir(remessa_id)
    except ValueError as e:
        return RedirectResponse(_volta_remessa(remessa_id, "erro=" + quote(str(e))),
                                status_code=302)
    request.state.auditoria_detalhe = f"Remessa \"{r['nome']}\" reaberta"
    return RedirectResponse(_volta_remessa(remessa_id, "ok=reaberta"), status_code=302)


@app.post("/admin/producao/remessas/{remessa_id}/arquivar")
@require_permission("remessas_gerenciar")
async def admin_remessa_arquivar(request: Request, remessa_id: int):
    """Tira do caminho SEM apagar: o historico do envio continua inteiro."""
    form = await request.form()
    motivo = str(form.get("motivo", ""))
    try:
        r = remessas_mod.arquivar(remessa_id, motivo)
    except ValueError as e:
        return RedirectResponse("/admin/producao/remessas?erro=" + quote(str(e)),
                                status_code=302)
    request.state.auditoria_detalhe = (
        f"Remessa \"{r['nome']}\" arquivada" + (f" (motivo: {motivo})" if motivo else ""))
    return RedirectResponse("/admin/producao/remessas?ok=arquivada", status_code=302)


@app.post("/admin/producao/remessas/{remessa_id}/excluir")
@require_permission("remessas_gerenciar")
async def admin_remessa_excluir(request: Request, remessa_id: int):
    """So apaga remessa VAZIA. Com itens, recusa e manda arquivar — apagar
    levaria junto o registro de que aqueles kits foram enviados."""
    try:
        r = remessas_mod.excluir(remessa_id)
    except ValueError as e:
        return RedirectResponse("/admin/producao/remessas?erro=" + quote(str(e)),
                                status_code=302)
    request.state.auditoria_detalhe = f"Remessa \"{r['nome']}\" excluída"
    return RedirectResponse("/admin/producao/remessas?ok=excluida&n=" + quote(r["nome"]),
                            status_code=302)


@app.post("/admin/producao/remessas/transferir")
@require_permission("remessas_gerenciar")
async def admin_remessa_transferir(request: Request):
    """Move um veiculo de uma remessa pra outra numa operacao so — sem passar
    por "tirar de uma" e "por na outra", onde da pra esquecer o segundo passo."""
    form = await request.form()
    ref = str(form.get("ref", ""))
    destino = int(form.get("destino_id") or 0)
    origem = int(form.get("remessa_id") or 0)
    try:
        remessas_mod.transferir(ref, destino, (get_current_user(request) or {}).get("id"))
    except ValueError as e:
        return RedirectResponse(_volta_remessa(origem, "erro=" + quote(str(e))),
                                status_code=302)
    nome_destino = (remessas_mod.listar_uma(destino) or {}).get("nome", destino)
    request.state.auditoria_detalhe = (
        f"{remessas_mod.rotulo_do_ref(ref)} movido para a remessa \"{nome_destino}\"")
    return RedirectResponse(_volta_remessa(destino, "ok=transferido"), status_code=302)


@app.post("/admin/producao/remessas/{remessa_id}/adicionar")
@require_permission("remessas_gerenciar")
async def admin_remessa_adicionar(request: Request, remessa_id: int):
    """Acrescenta a mao os veiculos selecionados — de qualquer etapa."""
    form = await request.form()
    refs = [r for r in form.getlist("ref") if str(r).strip()]
    # Compatibilidade com o formulario antigo, que mandava kit_id puro.
    refs += ["kit:" + k for k in form.getlist("kit_id") if str(k).strip()]
    filtros = "&".join(
        f"{c}={quote(str(form.get(c, '')))}" for c in ("busca", "imp_ini", "imp_fim")
        if str(form.get(c, "")).strip())
    if not refs:
        return RedirectResponse(
            _volta_remessa(remessa_id, (filtros + "&" if filtros else "") + "erro=" + quote(
                "Selecione ao menos um veiculo para acrescentar.")), status_code=302)
    try:
        r = remessas_mod.adicionar_itens(remessa_id, refs,
                                         (get_current_user(request) or {}).get("id"))
    except ValueError as e:
        return RedirectResponse(
            _volta_remessa(remessa_id, (filtros + "&" if filtros else "") + "erro=" + quote(str(e))),
            status_code=302)
    partes = [f"ok=add&n={r['entraram']}"]
    if r["ja_em_outra"]:
        partes.append(f"ja={r['ja_em_outra']}")
    if r["cliente_errado"]:
        partes.append(f"outro_cliente={r['cliente_errado']}&cliente={quote(r['cliente'])}")
    if r["fechou"]:
        partes.append("fechada=1")
    nome_remessa = (remessas_mod.listar_uma(remessa_id) or {}).get("nome", remessa_id)
    request.state.auditoria_detalhe = (
        f"{r['entraram']} veículo(s) acrescentado(s) à remessa \"{nome_remessa}\"")
    return RedirectResponse(
        _volta_remessa(remessa_id, (filtros + "&" if filtros else "") + "&".join(partes)),
        status_code=302)


@app.post("/admin/producao/remessas/{remessa_id}/remover")
@require_permission("remessas_gerenciar")
async def admin_remessa_remover(request: Request, remessa_id: int):
    form = await request.form()
    ref = str(form.get("ref", "")) or ("kit:" + str(form.get("kit_id", "")))
    rotulo = remessas_mod.rotulo_do_ref(ref)
    try:
        remessas_mod.remover_item(remessa_id, ref)
    except ValueError as e:
        return RedirectResponse(_volta_remessa(remessa_id, "erro=" + quote(str(e))),
                                status_code=302)
    nome_remessa = (remessas_mod.listar_uma(remessa_id) or {}).get("nome", remessa_id)
    request.state.auditoria_detalhe = f"{rotulo} removido da remessa \"{nome_remessa}\""
    return RedirectResponse(_volta_remessa(remessa_id, "ok=removido"), status_code=302)


@app.post("/admin/producao/remessas/{remessa_id}/selecionados")
@require_permission("remessas_gerenciar")
async def admin_remessa_selecionados(request: Request, remessa_id: int):
    """Um formulário, duas ações: transferir os marcados pra outra remessa ou
    tirá-los desta.

    Ficam juntos porque a seleção é a mesma — separar em dois formulários
    obrigaria a remarcar tudo quando o operador mudasse de ideia entre
    "mover" e "tirar"."""
    form = await request.form()
    refs = [r for r in form.getlist("ref") if str(r).strip()]
    acao = str(form.get("acao", ""))
    if not refs:
        return RedirectResponse(_volta_remessa(remessa_id, "erro=" + quote(
            "Marque ao menos um veículo.")), status_code=302)
    try:
        if acao == "transferir":
            r = remessas_mod.transferir_varios(
                refs, int(form.get("destino_id") or 0),
                (get_current_user(request) or {}).get("id"))
            destino = r["destino"]["id"] if r.get("destino") else remessa_id
            extra = f"ok=movidos&n={r['movidos']}"
            if r["recusados"]:
                extra += "&recusados=" + quote(" · ".join(r["recusados"][:4]))
            nome_destino = (r["destino"] or {}).get("nome", destino)
            request.state.auditoria_detalhe = (
                f"{r['movidos']} veículo(s) movido(s) para a remessa \"{nome_destino}\"")
            return RedirectResponse(_volta_remessa(destino, extra), status_code=302)
        r = remessas_mod.remover_varios(remessa_id, refs)
        extra = f"ok=tirados&n={r['tirados']}"
        if r["recusados"]:
            extra += "&recusados=" + quote(" · ".join(r["recusados"][:4]))
        nome_remessa = (remessas_mod.listar_uma(remessa_id) or {}).get("nome", remessa_id)
        request.state.auditoria_detalhe = (
            f"{r['tirados']} veículo(s) removido(s) da remessa \"{nome_remessa}\"")
        return RedirectResponse(_volta_remessa(remessa_id, extra), status_code=302)
    except ValueError as e:
        return RedirectResponse(_volta_remessa(remessa_id, "erro=" + quote(str(e))),
                                status_code=302)


@app.post("/admin/producao/remessas/abrir")
@require_permission("remessas_gerenciar")
async def admin_remessa_abrir(request: Request):
    form = await request.form()
    user = get_current_user(request)
    cliente = str(form.get("cliente", ""))
    try:
        r = remessas_mod.abrir(str(form.get("nome", "")), form.get("alvo", ""),
                               cliente, user["id"] if user else None)
    except ValueError as e:
        return RedirectResponse("/admin/producao/remessas?erro=" + quote(str(e)), status_code=302)
    request.state.auditoria_detalhe = (
        f"Remessa \"{r['nome']}\" aberta (alvo {r['alvo']}"
        + (f", cliente {cliente}" if cliente else "") + ")")
    return RedirectResponse("/admin/producao/remessas?ok=aberta", status_code=302)


@app.post("/admin/producao/remessas/{remessa_id}/alvo")
@require_permission("remessas_gerenciar")
async def admin_remessa_alvo(request: Request, remessa_id: int):
    """Muda a quantidade no meio do caminho — o "0/90 → 80/80"."""
    form = await request.form()
    voltar = str(form.get("voltar", "")) or "/admin/producao/remessas"
    antes = remessas_mod.listar_uma(remessa_id)
    try:
        r = remessas_mod.definir_alvo(remessa_id, form.get("alvo", ""))
    except ValueError as e:
        return RedirectResponse(voltar + "?erro=" + quote(str(e)), status_code=302)
    fechou = "&fechada=1" if r and r["status"] == "fechada" else ""
    nome = (r or antes or {}).get("nome", remessa_id)
    alvo_antes = (antes or {}).get("alvo")
    request.state.auditoria_detalhe = (
        f"Remessa \"{nome}\": alvo alterado de {alvo_antes} para {r['alvo']}"
        if r else f"Remessa \"{nome}\": alvo alterado")
    return RedirectResponse(f"{voltar}?ok=alvo{fechou}", status_code=302)


def _planilha_kits(titulo: str, subtitulo: str, linhas: list[dict],
                   colunas: list[tuple]) -> bytes:
    """Uma planilha de acompanhamento, do mesmo jeito nas três telas.

    Existe pra as três não divergirem: quem exporta Produção, Trânsito e a
    Remessa espera a mesma cara e as mesmas colunas — e três montagens
    separadas vão ficando diferentes commit a commit."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from io import BytesIO

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = titulo[:31]
    azul, branco, cinza = "1A3A5C", "FFFFFF", "F4F7FB"

    ws.cell(1, 1, titulo).font = Font(bold=True, size=14, color=azul)
    ws.cell(2, 1, subtitulo).font = Font(size=10, color="7A8794")
    ws.cell(3, 1, f"Gerado em {now_brt()}").font = Font(size=9, color="9AA7B4")

    cab = 5
    for i, (rotulo, _chave, _larg) in enumerate(colunas, 1):
        c = ws.cell(cab, i, rotulo)
        c.font = Font(bold=True, color=branco)
        c.fill = PatternFill("solid", fgColor=azul)
        c.alignment = Alignment(horizontal="center", vertical="center")
    for n, linha in enumerate(linhas):
        r = cab + 1 + n
        for i, (_rotulo, chave, _larg) in enumerate(colunas, 1):
            ws.cell(r, i, linha.get(chave) or "")
        if n % 2 == 0:
            for i in range(1, len(colunas) + 1):
                ws.cell(r, i).fill = PatternFill("solid", fgColor=cinza)
    for i, (_rotulo, _chave, larg) in enumerate(colunas, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = larg
    ws.freeze_panes = ws.cell(cab + 1, 1)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# Colunas do acompanhamento — as mesmas nas três planilhas.
_COLS_ACOMP = [
    ("Veículo", "veiculo", 16), ("Cliente", "cliente", 24), ("Garagem", "garagem", 20),
    ("Kit", "kit_nome", 30), ("Modelo", "modelo", 26),
    ("Nota fiscal", "nota_fiscal", 16), ("Data da NF", "nota_fiscal_data", 14),
    ("Operador", "operador_nome", 22),
    ("Montado em", "finalizado_em", 20), ("Em trânsito desde", "transito_em", 20),
]

# Etapas que ainda não viraram kit fechado: sem NF, sem datas de esteira.
_COLS_FILA = [
    ("Veículo", "veiculo", 16), ("Cliente", "cliente", 24), ("Garagem", "garagem", 20),
    ("Kit", "kit_nome", 30), ("Modelo", "modelo", 26),
    ("Operador", "operador_nome", 22), ("Começou em", "iniciado_em", 20),
]


@app.get("/admin/producao/excel")
@require_login
async def admin_producao_excel(request: Request, etapa: str = "transito"):
    """Acompanhamento de uma etapa da esteira em Excel — é o que se manda
    pro cliente ou pro time de campo sem dar acesso ao sistema."""
    from fastapi.responses import Response as _Resp
    # Cada etapa exporta as colunas que ela REALMENTE tem: "a produzir" e
    # "em produção" não têm nota fiscal nem data de trânsito, e colunas
    # sempre vazias fazem quem recebe a planilha achar que faltou dado.
    etapas = {
        "produzido": ("Produzido", producao_mod.listar_produzido, _COLS_ACOMP),
        "transito": ("Em trânsito", producao_mod.listar_transito, _COLS_ACOMP),
        "em_producao": ("Em produção", producao_mod.listar_em_producao, _COLS_FILA),
        "a_produzir": ("A produzir", producao_mod.listar_a_produzir, _COLS_FILA),
        # limite=None: a tela mostra só os últimos 30, mas a planilha exporta
        # tudo que já chegou no cliente — é justamente pra ver o que não
        # coube na tela que alguém baixa essa aba.
        "cliente": ("No cliente", lambda: producao_mod.listar_no_cliente(limite=None), _COLS_ACOMP),
    }
    rotulo, fonte, colunas = etapas.get(etapa, etapas["transito"])
    linhas = fonte()
    corpo = _planilha_kits(rotulo, f"{len(linhas)} kit(s) nesta etapa", linhas, colunas)
    nome = f"producao-{etapa}-{now_brt()[:10]}.xlsx"
    return _Resp(content=corpo,
                 media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                 headers={"Content-Disposition": f'attachment; filename="{nome}"'})


@app.get("/admin/producao/remessas/{remessa_id}/excel")
@require_login
async def admin_remessa_excel(request: Request, remessa_id: int):
    from fastapi.responses import Response as _Resp
    r = remessas_mod.listar_uma(remessa_id)
    if not r:
        return PlainTextResponse("Remessa não encontrada", status_code=404)
    linhas = remessas_mod.kits_da_remessa(remessa_id)
    sub = (f"{r['rotulo']} enviados"
           + (f" · cliente {r['cliente']}" if r["cliente"] else " · todos os clientes")
           + (" · FECHADA em " + (r["fechada_em"] or "")[:16] if r["status"] == "fechada"
              else " · em andamento"))
    colunas = [("Entrou na remessa", "entrou_em", 20)] + _COLS_ACOMP
    corpo = _planilha_kits(r["nome"], sub, linhas, colunas)
    nome = f"{r['nome'].replace(' ', '-').lower()}-{now_brt()[:10]}.xlsx"
    return _Resp(content=corpo,
                 media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                 headers={"Content-Disposition": f'attachment; filename="{nome}"'})


@app.post("/admin/producao/remessas/{remessa_id}/fechar")
@require_permission("remessas_gerenciar")
async def admin_remessa_fechar(request: Request, remessa_id: int):
    form = await request.form()
    motivo = str(form.get("motivo", "")).strip() or "fechada à mão"
    nome = (remessas_mod.listar_uma(remessa_id) or {}).get("nome", remessa_id)
    remessas_mod.fechar(remessa_id, motivo=motivo)
    request.state.auditoria_detalhe = f"Remessa \"{nome}\" fechada ({motivo})"
    return RedirectResponse("/admin/producao/remessas?ok=fechada", status_code=302)


@app.post("/admin/producao/tv-config")
@require_admin
async def admin_producao_tv_config(request: Request):
    form = await request.form()
    producao_mod.salvar_tv_config(dict(form))
    return RedirectResponse("/admin/producao?ok=tv_config", status_code=302)


@app.post("/admin/producao/zerar-sequencia")
# Liberado pra todos: zera só o contador impresso na etiqueta "Em
# Andamento". Não altera kit, estoque, produção nem histórico — quem está
# no chão de fábrica precisa reiniciar a numeração sem depender de admin.
@require_login
async def admin_producao_zerar_sequencia(request: Request):
    producao_mod.zerar_sequencia()
    return RedirectResponse("/admin/producao?ok=sequencia_zerada", status_code=302)


@app.post("/admin/producao/transito")
@require_permission("producao_mover_estagio")
async def admin_producao_transito(request: Request):
    form = await request.form()
    kit_ids = form.getlist("kit_ids")
    n = producao_mod.marcar_transito(kit_ids)
    request.state.auditoria_detalhe = f"{n} kit(s) movido(s) para Em trânsito"
    return RedirectResponse(f"/admin/producao?ok=transito&n={n}", status_code=302)


@app.post("/admin/producao/nota-lote")
# Mesma permissão da nota individual: o lote não pode ser um caminho mais
# frouxo pra fazer o que a rota unitária protege.
@require_permission("producao_nota_fiscal")
async def admin_producao_nota_lote(request: Request):
    """Mesma nota/data pra vários kits selecionados. Só os selecionados são
    tocados — o form manda os kit_ids marcados, nada de "todos do filtro"."""
    form = await request.form()
    kit_ids = [k for k in form.getlist("kit_ids") if str(k).strip()]
    if not kit_ids:
        return RedirectResponse("/admin/producao?erro=nota_lote_vazio", status_code=302)
    nota = str(form.get("nota_fiscal", ""))
    r = producao_mod.atribuir_nota_em_lote(
        kit_ids,
        nota,
        str(form.get("nota_fiscal_data", "")),
        str(form.get("motivo", "")),
    )
    q = f"?ok=nota_lote&n={len(r['atualizados'])}"
    if r["bloqueados"]:
        q += f"&bloqueados={len(r['bloqueados'])}"
    request.state.auditoria_detalhe = (
        f"{len(r['atualizados'])} kit(s): nota fiscal definida para {nota}")
    return RedirectResponse("/admin/producao" + q, status_code=302)


@app.post("/admin/producao/{kit_id}/cliente-instalando")
@require_permission("producao_mover_estagio")
async def admin_producao_cliente_instalando(request: Request, kit_id: str):
    veiculo = _veiculo_do_kit(kit_id)
    producao_mod.marcar_cliente_instalando(kit_id)
    request.state.auditoria_detalhe = f"Veículo {veiculo}: instalação iniciada no cliente"
    return RedirectResponse("/admin/producao?ok=instalando", status_code=302)


@app.post("/admin/producao/{kit_id}/cliente-concluido")
@require_permission("producao_mover_estagio")
async def admin_producao_cliente_concluido(request: Request, kit_id: str):
    veiculo = _veiculo_do_kit(kit_id)
    producao_mod.marcar_cliente_concluido(kit_id)
    request.state.auditoria_detalhe = f"Veículo {veiculo}: instalação concluída"
    return RedirectResponse("/admin/producao?ok=concluido", status_code=302)


@app.post("/admin/producao/cliente-concluido-lote")
@require_permission("producao_mover_estagio")
async def admin_producao_cliente_concluido_lote(request: Request):
    form = await request.form()
    kit_ids = form.getlist("kit_ids")
    n = producao_mod.marcar_cliente_concluido_lote(kit_ids)
    request.state.auditoria_detalhe = f"{n} kit(s): instalação concluída"
    return RedirectResponse(f"/admin/producao?ok=concluido_lote&n={n}", status_code=302)


@app.post("/admin/producao/{kit_id}/voltar")
@require_permission("producao_mover_estagio")
async def admin_producao_voltar(request: Request, kit_id: str):
    veiculo = _veiculo_do_kit(kit_id)
    producao_mod.voltar_estagio(kit_id)
    request.state.auditoria_detalhe = f"Veículo {veiculo}: voltou para a etapa anterior"
    return RedirectResponse("/admin/producao?ok=voltou", status_code=302)


@app.post("/admin/producao/{kit_id}/nota-fiscal")
@require_permission("producao_nota_fiscal")
async def admin_producao_nota_fiscal(request: Request, kit_id: str):
    form = await request.form()
    nota = str(form.get("nota_fiscal", ""))
    ok = producao_mod.atualizar_nota_fiscal(
        kit_id,
        nota,
        str(form.get("nota_fiscal_data", "")),
        str(form.get("motivo", "")),
    )
    if not ok:
        return RedirectResponse("/admin/producao?erro=nf_motivo", status_code=302)
    veiculo = _veiculo_do_kit(kit_id)
    request.state.auditoria_detalhe = f"Veículo {veiculo}: nota fiscal definida para {nota}"
    return RedirectResponse("/admin/producao?ok=nota_fiscal", status_code=302)


@app.get("/admin/producao/historico", response_class=HTMLResponse)
@require_login
async def admin_producao_historico(request: Request,
                                    data_ini: str = "", data_fim: str = "",
                                    pagina: int = 1):
    # Pagina em vez de cortar: o teto (LIMITE_HISTORICO) era o último corte
    # silencioso que sobrava — passando dele, o excedente sumia sem aviso.
    # Agora a lista inteira do período é paginada e a contagem diz o total.
    # A exportação continua trazendo tudo, com o mesmo filtro.
    todos = producao_mod.listar_historico(data_ini, data_fim,
                                          limite=producao_mod.LIMITE_HISTORICO)
    pag = paginacao_mod.paginar(todos, pagina)
    return render(request, "admin_producao_historico.html", {
        "registros": pag["itens"],
        "pag": pag,
        "atingiu_teto": len(todos) >= producao_mod.LIMITE_HISTORICO,
        "limite": producao_mod.LIMITE_HISTORICO,
        "data_ini": data_ini, "data_fim": data_fim,
    })


@app.get("/admin/producao/historico/exportar.xlsx")
@require_login
async def admin_producao_historico_exportar(request: Request,
                                             data_ini: str = "", data_fim: str = ""):
    from fastapi.responses import Response as _Resp
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from io import BytesIO

    registros = producao_mod.listar_historico(data_ini, data_fim,
                                          limite=producao_mod.LIMITE_HISTORICO)

    wb = openpyxl.Workbook()
    azul, branco = "1A3A5C", "FFFFFF"
    ws = wb.active
    ws.title = "Historico Producao"

    def hdr_cell(row, col, value):
        c = ws.cell(row=row, column=col, value=value)
        c.font = Font(bold=True, color=branco)
        c.fill = PatternFill("solid", fgColor=azul)
        c.alignment = Alignment(horizontal="center", vertical="center")
        return c

    for col, h in enumerate(
        ["Data/Hora", "Usuário", "Ação", "Kit", "Detalhe", "IP", "Status"], 1):
        hdr_cell(1, col, h)
    for i, r in enumerate(registros):
        row = i + 2
        ws.cell(row, 1, (r["criado_em"] or "")[:16])
        ws.cell(row, 2, r["user_nome"] or "—")
        ws.cell(row, 3, r["acao"])
        ws.cell(row, 4, r["kit_desc"])
        ws.cell(row, 5, r["resumo"])
        ws.cell(row, 6, r["ip"] or "")
        ws.cell(row, 7, r["status"])
    larguras = [17, 18, 26, 20, 60, 15, 8]
    for col, largura in enumerate(larguras, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = largura

    buf = BytesIO()
    wb.save(buf)
    return _Resp(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="historico_producao.xlsx"'},
    )


# ── Estoque ───────────────────────────────────────────────────────────────────

@app.get("/admin/estoque/ajuste-modelo.xlsx")
@require_permission("estoque_editar")
async def admin_estoque_ajuste_modelo(request: Request):
    """Planilha do estoque ATUAL, pronta pra editar e devolver. Sai com os
    valores de hoje preenchidos: o operador muda só as células que quiser e
    reenvia — o que não mudar não vira movimento.

    A coluna ID é a chave da volta. Fica visível (e travada por um aviso no
    cabeçalho) em vez de escondida, pra ninguém apagar a coluna sem saber o
    que está fazendo e depois a importação não achar o item."""
    from fastapi.responses import Response as _Resp
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.worksheet.datavalidation import DataValidation
    from io import BytesIO

    itens = estoque_mod.listar_estoque()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ajuste de Estoque"
    azul, branco, cinza = "1A3A5C", "FFFFFF", "F4F7FB"
    colunas = ["ID", "Tipo de Item", "Código de Barras",
               "Quantidade Atual", "Quantidade Mínima", "Status de Compra"]
    larguras = [8, 34, 22, 18, 18, 22]
    for col, (h, w) in enumerate(zip(colunas, larguras), 1):
        c = ws.cell(1, col, h)
        c.font = Font(bold=True, color=branco)
        c.fill = PatternFill("solid", fgColor=azul)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[c.column_letter].width = w
    for i, item in enumerate(itens):
        r = i + 2
        ws.cell(r, 1, item["id"])
        ws.cell(r, 2, item["tipo_nome"])
        ws.cell(r, 3, item["codigo_barra"])
        ws.cell(r, 4, item["quantidade_atual"])
        ws.cell(r, 5, item["quantidade_minima"])
        ws.cell(r, 6, item.get("status_compra") or "")
        if i % 2 == 0:
            for col in range(1, len(colunas) + 1):
                ws.cell(r, col).fill = PatternFill("solid", fgColor=cinza)
    # ID, tipo e código são só referência — quem edita mexe nas três últimas.
    for col in (1, 2, 3):
        for r in range(2, len(itens) + 2):
            ws.cell(r, col).font = Font(color="808080")
    ws.freeze_panes = "A2"

    if itens:
        dv = DataValidation(type="list",
                            formula1='"' + ",".join(estoque_mod.STATUS_COMPRA) + '"',
                            allow_blank=True, showErrorMessage=False)
        dv.promptTitle = "Status de Compra"
        dv.prompt = "Deixe vazio se não há pendência de compra."
        dv.showInputMessage = True
        ws.add_data_validation(dv)
        dv.add(f"F2:F{len(itens) + 1}")

    ws2 = wb.create_sheet("Como usar")
    instrucoes = [
        "Como atualizar o estoque em massa",
        "",
        "1. Altere as colunas Quantidade Atual, Quantidade Mínima e/ou Status de Compra.",
        "2. NÃO apague nem reordene a coluna ID — é por ela que o sistema reencontra cada item.",
        "3. Salve e envie o arquivo em Itens & Estoque › Importar Ajuste.",
        "",
        "Célula deixada em BRANCO significa 'não mexer nesse campo'.",
        "Para zerar um estoque, escreva 0 — apagar a célula não zera nada.",
        "",
        "Linha sem alteração nenhuma não gera movimento: reenviar a mesma",
        "planilha duas vezes não duplica nada no histórico.",
        "",
        "IMPORTANTE: baixe a planilha na hora de usar. Ela leva os números do",
        "momento do download, então enviar um arquivo baixado dias atrás faz o",
        "estoque VOLTAR para os valores daquele dia, desfazendo o que aconteceu",
        "no meio do caminho.",
        "",
        "Toda alteração entra no histórico do item com seu usuário, igual ao",
        "ajuste feito na tela — é possível auditar tudo depois.",
        "",
        "Status de Compra aceita: " + ", ".join(estoque_mod.STATUS_COMPRA) + " (ou vazio).",
    ]
    ws2.column_dimensions["A"].width = 92
    for i, linha in enumerate(instrucoes, 1):
        c = ws2.cell(i, 1, linha)
        if i == 1:
            c.font = Font(bold=True, size=13, color=branco)
            c.fill = PatternFill("solid", fgColor=azul)

    buf = BytesIO()
    wb.save(buf); buf.seek(0)
    return _Resp(content=buf.read(),
                 media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                 headers={"Content-Disposition":
                          'attachment; filename="ajuste_estoque.xlsx"'})


@app.post("/admin/estoque/importar-ajuste")
@require_permission("estoque_editar")
async def admin_estoque_importar_ajuste(request: Request):
    """Aplica a planilha de ajuste. Passa pelas mesmas funções da tela, então
    cada mudança gera o mesmo movimento e o mesmo log do ajuste manual."""
    user = get_current_user(request)
    form = await request.form()
    arquivo = form.get("arquivo")
    if not arquivo or not getattr(arquivo, "filename", ""):
        return RedirectResponse(
            "/admin/items?tab=catalogo&erro=" + quote("Selecione um arquivo .xlsx."),
            status_code=302)
    try:
        conteudo = await _ler_upload(request, arquivo)
        r = estoque_mod.importar_ajustes_xlsx(conteudo, user["id"])
    except Exception as e:
        return RedirectResponse(
            "/admin/items?tab=catalogo&erro=" + quote(_erro_usuario(e, "Não foi possível ler a planilha.")),
            status_code=302)

    partes = [f"{len(r['atualizados'])} item(ns) atualizado(s)"]
    if r["ignorados"]:
        partes.append(f"{r['ignorados']} sem alteração")
    if r["erros"]:
        partes.append(f"{len(r['erros'])} com problema: " + " | ".join(r["erros"][:3]))
    return RedirectResponse(
        "/admin/items?tab=catalogo&ok_ajuste=" + quote(" · ".join(partes)),
        status_code=302)


@app.get("/admin/estoque/exportar.xlsx")
@require_login
async def admin_estoque_exportar(request: Request):
    from fastapi.responses import Response as _Resp
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from io import BytesIO

    itens = estoque_mod.listar_estoque()
    historico = estoque_mod.listar_historico_completo()

    wb = openpyxl.Workbook()
    azul, branco, cinza = "1A3A5C", "FFFFFF", "F4F7FB"

    def hdr_cell(ws, row, col, value):
        c = ws.cell(row=row, column=col, value=value)
        c.font = Font(bold=True, color=branco)
        c.fill = PatternFill("solid", fgColor=azul)
        c.alignment = Alignment(horizontal="center", vertical="center")
        return c

    # ── Aba Estoque ──────────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Estoque"
    for col, h in enumerate(
        ["Tipo de Item", "Código de Barras", "Quantidade Atual",
         "Quantidade Mínima", "Status", "Cadastrado em"], 1):
        hdr_cell(ws1, 1, col, h)
    for i, item in enumerate(itens):
        row = i + 2
        abaixo = item["quantidade_atual"] <= item["quantidade_minima"]
        proximo = item["quantidade_atual"] <= item["quantidade_minima"] * 2 and not abaixo
        status = "Abaixo do mínimo" if abaixo else ("Próximo do mínimo" if proximo else "OK")
        ws1.cell(row, 1, item["tipo_nome"])
        ws1.cell(row, 2, item["codigo_barra"])
        ws1.cell(row, 3, item["quantidade_atual"])
        ws1.cell(row, 4, item["quantidade_minima"])
        ws1.cell(row, 5, status)
        ws1.cell(row, 6, item.get("criado_em") or "")
        if i % 2 == 0:
            for col in range(1, 7):
                ws1.cell(row, col).fill = PatternFill("solid", fgColor=cinza)
    for col, w in zip("ABCDEF", (32, 26, 16, 16, 20, 20)):
        ws1.column_dimensions[col].width = w
    ws1.freeze_panes = "A2"

    # ── Aba Histórico ────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Histórico")
    for col, h in enumerate(
        ["Tipo de Item", "Código de Barras", "Movimento", "Quantidade",
         "Observação", "Operador", "Data"], 1):
        hdr_cell(ws2, 1, col, h)
    movimento_labels = {
        "entrada": "Entrada", "saida": "Saída", "saida_cancelada": "Saída cancelada",
        "correcao": "Correção", "ajuste_minimo": "Ajuste mínimo",
        "status_compra": "Status de compra",
    }
    for i, m in enumerate(historico):
        row = i + 2
        ws2.cell(row, 1, m["tipo_nome"])
        ws2.cell(row, 2, m["codigo_barra"])
        ws2.cell(row, 3, movimento_labels.get(m["tipo"], m["tipo"]))
        ws2.cell(row, 4, m["quantidade"])
        ws2.cell(row, 5, m.get("observacao") or "")
        ws2.cell(row, 6, m.get("operador_nome") or "—")
        ws2.cell(row, 7, m.get("criado_em") or "")
        if i % 2 == 0:
            for col in range(1, 8):
                ws2.cell(row, col).fill = PatternFill("solid", fgColor=cinza)
    for col, w in zip("ABCDEFG", (32, 26, 18, 14, 32, 20, 20)):
        ws2.column_dimensions[col].width = w
    ws2.freeze_panes = "A2"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"estoque_{now_brt()[:10]}.xlsx"
    return _Resp(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.post("/admin/estoque/alertas")
@require_permission("estoque_editar")
async def admin_estoque_alertas(request: Request):
    """Salva como a faixa de aviso de estoque baixo se comporta. Mesma
    permissão de mexer no estoque: quem define o mínimo é quem define o
    que conta como falta."""
    form = await request.form()
    estoque_mod.salvar_alerta_config({
        # Checkbox desmarcado não é enviado — a ausência é o "desligado".
        "alerta_ativo": "1" if form.get("alerta_ativo") else "0",
        "alerta_margem": form.get("alerta_margem", ""),
        "alerta_limite": form.get("alerta_limite", ""),
        "alerta_segundos": form.get("alerta_segundos", ""),
        "alerta_telas": str(form.get("alerta_telas", "")),
        "alerta_cor_critico": str(form.get("alerta_cor_critico", "")),
    })
    return RedirectResponse("/admin/items?tab=catalogo&ok=alertas", status_code=302)


@app.get("/admin/estoque", response_class=HTMLResponse)
@require_login
async def admin_estoque(request: Request):
    import app.zpl as _zpl
    itens = estoque_mod.listar_estoque()
    alertas = estoque_mod.alertas_abaixo_minimo()
    url_http = getattr(app.state, "url_http", _zpl.SERVIDOR_URL)
    return render(request, "admin_estoque.html", {
        "itens": itens,
        "alertas": alertas,
        "url_http_base": url_http,
        "status_compra_opcoes": estoque_mod.STATUS_COMPRA,
    })


@app.post("/admin/estoque")
@require_login
async def admin_estoque_post(request: Request):
    # Chamado a partir do popup de configuração de um tipo em /admin/items
    # ("+ Adicionar ao estoque") — a criação de estoque para um tipo novo
    # acontece na aba "Novo Item", que usa /admin/tipos/completo.
    user = get_current_user(request)
    form = await request.form()
    try:
        item_tipo_id = int(form.get("item_tipo_id", 0) or 0)
        codigo_barra = form.get("codigo_barra", "").strip()
        quantidade = max(0, int(form.get("quantidade", 0) or 0))
        quantidade_minima = max(0, int(form.get("quantidade_minima", 0) or 0))
        if not item_tipo_id or not codigo_barra:
            raise ValueError("Tipo e código de barras são obrigatórios.")
        estoque_mod.criar_estoque(item_tipo_id, codigo_barra, quantidade,
                                   quantidade_minima, user["id"])
    except Exception as e:
        return RedirectResponse(
            "/admin/items?erro=" + quote(_erro_usuario(e, "Não foi possível adicionar ao estoque.")),
            status_code=302)
    return RedirectResponse("/admin/items?ok=estoque_criado", status_code=302)


@app.post("/admin/estoque/{estoque_id}/repor")
@require_permission("estoque_editar")
async def admin_estoque_repor(request: Request, estoque_id: int):
    user = get_current_user(request)
    form = await request.form()
    quantidade = max(1, int(form.get("quantidade", 1) or 1))
    observacao = form.get("observacao", "").strip()
    estoque_mod.repor_estoque(estoque_id, quantidade, user["id"], observacao)
    return RedirectResponse("/admin/items?ok=reposto", status_code=302)


@app.post("/admin/estoque/reconciliar-producao")
@require_admin
async def admin_estoque_reconciliar_producao(request: Request):
    user = get_current_user(request)
    resumo = estoque_mod.reconciliar_saidas_producao(user["id"])
    if not resumo:
        return RedirectResponse("/admin/items?tab=catalogo&ok=reconciliado_vazio", status_code=302)
    detalhe = "; ".join(f"{r['tipo_nome']} ×{r['quantidade']}" for r in resumo)
    return RedirectResponse(
        "/admin/items?tab=catalogo&ok=reconciliado&detalhe=" + quote(detalhe),
        status_code=302)


@app.post("/admin/sobressalente")
@require_permission("estoque_editar")
async def admin_sobressalente_enviar(request: Request):
    """Envia um OU VÁRIOS sobressalentes de uma vez. O formulário manda
    listas paralelas (estoque_id[], quantidade[], observacao[]), então uma
    linha só continua funcionando igual — é o mesmo caminho, com uma linha."""
    user = get_current_user(request)
    form = await request.form()
    cliente = str(form.get("cliente", "")).strip()
    destino = f"/admin/items?tab=sobressalentes&cliente={quote(cliente)}"

    ids = form.getlist("estoque_id")
    qtds = form.getlist("quantidade")
    obs = form.getlist("observacao")
    linhas = [{"estoque_id": ids[i],
               "quantidade": qtds[i] if i < len(qtds) else 1,
               "observacao": obs[i] if i < len(obs) else ""}
              for i in range(len(ids))]
    try:
        r = estoque_mod.registrar_sobressalentes_em_lote(linhas, cliente, user["id"])
    except ValueError as e:
        return RedirectResponse(
            destino + "&erro=" + quote(f"Erro ao registrar sobressalente: {e}"),
            status_code=302)
    return RedirectResponse(
        destino + "&ok=sobressalente&itens=" + str(r["itens"])
        + "&unidades=" + str(r["unidades"]),
        status_code=302)


@app.post("/admin/estoque/{estoque_id}/corrigir")
@require_permission("estoque_editar")
async def admin_estoque_corrigir(request: Request, estoque_id: int):
    user = get_current_user(request)
    form = await request.form()
    try:
        nova_quantidade = max(0, int(form.get("quantidade_atual", 0) or 0))
        estoque_mod.corrigir_quantidade(estoque_id, nova_quantidade, user["id"])
    except Exception as e:
        return RedirectResponse(
            "/admin/items?erro=" + quote(_erro_usuario(e, "Não foi possível corrigir a quantidade.")),
            status_code=302)
    return RedirectResponse("/admin/items?ok=quantidade_corrigida", status_code=302)


@app.post("/admin/estoque/{estoque_id}/minimo")
@require_permission("estoque_editar")
async def admin_estoque_minimo(request: Request, estoque_id: int):
    user = get_current_user(request)
    form = await request.form()
    novo_minimo = max(0, int(form.get("quantidade_minima", 0) or 0))
    estoque_mod.atualizar_minimo(estoque_id, novo_minimo, user["id"])
    return RedirectResponse("/admin/items?ok=minimo", status_code=302)


@app.post("/admin/estoque/{estoque_id}/status-compra")
@require_permission("estoque_editar")
async def admin_estoque_status_compra(request: Request, estoque_id: int):
    user = get_current_user(request)
    form = await request.form()
    estoque_mod.atualizar_status_compra(estoque_id, str(form.get("status_compra", "")), user["id"])
    return RedirectResponse("/admin/items?ok=status_compra", status_code=302)


@app.get("/admin/estoque/{estoque_id}/historico", response_class=HTMLResponse)
@require_login
async def admin_estoque_historico(request: Request, estoque_id: int, pagina: int = 1):
    est = estoque_mod.buscar_por_id(estoque_id)
    if not est:
        return RedirectResponse("/admin/estoque", status_code=302)
    # Paginado no SQL (LIMIT/OFFSET), não em memória: um item muito
    # movimentado pode ter milhares de linhas — carregar tudo pra depois
    # fatiar em Python seria o mesmo desperdício que o LIMIT fixo de antes,
    # só que sem a página 2 pra compensar.
    total = estoque_mod.contar_historico(estoque_id)
    por_pagina = paginacao_mod.POR_PAGINA_PADRAO
    total_paginas = max(1, -(-total // por_pagina))
    pagina = max(1, min(pagina, total_paginas))
    historico = estoque_mod.listar_historico(
        estoque_id, limit=por_pagina, offset=(pagina - 1) * por_pagina)
    pag_historico = {
        "itens": historico, "pagina": pagina, "total_paginas": total_paginas,
        "total": total,
        "inicio": (pagina - 1) * por_pagina + 1 if historico else 0,
        "fim": (pagina - 1) * por_pagina + len(historico),
        "exibindo": len(historico),
        "paginas_visiveis": paginacao_mod.janela_paginas(pagina, total_paginas),
    }
    return render(request, "admin_estoque_historico.html", {
        "est": est, "historico": historico, "pag_historico": pag_historico,
        # O histórico é aberto pela janela de configuração do item, que vive na
        # aba Estoque — voltar pra /admin/items pelado caía na aba errada.
        "voltar_para": _voltar_para(request, "/admin/items?tab=catalogo"),
    })


@app.post("/admin/estoque/{estoque_id}/delete")
@require_admin
async def admin_estoque_delete(request: Request, estoque_id: int):
    estoque_mod.deletar_estoque(estoque_id)
    return RedirectResponse("/admin/estoque?ok=excluido", status_code=302)


@app.get("/admin/estoque/{estoque_id}/etiqueta", response_class=HTMLResponse)
@require_login
async def admin_estoque_etiqueta(request: Request, estoque_id: int):
    import app.zpl as _zpl
    est = estoque_mod.buscar_por_id(estoque_id)
    if not est:
        raise HTTPException(status_code=404)
    base = getattr(app.state, "servidor_url", _zpl.SERVIDOR_URL)
    url_qr = f"{base}/estoque/{estoque_id}"
    html = _zpl.generate_estoque_html_label(
        tipo_nome=est["tipo_nome"],
        codigo_barra=est["codigo_barra"],
        url_qr=url_qr,
    )
    return HTMLResponse(content=html)


@app.get("/admin/estoque/{estoque_id}/qrcode.svg")
@require_login
async def admin_estoque_qrcode(request: Request, estoque_id: int):
    from fastapi.responses import Response as FResponse
    import app.zpl as _zpl
    est = estoque_mod.buscar_por_id(estoque_id)
    if not est:
        raise HTTPException(status_code=404)
    base = getattr(app.state, "servidor_url", _zpl.SERVIDOR_URL)
    url = f"{base}/estoque/{estoque_id}"
    import segno, io as _io
    qr = segno.make(url, error="q")
    buf = _io.BytesIO()
    qr.save(buf, kind="svg", scale=8, border=3, xmldecl=True, nl=False)
    return FResponse(content=buf.getvalue(), media_type="image/svg+xml")


# ── Veículos ──────────────────────────────────────────────────────────────────

def _admin_veiculos_context(cliente: list[str] | None = None, pagina: int = 1,
                            busca: str = "", modelo: list[str] | None = None,
                            situacao: list[str] | None = None,
                            garagem: list[str] | None = None,
                            imp_ini: str = "", imp_fim: str = "",
                            ordem: str = "") -> dict:
    # Os filtros são LISTAS: dá pra ver duas garagens (ou três situações) ao
    # mesmo tempo. Vazio = sem filtro, como antes.
    cliente = [c for c in (cliente or []) if c]
    modelo = [m for m in (modelo or []) if m]
    situacao = [s for s in (situacao or []) if s]
    garagem = [g for g in (garagem or []) if g]
    todos = veiculos_mod.listar(ativo=True)
    total_geral = len(todos)
    # Localização de todos numa consulta só — usada tanto pra exibir a coluna
    # quanto pra filtrar por ela, então o filtro nunca discorda do que a
    # coluna mostra.
    localizacao = producao_mod.localizacao_dos_veiculos()
    for v in todos:
        loc = localizacao.get(v["id"])
        v["localizacao"] = loc["texto"] if loc else ""
        v["localizacao_estado"] = loc["estado"] if loc else ""

    # Veículo cujo kit está devendo item — a pendência que não pode sumir de
    # vista até alguém repor. Um mapa só pra todos (nada de N+1).
    incompletos = sessions_mod.veiculos_com_kit_incompleto()
    for v in todos:
        # Por id e, na falta dele, pelo número: kit antigo (anterior à coluna
        # veiculo_id) só tem o texto do veículo, e sem isso a marca não
        # aparecia justamente nos kits mais velhos.
        v["falta_item"] = (incompletos.get(v["id"])
                           or incompletos.get((v["numero"] or "").strip().upper()))

    veiculos = todos
    # Dentro do mesmo filtro as opções somam (duas garagens = as duas listas
    # juntas); entre filtros diferentes elas se cruzam (essas garagens E esse
    # modelo). É o que se espera de "filtrar por duas coisas".
    if cliente:
        alvo_cli = {c.strip() for c in cliente}
        veiculos = [v for v in veiculos if v["cliente"] in alvo_cli]
    # Garagem: comparação pelo nome INTEIRO (a busca livre casa por pedaço e
    # "GARAGEM 1" trazia junto a "GARAGEM 10"). __sem__ = os que estão sem
    # garagem — o furo de cadastro que tira o veículo de "Kits a produzir".
    if garagem:
        alvo_gar = {g.strip().upper() for g in garagem}
        sem_garagem = "__SEM__" in alvo_gar
        veiculos = [v for v in veiculos
                    if (sem_garagem and not (v["garagem"] or "").strip())
                    or (v["garagem"] or "").strip().upper() in alvo_gar]
    if modelo:
        alvo_mod = {m.strip().lower() for m in modelo}
        veiculos = [v for v in veiculos
                    if (v["modelo"] or "").strip().lower() in alvo_mod]
    # Situação cobre tanto a etapa do fluxo (localização) quanto os furos de
    # cadastro — as duas coisas que fazem o operador filtrar essa lista.
    if situacao:
        estados = set()
        for s in situacao:
            if s == "cliente":
                estados |= {"cliente_instalando", "cliente_concluido"}
            elif s in ("a_produzir", "em_producao", "produzido", "transito"):
                estados.add(s)

        def _cabe(v):
            if v["localizacao_estado"] in estados:
                return True
            if "sem_localizacao" in situacao and not v["localizacao_estado"]:
                return True
            if "sem_modelo" in situacao and not (v["modelo"] or "").strip():
                return True
            if "sem_garagem" in situacao and not (v["garagem"] or "").strip():
                return True
            if "sem_kits" in situacao and not v["total_kits"]:
                return True
            if "com_kits" in situacao and v["total_kits"]:
                return True
            if "falta_item" in situacao and v.get("falta_item"):
                return True
            return False

        veiculos = [v for v in veiculos if _cabe(v)]

    if busca:
        por_texto = paginacao_mod.filtrar(
            veiculos, busca, ("numero", "cliente", "garagem", "modelo", "localizacao"))
        # Busca por PATRIMÔNIO: acha o veículo pelo código de um item bipado
        # em qualquer kit dele. Mínimo de 4 caracteres — menos que isso o
        # LIKE varre demais e devolve veículo demais pra servir de busca.
        ids_texto = {v["id"] for v in por_texto}
        extras = []
        if len(busca.strip()) >= 4:
            with db() as conn:
                ids_patrimonio = {r["veiculo_id"] for r in conn.execute(
                    "SELECT DISTINCT kr.veiculo_id FROM scan_session_items ssi "
                    "JOIN kit_record kr ON kr.sessao_id = ssi.sessao_id "
                    "WHERE ssi.codigo_barra LIKE ? AND kr.veiculo_id IS NOT NULL "
                    "LIMIT 100",
                    (f"%{busca.strip()}%",)).fetchall()}
            extras = [v for v in veiculos
                      if v["id"] in ids_patrimonio and v["id"] not in ids_texto]
        veiculos = por_texto + extras

    # Data de IMPORTAÇÃO: quando o veículo entrou no sistema. Vem de
    # veiculos.criado_em — o campo já existia e já estava preenchido em todos,
    # então não há campo novo nem migração. Comparação pelos 10 primeiros
    # caracteres ("2026-08-25"), que é o formato gravado: assim "de 25 até 25"
    # pega o dia inteiro sem precisar somar hora.
    if imp_ini:
        veiculos = [v for v in veiculos if (v.get("criado_em") or "")[:10] >= imp_ini]
    if imp_fim:
        veiculos = [v for v in veiculos if (v.get("criado_em") or "")[:10] <= imp_fim]

    # Ordenação: o padrão continua cliente+número (é como o operador procura).
    # "importado_desc" responde "o que entrou por último?", que é a pergunta
    # de quem acabou de importar uma planilha.
    if ordem == "importado_desc":
        veiculos = sorted(veiculos, key=lambda v: (v.get("criado_em") or ""), reverse=True)
    elif ordem == "importado_asc":
        veiculos = sorted(veiculos, key=lambda v: (v.get("criado_em") or ""))

    # A lista de inativos segue um cliente só (é uma lista de apoio); com
    # vários filtrados, mostra todos e deixa o filtro pra lista principal.
    veiculos_inativos = veiculos_mod.listar(
        cliente=cliente[0] if len(cliente) == 1 else None, ativo=False)
    cadastrados_clientes = clientes_mod.listar()
    cadastradas_garagens = garagens_mod.listar()

    # Cliente ⇄ garagem pra as listas suspensas do topo: abrir um cliente
    # mostra as garagens dele (e vice-versa). O vínculo não existe em
    # cadastro nenhum — quem liga os dois é o veículo —, então sai daqui.
    _id_garagem = {g["nome"].upper(): g["id"] for g in cadastradas_garagens}
    garagens_por_cliente: dict[str, list] = {}
    clientes_por_garagem: dict[str, list] = {}
    for par in veiculos_mod.mapa_cliente_garagem():
        nome_g = par["garagem"]
        garagens_por_cliente.setdefault(par["cliente"], []).append({
            "nome": nome_g or "(sem garagem)",
            "id": _id_garagem.get(nome_g.upper()) if nome_g else None,
            "veiculos": par["veiculos"],
        })
        if nome_g:
            clientes_por_garagem.setdefault(nome_g.upper(), []).append({
                "nome": par["cliente"],
                "id": next((c["id"] for c in cadastrados_clientes
                            if c["nome"] == par["cliente"]), None),
                "veiculos": par["veiculos"],
            })
    return {
        # Onde cada veículo está agora no fluxo — derivado dos estados que a
        # Produção já usa, numa consulta só pra todos (nada de N+1).
        "localizacao": localizacao,
        "busca": busca,
        "filtro_modelo": modelo,
        "filtro_situacao": situacao,
        "filtro_garagem": garagem,
        # Levados de volta pra tela: sem isso o formulário voltava em branco
        # e o operador não sabia por que a lista estava curta.
        "imp_ini": imp_ini,
        "imp_fim": imp_fim,
        "ordem": ordem,
        "total_geral": total_geral,
        "tem_filtro": bool(cliente or garagem or modelo or situacao or busca
                              or imp_ini or imp_fim or ordem),
        "modelos": veiculos_mod.modelos_disponiveis(),
        "sem_modelo": veiculos_mod.contar_sem_modelo(
            cliente[0] if len(cliente) == 1 else None),
        "pag_veiculos": paginacao_mod.paginar(veiculos, pagina),
        "veiculos_inativos": veiculos_inativos,
        # Números cadastrados mais de uma vez (dados de antes da regra de
        # número único) — a tela avisa pra o admin limpar com a exclusão
        # em massa; o sistema não apaga sozinho porque cada cadastro pode
        # ter kits no histórico.
        "duplicados": veiculos_mod.numeros_duplicados(),
        "clientes": [c["nome"] for c in clientes_mod.listar()],
        "clientes_cadastrados": cadastrados_clientes,
        "garagens_cadastradas": cadastradas_garagens,
        # Nome → id, pra cliente e garagem na tabela virarem link pro
        # panorama de cada um sem uma consulta por linha.
        "id_cliente": {c["nome"]: c["id"] for c in cadastrados_clientes},
        "id_garagem": {g["nome"]: g["id"] for g in cadastradas_garagens},
        "garagens_por_cliente": garagens_por_cliente,
        "clientes_por_garagem": clientes_por_garagem,
        "filtro_cliente": cliente,
    }


@app.get("/admin/veiculos", response_class=HTMLResponse)
@require_login
async def admin_veiculos(request: Request, pagina: int = 1, busca: str = "",
                         cliente: list[str] = Query(default=[]),
                         modelo: list[str] = Query(default=[]),
                         situacao: list[str] = Query(default=[]),
                         garagem: list[str] = Query(default=[]),
                         imp_ini: str = "", imp_fim: str = "", ordem: str = ""):
    return render(request, "admin_veiculos.html",
                  _admin_veiculos_context(cliente, pagina, busca, modelo, situacao,
                                          garagem, imp_ini, imp_fim, ordem))


@app.post("/admin/veiculos", response_class=HTMLResponse)
@require_login
async def admin_veiculos_post(request: Request):
    form = await request.form()
    numero = str(form.get("numero", "")).strip()
    cliente = str(form.get("cliente", "")).strip()
    garagem = str(form.get("garagem", "")).strip()
    modelo = str(form.get("modelo", "")).strip()
    if not numero or not cliente:
        return render(request, "admin_veiculos.html", {
            **_admin_veiculos_context(),
            "erro": "Número e cliente são obrigatórios.",
        })
    try:
        veiculos_mod.criar(numero, cliente, garagem, modelo)
    except ValueError as e:
        # Número já em uso — o número do veículo é único no sistema inteiro.
        return render(request, "admin_veiculos.html", {
            **_admin_veiculos_context(),
            "erro": str(e),
        })
    return RedirectResponse("/admin/veiculos?ok=criado", status_code=302)


@app.post("/admin/veiculos/excluir-em-massa")
@require_permission("veiculos_excluir")
async def admin_veiculos_excluir_em_massa(request: Request):
    """Exclui vários veículos de uma vez — a tela deixa marcar alguns ou
    todos os veículos filtrados (ex: todos de um cliente digitado errado).
    Cada exclusão preserva o histórico de kits, igual à exclusão unitária:
    é o mesmo deletar() por baixo. Só admin, como toda exclusão."""
    form = await request.form()
    ids = [int(x) for x in form.getlist("veiculo_ids") if str(x).isdigit()]
    for vid in ids:
        veiculos_mod.deletar(vid)
    voltar = str(form.get("voltar", "")).strip()
    destino = "/admin/veiculos" + (f"?cliente={quote(voltar)}&" if voltar else "?")
    return RedirectResponse(destino + f"ok=excluidos&qtd={len(ids)}", status_code=302)


@app.get("/admin/veiculos/modelo.xlsx")
@require_login
async def admin_veiculos_modelo(request: Request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from io import BytesIO
    from fastapi.responses import Response as _Resp
    azul, branco = "1A3A5C", "FFFFFF"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Veículos"
    for col, h in enumerate(["Número do Veículo", "Cliente", "Garagem", "Modelo (Kit)"], 1):
        c = ws.cell(1, col, h)
        c.font = Font(bold=True, color=branco)
        c.fill = PatternFill("solid", fgColor=azul)
        ws.column_dimensions[ws.cell(1, col).column_letter].width = 30
    # Exemplos usam nomes de kit REAIS quando existirem — o modelo tem que
    # bater com o nome do kit, então mostrar o nome certo evita erro de
    # digitação na planilha.
    modelos = veiculos_mod.modelos_disponiveis()
    ex1 = modelos[0] if modelos else "Mercedes Euro 5 Diesel"
    ex2 = modelos[1] if len(modelos) > 1 else ex1
    ws.cell(2, 1, "VH-001"); ws.cell(2, 2, "Exemplo Cliente"); ws.cell(2, 3, "Base Norte"); ws.cell(2, 4, ex1)
    ws.cell(3, 1, "VH-002"); ws.cell(3, 2, "Outro Cliente");  ws.cell(3, 3, "Base Sul");   ws.cell(3, 4, ex2)

    # Aba de apoio: a lista exata dos modelos aceitos, pra copiar e colar.
    ws2 = wb.create_sheet("Modelos disponíveis")
    c = ws2.cell(1, 1, "Modelos válidos (nomes dos kits ativos)")
    c.font = Font(bold=True, color=branco)
    c.fill = PatternFill("solid", fgColor=azul)
    ws2.column_dimensions["A"].width = 44
    for i, m in enumerate(modelos, 2):
        ws2.cell(i, 1, m)
    if not modelos:
        ws2.cell(2, 1, "(nenhum kit cadastrado ainda)")

    # Lista suspensa na coluna Modelo, alimentada pela aba acima: o operador
    # escolhe o kit em vez de digitar, que é onde nascia o erro (o modelo é
    # comparado pelo nome). showErrorMessage=False de propósito — digitar à
    # mão continua valendo, porque a importação casa o texto com o kit MAIS
    # PRÓXIMO (caixa/abreviação/erro de digitação); só o que não parece com
    # kit nenhum é recusado e reportado na tela de importação.
    if modelos:
        from openpyxl.worksheet.datavalidation import DataValidation
        dv = DataValidation(
            type="list",
            formula1=f"='Modelos disponíveis'!$A$2:$A${len(modelos) + 1}",
            allow_blank=True,
            showErrorMessage=False,
        )
        dv.prompt = ("Escolha um kit da lista. Se digitar, o sistema casa com "
                     "o kit mais próximo — mas só entre os que já existem.")
        dv.promptTitle = "Modelo (Kit)"
        dv.showInputMessage = True
        ws.add_data_validation(dv)
        dv.add(f"D2:D{max(len(modelos), 500)}")

    buf = BytesIO()
    wb.save(buf); buf.seek(0)
    return _Resp(content=buf.read(),
                 media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                 headers={"Content-Disposition": "attachment; filename=modelo_veiculos.xlsx"})


@app.get("/admin/veiculos/import", response_class=HTMLResponse)
@require_login
async def admin_veiculos_import_form(request: Request):
    return render(request, "admin_veiculos_import.html", {
        "voltar_para": _voltar_para(request, "/admin/veiculos"),
    })


@app.post("/admin/veiculos/import", response_class=HTMLResponse)
@require_login
async def admin_veiculos_import_post(request: Request):
    form = await request.form()
    arquivo = form.get("arquivo")
    if not arquivo or not arquivo.filename:
        return render(request, "admin_veiculos_import.html",
                      {"erro": "Selecione um arquivo .xlsx."})
    try:
        file_bytes = await _ler_upload(request, arquivo)
    except ValueError as e:
        return render(request, "admin_veiculos_import.html", {"erro": str(e)})
    user = request.state.user
    resultado = veiculos_mod.importar_excel(file_bytes)
    # Grava a conferência e MANDA PRA ELA. Antes o resultado morria na resposta
    # do POST: um F5 e ninguém mais sabia o que a planilha tinha feito.
    imp_id = importacoes_mod.registrar(
        resultado, getattr(arquivo, "filename", "") or "", user["id"])
    return RedirectResponse(f"/admin/veiculos/import/{imp_id}", status_code=302)


def _conferencia_context(request: Request, importacao_id: int) -> dict | None:
    imp = importacoes_mod.uma(importacao_id)
    if not imp:
        return None
    situacao = (request.query_params.get("situacao") or "").strip()
    busca = (request.query_params.get("busca") or "").strip()
    so_avisos = request.query_params.get("avisos") == "1"
    return {
        "imp": imp,
        "contagens": importacoes_mod.contagens(importacao_id),
        "linhas": importacoes_mod.itens(importacao_id, situacao, busca, so_avisos),
        "situacao": situacao,
        "busca": busca,
        "so_avisos": so_avisos,
        "situacoes": importacoes_mod.SITUACOES,
        "voltar_para": _voltar_para(request, "/admin/veiculos"),
    }


@app.get("/admin/veiculos/importacoes", response_class=HTMLResponse)
@require_login
async def admin_veiculos_importacoes(request: Request):
    return render(request, "admin_veiculos_importacoes.html", {
        "importacoes": importacoes_mod.listar(),
        "voltar_para": _voltar_para(request, "/admin/veiculos"),
    })


@app.get("/admin/veiculos/import/{importacao_id:int}", response_class=HTMLResponse)
@require_login
async def admin_veiculos_conferencia(request: Request, importacao_id: int):
    ctx = _conferencia_context(request, importacao_id)
    if ctx is None:
        return RedirectResponse("/admin/veiculos/importacoes?erro=nao_encontrada",
                                status_code=302)
    return render(request, "admin_veiculos_conferencia.html", ctx)


def _painel_context(request: Request, veiculo_id: int | None = None,
                    codigo_barra: str = "", voltar: str = "",
                    modo: str = "veiculo") -> dict:
    """Contexto do PAINEL ÚNICO — o mesmo pedaço de tela aberto pelo número do
    veículo e pelo código do patrimônio.

    Antes eram dois popups com metade das ações cada um: pelo veículo dava pra
    adicionar item, pelo patrimônio dava pra mover/corrigir. Quem entrava pela
    porta errada tinha que sair e entrar de novo. Agora é um só: o veículo
    sempre aparece, e o item fica "em foco" quando se entra por ele.

    `codigo_barra` manda: se veio um patrimônio, o veículo é o do kit dele.

    `modo` diz PARA QUE a janela foi aberta, e é só a quantidade de tela:
    "veiculo" (clique no número, em Veículos e Clientes) é a janela inteira —
    itens do kit, pendência e adicionar item; "item" (clique no código, em
    Itens Cadastrados) é a janela curta de quem só quer resolver aquele
    patrimônio: veículo, item em foco, mover, corrigir e o histórico."""
    item_sel = items_mod.onde_esta(codigo_barra) if codigo_barra else None
    v = None
    if item_sel and item_sel.get("veiculo"):
        v = veiculos_mod.buscar_por_numero(item_sel["veiculo"])
    if v is None and veiculo_id:
        v = veiculos_mod.buscar(veiculo_id)
    kit = items_mod.kit_do_veiculo(v["numero"]) if v else None
    # Entrando pelo item, o kit em foco é o DELE — não o mais recente do
    # veículo, que pode ser outro se o veículo já teve mais de um kit.
    if item_sel and item_sel.get("kit_id"):
        kit = items_mod.kit_por_id(item_sel["kit_id"]) or kit
    return {
        "v": v,
        "kit": kit,
        "modo": modo,
        # A MESMA localização da lista de veículos — a janela não pode dizer
        # um estágio e a lista atrás dela dizer outro.
        "localizacao": (producao_mod.localizacao_dos_veiculos().get(v["id"])
                        if v else None),
        "codigo_barra": codigo_barra,
        "item_sel": item_sel,
        "conferencia": sessions_mod.conferencia_com_modelo(kit["kit_id"]) if kit else [],
        # A pendência sai da MESMA fonte das outras telas, pra janela e lista
        # nunca discordarem sobre o que está faltando.
        "faltas": (sessions_mod.kits_incompletos().get(kit["kit_id"], {}).get("faltas", [])
                   if kit else []),
        "mudancas": sessions_mod.historico_mudancas_do_kit(kit["kit_id"]) if kit else [],
        "itens": sessions_mod.itens_do_kit(kit["kit_id"]) if kit else [],
        "tipos_do_kit": (items_mod.listar_tipos_para_kit(kit["kit_template_id"])
                         if kit else []),
        "voltar": voltar,
        "previa_mover": None, "mover_destino": "", "mover_motivo": "",
        "previa_add": None, "add_codigo": "", "add_serial": "", "add_motivo": "",
    }


def _previa_adicionar(ctx: dict, codigo: str, motivo: str) -> dict:
    """O que vai acontecer ao adicionar este código no kit em foco — sem
    gravar nada. Dizer em QUAL veículo o código está é o ponto: sem isso o
    operador só descobria o conflito depois de tentar."""
    previa: dict = {"codigo": codigo, "ocupado": None, "erro": None}
    if not ctx["kit"]:
        previa["erro"] = "Este veículo não tem kit montado."
    elif len(motivo) < 5:
        previa["erro"] = ("Informe o motivo (pelo menos 5 letras) — ele fica gravado "
                          "no histórico do item.")
    elif not codigo:
        previa["erro"] = "Informe o código do patrimônio."
    else:
        ocupado = items_mod.onde_esta(codigo)
        if ocupado and ocupado.get("kit_id") == ctx["kit"]["kit_id"]:
            previa["erro"] = "Este patrimônio já está neste kit."
        else:
            previa["ocupado"] = ocupado
    return previa


def _preencher_previa(ctx: dict, form, codigo_em_foco: str) -> dict:
    """Preenche a prévia do formulário que foi enviado.

    A janela é uma só e tem dois formulários — mover (manda `destino`) e
    adicionar (manda `codigo_barra`). Como as duas portas abrem a mesma
    janela, as duas rotas precisam entender os dois; separar por rota era o
    que fazia "adicionar item" só funcionar entrando pelo veículo."""
    motivo = str(form.get("motivo", "")).strip()
    if form.get("destino") is not None and codigo_em_foco:
        destino = str(form.get("destino", "")).strip()
        ctx.update({"previa_mover": items_mod.previa_mover(codigo_em_foco, destino),
                    "mover_destino": destino, "mover_motivo": motivo})
    else:
        codigo = str(form.get("codigo_barra", "")).strip()
        ctx.update({"previa_add": _previa_adicionar(ctx, codigo, motivo),
                    "add_codigo": codigo,
                    "add_serial": str(form.get("serial", "")).strip(),
                    "add_motivo": motivo})
    return ctx


@app.get("/admin/veiculos/{veiculo_id}/painel", response_class=HTMLResponse)
@require_login
async def admin_veiculo_painel(request: Request, veiculo_id: int, voltar: str = "",
                               item: str = ""):
    """Janela do veículo — abre no clique do número, sem sair da lista."""
    return render(request, "_painel.html",
                  _painel_context(request, veiculo_id, item, voltar))


@app.post("/admin/veiculos/{veiculo_id}/painel", response_class=HTMLResponse)
@require_login
async def admin_veiculo_painel_previa(request: Request, veiculo_id: int):
    """Prévia dentro da janela, entrando pelo número do veículo — os mesmos
    dois formulários da porta do patrimônio."""
    form = await request.form()
    item = str(form.get("item", ""))
    ctx = _painel_context(request, veiculo_id, item, str(form.get("voltar", "")))
    return render(request, "_painel.html", _preencher_previa(ctx, form, item))


@app.get("/admin/veiculos/{veiculo_id}", response_class=HTMLResponse)
@require_login
async def admin_veiculo_detalhe(request: Request, veiculo_id: int):
    v = veiculos_mod.buscar(veiculo_id)
    if not v:
        raise HTTPException(status_code=404)
    historico = veiculos_mod.historico_kits(veiculo_id)
    clientes_cadastrados = clientes_mod.listar()
    garagens_cadastradas = garagens_mod.listar()
    # Pendência de patrimônio deste veículo — a mesma marca da lista, aqui
    # também: quem abre o veículo pra entender o que houve tem que ver.
    _pend = sessions_mod.veiculos_com_kit_incompleto()
    return render(request, "admin_veiculo_detalhe.html", {
        "voltar_para": _voltar_para(request, "/admin/veiculos"),
        "falta_item": (_pend.get(veiculo_id)
                       or _pend.get((v["numero"] or "").strip().upper())),
        "v": v, "historico": historico, "clientes": clientes_cadastrados,
        "garagens": garagens_cadastradas,
        "ocupado": veiculos_mod.esta_ocupado(veiculo_id),
        "kits_para_vincular": veiculos_mod.kits_para_vincular(veiculo_id),
        "modelos": veiculos_mod.modelos_disponiveis(),
        # Toda vez que uma planilha tocou neste veículo. É o que responde
        # "quem mudou a garagem dele?" sem abrir importação por importação.
        "importacoes": importacoes_mod.do_veiculo(veiculo_id),
        "ok": request.query_params.get("ok", ""),
    })


@app.post("/admin/veiculos/{veiculo_id}/editar")
@require_login
async def admin_veiculo_editar(request: Request, veiculo_id: int):
    form = await request.form()
    numero = str(form.get("numero", "")).strip()
    cliente = str(form.get("cliente", "")).strip()
    garagem = str(form.get("garagem", "")).strip()
    # O modelo tem formulário próprio (fica junto do vínculo de kit). Se o
    # campo não vier, modelo=None manda atualizar() preservar o que está
    # gravado — salvar os dados do veículo não pode zerar o modelo.
    modelo_bruto = form.get("modelo")
    modelo = str(modelo_bruto).strip() if modelo_bruto is not None else None
    if not numero or not cliente:
        v = veiculos_mod.buscar(veiculo_id)
        clientes = clientes_mod.listar()
        garagens_cadastradas = garagens_mod.listar()
        return render(request, "admin_veiculo_detalhe.html", {
            "voltar_para": "/admin/veiculos",
            "v": v, "historico": veiculos_mod.historico_kits(veiculo_id),
            "clientes": clientes, "garagens": garagens_cadastradas,
            "ocupado": veiculos_mod.esta_ocupado(veiculo_id),
            "modelos": veiculos_mod.modelos_disponiveis(),
            "kits_para_vincular": veiculos_mod.kits_para_vincular(veiculo_id),
            "erro": "Número e cliente são obrigatórios.",
        })
    try:
        veiculos_mod.atualizar(veiculo_id, numero, cliente, garagem, modelo)
    except ValueError as e:
        # Número já em uso por outro veículo — único no sistema inteiro.
        v = veiculos_mod.buscar(veiculo_id)
        return render(request, "admin_veiculo_detalhe.html", {
            "voltar_para": "/admin/veiculos",
            "v": v, "historico": veiculos_mod.historico_kits(veiculo_id),
            "clientes": clientes_mod.listar(), "garagens": garagens_mod.listar(),
            "ocupado": veiculos_mod.esta_ocupado(veiculo_id),
            "modelos": veiculos_mod.modelos_disponiveis(),
            "kits_para_vincular": veiculos_mod.kits_para_vincular(veiculo_id),
            "erro": str(e),
        })
    return RedirectResponse(f"/admin/veiculos/{veiculo_id}?ok=atualizado", status_code=302)


@app.post("/admin/veiculos/{veiculo_id}/modelo")
@require_login
async def admin_veiculo_modelo(request: Request, veiculo_id: int):
    """Só o modelo. Fica separado do formulário de dados porque na tela ele
    vive junto do vínculo de kit — é ele que decide em qual bipagem este
    veículo aparece, então o operador precisa ver as duas coisas juntas."""
    form = await request.form()
    veiculos_mod.definir_modelo(veiculo_id, str(form.get("modelo", "")).strip())
    return RedirectResponse(f"/admin/veiculos/{veiculo_id}?ok=modelo", status_code=302)


@app.post("/admin/veiculos/{veiculo_id}/liberar")
@require_admin
async def admin_veiculo_liberar(request: Request, veiculo_id: int):
    """Destrava o veículo pra uma nova bipagem mesmo já tendo kit(s)
    associados. Não apaga nem altera nada do kit/sessão existente — só
    permite atribuir esse veículo de novo. É consumido automaticamente
    assim que essa nova atribuição acontecer."""
    veiculos_mod.liberar(veiculo_id)
    return RedirectResponse(f"/admin/veiculos/{veiculo_id}?ok=liberado", status_code=302)


@app.post("/admin/veiculos/{veiculo_id}/desativar")
@require_login
async def admin_veiculo_desativar(request: Request, veiculo_id: int):
    veiculos_mod.desativar(veiculo_id)
    return RedirectResponse("/admin/veiculos?ok=desativado", status_code=302)


@app.post("/admin/veiculos/{veiculo_id}/reativar")
@require_login
async def admin_veiculo_reativar(request: Request, veiculo_id: int):
    veiculos_mod.reativar(veiculo_id)
    return RedirectResponse(f"/admin/veiculos/{veiculo_id}?ok=reativado", status_code=302)


@app.post("/admin/veiculos/{veiculo_id}/delete")
@require_permission("veiculos_excluir")
async def admin_veiculo_delete(request: Request, veiculo_id: int):
    veiculos_mod.deletar(veiculo_id)
    return RedirectResponse("/admin/veiculos?ok=excluido", status_code=302)


@app.post("/admin/clientes")
@require_login
async def admin_clientes_post(request: Request):
    form = await request.form()
    nome = str(form.get("nome", "")).strip()
    if not nome:
        return RedirectResponse("/admin/veiculos?erro_cliente=vazio", status_code=302)
    resultado = clientes_mod.criar(nome)
    if resultado is None:
        return RedirectResponse("/admin/veiculos?erro_cliente=duplicado", status_code=302)
    return RedirectResponse("/admin/veiculos?ok=cliente", status_code=302)


@app.post("/admin/clientes/{cliente_id}/delete")
@require_admin
async def admin_cliente_delete(request: Request, cliente_id: int):
    clientes_mod.deletar(cliente_id)
    return RedirectResponse("/admin/veiculos?ok=cliente_excluido", status_code=302)


@app.post("/admin/garagens")
@require_login
async def admin_garagens_post(request: Request):
    form = await request.form()
    nome = str(form.get("nome", "")).strip()
    if not nome:
        return RedirectResponse("/admin/veiculos?erro_garagem=vazio", status_code=302)
    resultado = garagens_mod.criar(nome)
    if resultado is None:
        return RedirectResponse("/admin/veiculos?erro_garagem=duplicado", status_code=302)
    return RedirectResponse("/admin/veiculos?ok=garagem", status_code=302)


@app.post("/admin/garagens/{garagem_id}/delete")
@require_admin
async def admin_garagem_delete(request: Request, garagem_id: int):
    garagens_mod.deletar(garagem_id)
    return RedirectResponse("/admin/veiculos?ok=garagem_excluida", status_code=302)




# ── Cliente e Garagem — páginas de detalhe ────────────────────────────────────
# Cliente e garagem não têm vínculo próprio no banco: quem liga os dois é o
# VEÍCULO. Estas telas juntam num lugar só tudo que hoje estava espalhado
# (veículos, garagens usadas, modelos, kits e a esteira) e dão a única ação
# que de fato amarra os dois — mover veículos de uma garagem pra outra.

@app.get("/admin/clientes/{cliente_id}", response_class=HTMLResponse)
@require_login
async def admin_cliente_detalhe(request: Request, cliente_id: int, ok: str = "",
                                erro: str = ""):
    cliente = clientes_mod.buscar(cliente_id)
    if not cliente:
        return RedirectResponse("/admin/veiculos?erro_cliente=nao_encontrado",
                                status_code=302)
    dados = clientes_mod.panorama(cliente["nome"])
    return render(request, "admin_cliente.html", {
        **dados,
        "garagens_cadastradas": garagens_mod.listar(),
        "voltar_para": _voltar_para(request, "/admin/veiculos"),
        "ok": ok, "erro": erro,
    })


@app.post("/admin/clientes/{cliente_id}/prefixo")
@require_login
async def admin_cliente_prefixo(request: Request, cliente_id: int):
    """Prefixo do número de frota (ex: REDEMOB = 31001) — usado pra montar
    o número sozinho (PREFIXO-00001) quando cadastro manual ou planilha
    mandam só o número cru."""
    form = await request.form()
    clientes_mod.atualizar_prefixo(cliente_id, str(form.get("prefixo", "")))
    return RedirectResponse(f"/admin/clientes/{cliente_id}?ok=prefixo", status_code=302)


@app.get("/admin/garagens/{garagem_id}", response_class=HTMLResponse)
@require_login
async def admin_garagem_detalhe(request: Request, garagem_id: int, ok: str = "",
                                erro: str = ""):
    garagem = garagens_mod.buscar(garagem_id)
    if not garagem:
        return RedirectResponse("/admin/veiculos?erro_garagem=nao_encontrada",
                                status_code=302)
    dados = garagens_mod.panorama(garagem["nome"])
    return render(request, "admin_garagem.html", {
        **dados,
        "garagens_cadastradas": garagens_mod.listar(),
        "voltar_para": _voltar_para(request, "/admin/veiculos"),
        "ok": ok, "erro": erro,
    })


@app.post("/admin/clientes/{cliente_id}/mover-garagem")
@require_login
async def admin_cliente_mover_garagem(request: Request, cliente_id: int):
    """Passa os veículos deste cliente de uma garagem para outra."""
    cliente = clientes_mod.buscar(cliente_id)
    if not cliente:
        return RedirectResponse("/admin/veiculos?erro_cliente=nao_encontrado",
                                status_code=302)
    destino = await _garagem_destino_valida(request)
    if destino is None:
        return RedirectResponse(f"/admin/clientes/{cliente_id}?erro=destino",
                                status_code=302)
    form = await request.form()
    origem = str(form.get("origem", "")).strip()
    movidos = veiculos_mod.mover_veiculos_de_garagem(origem, destino, cliente["nome"])
    return RedirectResponse(f"/admin/clientes/{cliente_id}?ok=movidos_{movidos}",
                            status_code=302)


@app.post("/admin/garagens/{garagem_id}/mover-garagem")
@require_login
async def admin_garagem_mover(request: Request, garagem_id: int):
    """Tira veículos desta garagem e joga em outra — de um cliente só ou de
    todos, conforme o que foi escolhido na tela."""
    garagem = garagens_mod.buscar(garagem_id)
    if not garagem:
        return RedirectResponse("/admin/veiculos?erro_garagem=nao_encontrada",
                                status_code=302)
    destino = await _garagem_destino_valida(request)
    if destino is None:
        return RedirectResponse(f"/admin/garagens/{garagem_id}?erro=destino",
                                status_code=302)
    form = await request.form()
    cliente = str(form.get("cliente", "")).strip()
    movidos = veiculos_mod.mover_veiculos_de_garagem(
        garagem["nome"], destino, cliente or None)
    return RedirectResponse(f"/admin/garagens/{garagem_id}?ok=movidos_{movidos}",
                            status_code=302)


async def _garagem_destino_valida(request: Request) -> str | None:
    """O destino tem que ser uma garagem CADASTRADA — senão a tela viraria
    uma porta de entrada pra garagem digitada errada, que é exatamente o que
    o cadastro existe pra evitar. Devolve None quando não serve."""
    form = await request.form()
    destino = str(form.get("destino", "")).strip()
    if not destino:
        return None
    nomes = {g["nome"].upper() for g in garagens_mod.listar()}
    return destino if destino.upper() in nomes else None


# ── Estoque — página mobile (acesso via QR code) ──────────────────────────────

@app.get("/estoque", response_class=HTMLResponse)
@require_login
async def estoque_lista_mobile(request: Request):
    """Lista de estoque somente leitura, otimizada para celular — sem
    formulários de ajuste. Edição continua em /admin/estoque (computador)
    e em /estoque/{id} (via QR da etiqueta, no local)."""
    itens = estoque_mod.listar_estoque()
    return render(request, "estoque_lista_mobile.html", {"itens": itens})


@app.get("/estoque/buscar")
async def estoque_buscar(request: Request, codigo: str = ""):
    """Resolve um código de barras ou o texto de um QR de estoque para a
    página de consulta correspondente — usado pelo scanner do /mobile."""
    est = estoque_mod.buscar_por_referencia(codigo)
    if not est:
        return RedirectResponse("/mobile?erro=estoque_nao_encontrado", status_code=302)
    return RedirectResponse(f"/estoque/{est['id']}", status_code=302)


@app.get("/estoque/{estoque_id}", response_class=HTMLResponse)
async def estoque_mobile(request: Request, estoque_id: int):
    # Consulta de quantidade é pública (como a verificação de kit) —
    # só o ajuste de estoque exige login.
    est = estoque_mod.buscar_por_id(estoque_id)
    if not est:
        return RedirectResponse("/mobile?erro=estoque_nao_encontrado", status_code=302)
    historico = estoque_mod.listar_historico(estoque_id, limit=8)
    return render(request, "estoque_mobile.html", {
        "est": est,
        "historico": historico,
        "ok": request.query_params.get("ok"),
    })


@app.post("/estoque/{estoque_id}/ajustar")
# Mesma permissão do ajuste pela tela de admin: é a mesma operação de
# estoque, e o celular não pode ser a porta dos fundos.
@require_permission("estoque_editar")
async def estoque_mobile_ajustar(request: Request, estoque_id: int):
    user = get_current_user(request)
    form = await request.form()
    tipo = (form.get("tipo") or "").strip()
    motivo = (form.get("motivo") or "").strip()
    try:
        quantidade = max(1, int(form.get("quantidade") or 1))
    except (ValueError, TypeError):
        quantidade = 1

    def _erro(msg):
        est = estoque_mod.buscar_por_id(estoque_id)
        historico = estoque_mod.listar_historico(estoque_id, limit=8)
        return render(request, "estoque_mobile.html", {
            "est": est, "historico": historico,
            "erro": msg, "tipo_sel": tipo, "qtd_sel": quantidade,
        })

    if tipo not in ("entrada", "saida"):
        return _erro("Selecione Adicionar ou Subtrair.")
    if not motivo:
        return _erro("Motivo é obrigatório.")

    try:
        estoque_mod.ajustar_quantidade(estoque_id, tipo, quantidade, motivo, user["id"])
    except ValueError as e:
        return _erro(str(e))

    return RedirectResponse(f"/estoque/{estoque_id}?ok=1", status_code=302)


# ── Reset do banco (apenas admin) ─────────────────────────────────────────────

@app.get("/admin/reset", response_class=HTMLResponse)
@require_admin
async def reset_page(request: Request):
    return render(request, "admin_reset.html")


@app.post("/admin/reset")
@require_admin
async def reset_confirm(request: Request, confirmacao: str = Form("")):
    if confirmacao != "CONFIRMAR":
        return render(request, "admin_reset.html", {"erro": "Digite CONFIRMAR para prosseguir."})
    with db() as conn:
        conn.execute("DELETE FROM print_queue")
        conn.execute("DELETE FROM scan_session_items")
        conn.execute("DELETE FROM scan_session")
        conn.execute("DELETE FROM kit_record")
        conn.execute("DELETE FROM item_master")
        conn.execute("DELETE FROM kit_template_items")
        conn.execute("DELETE FROM kit_template")
        conn.execute("DELETE FROM item_tipo")
        conn.execute("DELETE FROM users")
        # Reseta os autoincrement
        conn.execute("DELETE FROM sqlite_sequence WHERE name != 'sqlite_sequence'")
    # Limpa a sessão (o próprio usuário foi apagado)
    request.session.clear()
    return RedirectResponse("/login?ok=reset", status_code=302)


if __name__ == "__main__":
    import asyncio
    import uvicorn

    _tem_ssl = os.path.exists("certs/cert.pem") and os.path.exists("certs/key.pem")

    # SOMENTE_HTTPS=1 desliga a porta 8080 (HTTP puro) quando o certificado
    # existe, deixando só a 8011 (HTTPS) no ar. Fica atrás de uma flag —
    # desligado por padrão — porque sem 8080 quem ainda depende de HTTP na
    # LAN (ou não tem o certificado instalado/confiável no aparelho) perde
    # acesso.
    _somente_https = os.getenv("SOMENTE_HTTPS", "").strip() in ("1", "true", "True")

    if _tem_ssl and _somente_https:
        uvicorn.run(
            "main:app", host="0.0.0.0", port=8011, reload=False,
            ssl_certfile="certs/cert.pem", ssl_keyfile="certs/key.pem",
        )
    elif _tem_ssl:
        async def _serve_dual():
            cfg_https = uvicorn.Config(
                "main:app", host="0.0.0.0", port=8011, reload=False,
                ssl_certfile="certs/cert.pem", ssl_keyfile="certs/key.pem",
            )
            cfg_http = uvicorn.Config(
                "main:app", host="0.0.0.0", port=8080, reload=False,
            )
            await asyncio.gather(
                uvicorn.Server(cfg_https).serve(),
                uvicorn.Server(cfg_http).serve(),
            )
        asyncio.run(_serve_dual())
    else:
        uvicorn.run("main:app", host="0.0.0.0", port=8080, reload=False)
